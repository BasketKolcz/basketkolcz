from flask import Flask, render_template_string, request, redirect, url_for, flash, session, send_file, g
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re, io, json, os
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "basketkolcz2025secret")
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = False
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

DATABASE_URL = os.environ.get("DATABASE_URL", "")

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════

def get_db():
    if "db" not in g:
        g.db = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
        g.db.autocommit = False
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db:
        try: db.rollback()
        except: pass
        try: db.close()
        except: pass

def init_db():
    db = get_db()
    cur = db.cursor()
    # Tabele główne
    cur.execute("""
    CREATE TABLE IF NOT EXISTS matches (
        id SERIAL PRIMARY KEY,
        sezon VARCHAR(20) NOT NULL DEFAULT '2024/25',
        data_meczu DATE,
        przeciwnik VARCHAR(100) NOT NULL,
        nazwa_gtk VARCHAR(100) DEFAULT '',
        rozgrywki VARCHAR(100) DEFAULT '',
        runda VARCHAR(50) DEFAULT '',
        miejsce VARCHAR(20) DEFAULT '',
        wynik_gtk INTEGER DEFAULT 0,
        wynik_opp INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS match_stats (
        id SERIAL PRIMARY KEY,
        match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
        druzyna VARCHAR(10) NOT NULL,
        kwarta INTEGER NOT NULL,
        pts INTEGER DEFAULT 0, poss INTEGER DEFAULT 0,
        p2m INTEGER DEFAULT 0, p2a INTEGER DEFAULT 0,
        p3m INTEGER DEFAULT 0, p3a INTEGER DEFAULT 0,
        ftm INTEGER DEFAULT 0, fta INTEGER DEFAULT 0,
        br INTEGER DEFAULT 0, fd INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS player_stats (
        id SERIAL PRIMARY KEY,
        match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
        druzyna VARCHAR(10) NOT NULL,
        nr INTEGER NOT NULL,
        pts INTEGER DEFAULT 0, p2m INTEGER DEFAULT 0, p2a INTEGER DEFAULT 0,
        p3m INTEGER DEFAULT 0, p3a INTEGER DEFAULT 0,
        ftm INTEGER DEFAULT 0, fta INTEGER DEFAULT 0,
        ast INTEGER DEFAULT 0, oreb INTEGER DEFAULT 0, dreb INTEGER DEFAULT 0,
        br INTEGER DEFAULT 0, fd INTEGER DEFAULT 0, finishes INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS timing_stats (
        id SERIAL PRIMARY KEY,
        match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
        druzyna VARCHAR(10) NOT NULL, bucket VARCHAR(10) NOT NULL,
        made2 INTEGER DEFAULT 0, att2 INTEGER DEFAULT 0,
        made3 INTEGER DEFAULT 0, att3 INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS roster (
        id SERIAL PRIMARY KEY,
        imie VARCHAR(50) NOT NULL,
        nazwisko VARCHAR(50) NOT NULL DEFAULT '',
        pseudonim VARCHAR(30) DEFAULT '',
        aktywny BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS player_aliases (
        id SERIAL PRIMARY KEY,
        roster_id INTEGER REFERENCES roster(id) ON DELETE CASCADE,
        nr INTEGER NOT NULL,
        sezon VARCHAR(20) DEFAULT '',
        UNIQUE(roster_id, nr, sezon)
    );
    CREATE TABLE IF NOT EXISTS settings (
        key VARCHAR(50) PRIMARY KEY,
        value VARCHAR(200)
    );
    CREATE TABLE IF NOT EXISTS lineup_stats (
        id SERIAL PRIMARY KEY,
        match_id INTEGER REFERENCES matches(id) ON DELETE CASCADE,
        druzyna VARCHAR(10) NOT NULL,
        lineup VARCHAR(30) NOT NULL,
        pts INTEGER DEFAULT 0,
        poss INTEGER DEFAULT 0,
        p2m INTEGER DEFAULT 0, p2a INTEGER DEFAULT 0,
        p3m INTEGER DEFAULT 0, p3a INTEGER DEFAULT 0,
        ftm INTEGER DEFAULT 0, fta INTEGER DEFAULT 0,
        br INTEGER DEFAULT 0, fd INTEGER DEFAULT 0
    );
    INSERT INTO settings (key, value) VALUES ('gtk_name', 'GTK') ON CONFLICT DO NOTHING;
    INSERT INTO settings (key, value) VALUES ('current_season', '2024/25') ON CONFLICT DO NOTHING;
    """)
    db.commit()

    # ALTER TABLE — każdy osobno z savepoint
    alters = [
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS nazwa_gtk VARCHAR(100) DEFAULT ''",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS rozgrywki VARCHAR(100) DEFAULT ''",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS runda VARCHAR(50) DEFAULT ''",
        "ALTER TABLE matches ADD COLUMN IF NOT EXISTS miejsce VARCHAR(20) DEFAULT ''",
        "ALTER TABLE player_stats ADD COLUMN IF NOT EXISTS roster_id INTEGER REFERENCES roster(id) ON DELETE SET NULL",
    ]
    for sql in alters:
        try:
            cur.execute("SAVEPOINT sp")
            cur.execute(sql)
            cur.execute("RELEASE SAVEPOINT sp")
            db.commit()
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT sp")
            db.commit()
    cur.close()


def get_setting(key):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT value FROM settings WHERE key=%s", (key,))
    row = cur.fetchone()
    cur.close()
    return row['value'] if row else None

def set_setting(key, value):
    db = get_db()
    cur = db.cursor()
    cur.execute("INSERT INTO settings (key,value) VALUES (%s,%s) ON CONFLICT (key) DO UPDATE SET value=%s",
                (key, value, value))
    db.commit()
    cur.close()

# ══════════════════════════════════════════════════════════════════════════════
# PARSER (identyczny jak w app_v2.py)
# ══════════════════════════════════════════════════════════════════════════════

ACTION_2PM = {"2","2+1","2+0","2D","2D+1","2D+0/1W","2D+1/1W"}
ACTION_3PM = {"3","3+1","3+0"}
ACTION_BR  = {"BR"}
ACTION_F   = {"F"}
BUCKETS    = ["0s","1-4s","5-8s","9-12s","13-16s","17-20s","21-24s"]

def extract_ft(code):
    # Format X/YW — np. 1/2W, 2/2W, 0/3W
    m = re.match(r'^(\d+)/(\d+)W', code)
    if m: return int(m.group(1)), int(m.group(2))
    m2 = re.search(r'(\d+)/(\d+)W', code)
    if m2: return int(m2.group(1)), int(m2.group(2))
    # Plus-one: 2+1, 3+1, 2D+1 → celny 1/1 RW
    if re.search(r'\+1$', code): return 1, 1
    # Plus-zero: 2+0, 3+0, 2D+0 → niecelny 0/1 RW
    if re.search(r'\+0$', code): return 0, 1
    return 0, 0

def time_bucket(t):
    if t == 0:    return "0s"
    if t <= 4:    return "1-4s"
    if t <= 8:    return "5-8s"
    if t <= 12:   return "9-12s"
    if t <= 16:   return "13-16s"
    if t <= 20:   return "17-20s"
    return "21-24s"

def parse_team_sheet(ws):
    stats = {
        "quarter": defaultdict(lambda: {"ftm":0,"fta":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"br":0,"fd":0,"poss":0,"pts":0}),
        "players": defaultdict(lambda: {"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"fd":0,"br":0,"finishes":0,"ast":0,"oreb":0,"dreb":0}),
        "timing":  {b: {"2PT":{"made":0,"miss":0},"3PT":{"made":0,"miss":0}} for b in BUCKETS},
        "lineups": defaultdict(lambda: {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"fd":0}),
    }
    current_q = 1
    current_lineup = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row[:4]): continue
        if row[0] is not None:
            try: current_q = int(str(row[0]).replace("*","").strip())
            except: current_q = 1

        for i in range(4, 9):
            val = row[i] if len(row) > i and row[i] is not None else None
            if val is not None:
                s = str(val).strip()
                if ";" in s:
                    parts = s.split(";")
                    try:
                        old_p, new_p = int(parts[0].strip()), int(parts[1].strip())
                        try: current_lineup[current_lineup.index(old_p)] = new_p
                        except ValueError: current_lineup.append(new_p)
                    except: pass
                else:
                    try:
                        p = int(s)
                        if p not in current_lineup: current_lineup.append(p)
                    except: pass
        if len(current_lineup) > 5: current_lineup = current_lineup[-5:]

        raw_b = str(row[1]) if row[1] is not None else ""
        raw_c = str(row[2]) if row[2] is not None else ""
        raw_d = str(row[3]) if row[3] is not None else ""
        raw_k = str(row[10]) if len(row)>10 and row[10] is not None else ""
        raw_l = str(row[11]) if len(row)>11 and row[11] is not None else ""
        raw_m = str(row[12]) if len(row)>12 and row[12] is not None else ""
        raw_n = str(row[13]) if len(row)>13 and row[13] is not None else ""

        codes     = [c.strip().upper() for c in raw_c.split(";") if c.strip()]
        times     = [t.strip() for t in raw_b.replace(";",",").split(",") if t.strip()]
        finishers = [f.strip() for f in raw_k.split(";") if f.strip()]
        assists   = [a.strip() for a in raw_l.split(";") if a.strip()]
        orebs     = [o.strip() for o in raw_m.split(";") if o.strip()]
        drebs     = [d.strip() for d in raw_n.split(";") if d.strip()]

        q = stats["quarter"][current_q]
        q["poss"] += 1

        for ai, code in enumerate(codes):
            t_val = 0
            if ai < len(times):
                try: t_val = float(times[ai])
                except: pass
            bucket = time_bucket(t_val)

            finisher = None
            if ai < len(finishers):
                try: finisher = int(finishers[ai])
                except: pass

            assister = None
            if ai < len(assists):
                try: assister = int(assists[ai])
                except: pass

            orebler = None
            if ai < len(orebs):
                try: orebler = int(orebs[ai])
                except: pass

            drebler = None
            if ai < len(drebs):
                try: drebler = int(drebs[ai])
                except: pass

            pts = 0

            if code in ACTION_2PM:
                q["p2m"]+=1; q["p2a"]+=1; pts=2
                stats["timing"][bucket]["2PT"]["made"]+=1
                if finisher is not None:
                    stats["players"][finisher]["p2m"]+=1; stats["players"][finisher]["p2a"]+=1
                    stats["players"][finisher]["finishes"]+=1
                    if assister is not None: stats["players"][assister]["ast"]+=1
            elif code in ("0/2","0/2D","0D+0/2W","0D+1/2W"):
                q["p2a"]+=1
                stats["timing"][bucket]["2PT"]["miss"]+=1
                if finisher is not None: stats["players"][finisher]["p2a"]+=1; stats["players"][finisher]["finishes"]+=1
                if orebler is not None: stats["players"][orebler]["oreb"]+=1
            elif code in ACTION_3PM:
                q["p3m"]+=1; q["p3a"]+=1; pts=3
                stats["timing"][bucket]["3PT"]["made"]+=1
                if finisher is not None:
                    stats["players"][finisher]["p3m"]+=1; stats["players"][finisher]["p3a"]+=1
                    stats["players"][finisher]["finishes"]+=1
                    if assister is not None: stats["players"][assister]["ast"]+=1
            elif code == "0/3":
                q["p3a"]+=1
                stats["timing"][bucket]["3PT"]["miss"]+=1
                if finisher is not None: stats["players"][finisher]["p3a"]+=1; stats["players"][finisher]["finishes"]+=1
                if orebler is not None: stats["players"][orebler]["oreb"]+=1
            elif code in ACTION_BR:
                q["br"]+=1
                if finisher is not None: stats["players"][finisher]["br"]+=1
                if drebler is not None: stats["players"][drebler]["dreb"]+=1
            elif code in ACTION_F:
                q["fd"]+=1

            ftm, fta = extract_ft(code)
            if fta > 0:
                q["ftm"]+=ftm; q["fta"]+=fta; pts+=ftm
                if finisher is not None:
                    stats["players"][finisher]["ftm"]+=ftm; stats["players"][finisher]["fta"]+=fta
                    stats["players"][finisher]["fd"]+=1

            q["pts"] += pts

            # Lineup tracking
            if len(current_lineup) == 5:
                lk = "-".join(str(x) for x in sorted(current_lineup))
                lu = stats["lineups"][lk]
                lu["poss"] += 1
                lu["pts"]  += pts
                if code in ACTION_2PM:
                    lu["p2m"]+=1; lu["p2a"]+=1
                elif code in ("0/2","0/2D","0D+0/2W","0D+1/2W"):
                    lu["p2a"]+=1
                elif code in ACTION_3PM:
                    lu["p3m"]+=1; lu["p3a"]+=1
                elif code == "0/3":
                    lu["p3a"]+=1
                elif code in ACTION_BR:
                    lu["br"]+=1
                elif code in ACTION_F:
                    lu["fd"]+=1
                if fta > 0:
                    lu["ftm"]+=ftm; lu["fta"]+=fta

    return stats

def suma_quarters(stats):
    s = defaultdict(int)
    for qn in [1,2,3,4]:
        for k,v in stats["quarter"].get(qn,{}).items():
            s[k] += v
    return dict(s)

def calc_kpi(d):
    fga = d.get("p2a",0) + d.get("p3a",0)
    pts = d.get("pts",0)
    poss = max(d.get("poss",1),1)
    fta = d.get("fta",0); ftm = d.get("ftm",0)
    pm2 = d.get("p2m",0); pa2 = d.get("p2a",0)
    pm3 = d.get("p3m",0); pa3 = d.get("p3a",0)
    def pct(n,d): return f"{n/d:.1%}" if d else "-"
    efg  = (pm2+1.5*pm3)/fga if fga else None
    ts   = pts/(2*(fga+0.44*fta)) if (fga+fta) else None
    return {
        "efg":   pct(pm2+1.5*pm3,fga) if fga else "-",
        "ts":    f"{ts:.1%}" if ts else "-",
        "ortg":  f"{pts*100/poss:.1f}",
        "ppp":   f"{pts/poss:.2f}",
        "topct": f"{d.get('br',0)/poss:.1%}",
        "ftr":   f"{fta/fga:.2f}" if fga else "-",
        "p2_pct":pct(pm2,pa2),
        "p3_pct":pct(pm3,pa3),
        "ft_pct":pct(ftm,fta),
    }

def save_match_to_db(przeciwnik, nazwa_gtk, sezon, data_meczu, stats_gtk, stats_opp,
                     rozgrywki="", runda="", miejsce=""):
    db = get_db()
    cur = db.cursor()
    suma_gtk = suma_quarters(stats_gtk)
    suma_opp = suma_quarters(stats_opp)

    # Wstaw mecz
    cur.execute("""
        INSERT INTO matches (sezon, data_meczu, przeciwnik, nazwa_gtk, rozgrywki, runda, miejsce, wynik_gtk, wynik_opp)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
    """, (sezon, data_meczu, przeciwnik, nazwa_gtk,
          rozgrywki or "", runda or "", miejsce or "",
          suma_gtk.get("pts",0), suma_opp.get("pts",0)))
    match_id = cur.fetchone()["id"]

    # Statystyki per kwarta
    for qn in [1,2,3,4]:
        qg = stats_gtk["quarter"].get(qn,{})
        qo = stats_opp["quarter"].get(qn,{})
        for druzyna, qd in [("gtk", qg), ("opp", qo)]:
            cur.execute("""
                INSERT INTO match_stats (match_id,druzyna,kwarta,pts,poss,p2m,p2a,p3m,p3a,ftm,fta,br,fd)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (match_id, druzyna, qn,
                  qd.get("pts",0), qd.get("poss",0),
                  qd.get("p2m",0), qd.get("p2a",0),
                  qd.get("p3m",0), qd.get("p3a",0),
                  qd.get("ftm",0), qd.get("fta",0),
                  qd.get("br",0),  qd.get("fd",0)))

    # Zawodnicy
    for druzyna, stats in [("gtk", stats_gtk), ("opp", stats_opp)]:
        for nr, pd in stats["players"].items():
            pts = pd.get("p2m",0)*2 + pd.get("p3m",0)*3 + pd.get("ftm",0)
            cur.execute("""
                INSERT INTO player_stats (match_id,druzyna,nr,pts,p2m,p2a,p3m,p3a,ftm,fta,ast,oreb,dreb,br,fd,finishes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (match_id, druzyna, int(nr), pts,
                  pd.get("p2m",0), pd.get("p2a",0),
                  pd.get("p3m",0), pd.get("p3a",0),
                  pd.get("ftm",0), pd.get("fta",0),
                  pd.get("ast",0), pd.get("oreb",0),
                  pd.get("dreb",0), pd.get("br",0),
                  pd.get("fd",0), pd.get("finishes",0)))

    # Shot timing
    for druzyna, stats in [("gtk", stats_gtk), ("opp", stats_opp)]:
        for b in BUCKETS:
            td = stats["timing"][b]
            cur.execute("""
                INSERT INTO timing_stats (match_id,druzyna,bucket,made2,att2,made3,att3)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (match_id, druzyna, b,
                  td["2PT"]["made"], td["2PT"]["made"]+td["2PT"]["miss"],
                  td["3PT"]["made"], td["3PT"]["made"]+td["3PT"]["miss"]))

    # Piątki (lineup stats) — tylko GTK
    for lineup_key, ld in stats_gtk["lineups"].items():
        cur.execute("""
            INSERT INTO lineup_stats (match_id,druzyna,lineup,pts,poss,p2m,p2a,p3m,p3a,ftm,fta,br,fd)
            VALUES (%s,'gtk',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (match_id, lineup_key,
              ld["pts"], ld["poss"],
              ld["p2m"], ld["p2a"],
              ld["p3m"], ld["p3a"],
              ld["ftm"], ld["fta"],
              ld["br"],  ld["fd"]))

    db.commit()
    cur.close()
    return match_id

# ══════════════════════════════════════════════════════════════════════════════
# HTML COMPONENTS
# ══════════════════════════════════════════════════════════════════════════════

CSS = """
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
/* ── ZMIENNE ── */
:root{
  --navy:#1a2b4a; --navy2:#2e5090;
  --gtk:#1a6b3c;  --gtk-light:#e8f5e9;
  --opp:#8b1a1a;  --opp-light:#ffebee;
  --gold:#EF9F27;
  --bg:#f0f2f7;
  --card:#fff;
  --radius:12px;
  --shadow:0 1px 6px rgba(0,0,0,.08);
}

/* ── RESET ── */
*{box-sizing:border-box}
body{margin:0;background:var(--bg);font-family:'Segoe UI',system-ui,sans-serif;font-size:.9rem;color:#222}

/* ══════════════════════════════════════════
   SIDEBAR (desktop)
══════════════════════════════════════════ */
.sidebar{
  position:fixed;top:0;left:0;height:100vh;width:240px;
  background:var(--navy);z-index:1000;
  display:flex;flex-direction:column;
  overflow-y:auto;overflow-x:hidden;
  transition:transform .3s ease;
}
.sidebar-logo{
  padding:1.1rem 1rem .9rem;
  border-bottom:1px solid #ffffff18;
  flex-shrink:0;
}
.sidebar-logo .brand{font-size:1.05rem;font-weight:700;color:#fff;letter-spacing:.3px}
.sidebar-logo .brand span{color:var(--gold)}
.sidebar-logo .sub{font-size:.68rem;color:#ffffff55;margin-top:2px}
.nav-season{
  margin:.6rem .6rem .3rem;
  padding:.5rem .7rem;
  background:#ffffff0d;
  border-radius:8px;
  font-size:.73rem;color:#ffffff77;
}
.nav-season strong{color:var(--gold);display:block;font-size:.8rem;margin-bottom:1px}
.nav-section{
  padding:.5rem 1rem .2rem;
  font-size:.6rem;text-transform:uppercase;
  letter-spacing:1.2px;color:#ffffff33;font-weight:700;
}
.nav-item-link{
  display:flex;align-items:center;gap:.6rem;
  padding:.58rem .9rem;
  color:#ffffffaa;text-decoration:none;
  font-size:.84rem;border-radius:8px;
  margin:1px .5rem;transition:.15s;
  white-space:nowrap;
}
.nav-item-link:hover{background:#ffffff14;color:#fff}
.nav-item-link.active{background:#EF9F2720;color:var(--gold);font-weight:600}
.nav-item-link .icon{width:20px;text-align:center;font-size:.95rem;flex-shrink:0}

/* ══════════════════════════════════════════
   TOPBAR (mobile)
══════════════════════════════════════════ */
.topbar{
  display:none;
  position:fixed;top:0;left:0;right:0;height:56px;
  background:var(--navy);z-index:1001;
  align-items:center;padding:0 1rem;gap:.75rem;
}
.topbar .t-brand{font-size:1rem;font-weight:700;color:#fff}
.topbar .t-brand span{color:var(--gold)}
.hamburger{
  background:none;border:none;cursor:pointer;
  padding:6px;border-radius:6px;
  display:flex;flex-direction:column;gap:5px;
}
.hamburger span{display:block;width:22px;height:2px;background:#fff;border-radius:2px;transition:.25s}
.hamburger.open span:nth-child(1){transform:translateY(7px) rotate(45deg)}
.hamburger.open span:nth-child(2){opacity:0}
.hamburger.open span:nth-child(3){transform:translateY(-7px) rotate(-45deg)}
.sidebar-overlay{
  display:none;position:fixed;inset:0;background:#00000055;z-index:999;
}

/* ══════════════════════════════════════════
   MAIN CONTENT
══════════════════════════════════════════ */
.main-content{
  margin-left:240px;
  min-height:100vh;
  padding:1.5rem 1.25rem;
  max-width:1400px;
}

/* ══════════════════════════════════════════
   CARDS & STATS
══════════════════════════════════════════ */
.card{
  border:none;border-radius:var(--radius);
  box-shadow:var(--shadow);background:var(--card);
}
.stat-card{
  background:var(--card);border-radius:var(--radius);
  padding:.85rem .75rem;text-align:center;
  box-shadow:var(--shadow);
}
.stat-val{font-size:1.5rem;font-weight:700;color:var(--navy);line-height:1.1}
.stat-val.sm{font-size:1.05rem}
.stat-lbl{font-size:.65rem;color:#999;text-transform:uppercase;letter-spacing:.5px;margin-top:.2rem}

/* ══════════════════════════════════════════
   TABLES
══════════════════════════════════════════ */
.table th{
  background:var(--navy);color:#fff;
  font-size:.75rem;font-weight:600;
  border:none;padding:.42rem .55rem;
  white-space:nowrap;
}
.table td{
  font-size:.8rem;vertical-align:middle;
  padding:.36rem .55rem;
}
.table-hover tbody tr:hover{background:#f0f4ff}
.table-responsive{-webkit-overflow-scrolling:touch}

/* ══════════════════════════════════════════
   MISC UI
══════════════════════════════════════════ */
.hero{
  background:linear-gradient(135deg,var(--navy),var(--navy2));
  color:#fff;border-radius:14px;padding:1.25rem 1.5rem;
}
.page-title{font-size:1.2rem;font-weight:700;color:var(--navy);margin-bottom:.75rem}
.badge-win{background:#c8e6c9;color:var(--gtk);font-size:.72rem;padding:3px 10px;border-radius:20px;font-weight:700}
.badge-loss{background:#ffcdd2;color:var(--opp);font-size:.72rem;padding:3px 10px;border-radius:20px;font-weight:700}
.badge-draw{background:#e0e0e0;color:#555;font-size:.72rem;padding:3px 10px;border-radius:20px;font-weight:700}
.gtk-color{color:var(--gtk);font-weight:700}
.opp-color{color:var(--opp);font-weight:700}
.upload-zone{
  border:2px dashed #c5cfe8;border-radius:14px;
  padding:2rem 1.5rem;text-align:center;
  background:#fff;cursor:pointer;transition:.2s;
}
.upload-zone:hover{background:#f0f4ff;border-color:var(--navy)}
.nav-tabs .nav-link{color:#666;font-size:.82rem;padding:.4rem .8rem}
.nav-tabs .nav-link.active{color:var(--navy);font-weight:600;border-bottom:2px solid var(--navy)}
.section-hdr{
  font-size:.66rem;text-transform:uppercase;letter-spacing:1px;
  color:#aaa;font-weight:700;margin:.6rem 0 .35rem;
  padding-bottom:.25rem;border-bottom:1px solid #f0f0f0;
}
.flash-msg{padding:.6rem 1rem;border-radius:8px;margin-bottom:.75rem;font-size:.85rem}
.flash-success{background:#e8f5e9;color:#1a6b3c;border:1px solid #a5d6a7}
.flash-error{background:#ffebee;color:#8b1a1a;border:1px solid #ef9a9a}

/* ══════════════════════════════════════════
   RESPONSIVE BREAKPOINTS
══════════════════════════════════════════ */

/* ══════════════════════════════════════════
   RESPONSIVE BREAKPOINTS
══════════════════════════════════════════ */

/* Tablet / pół ekranu (< 992px) — sidebar tylko ikony */
@media(max-width:991px){
  .sidebar{width:56px}
  .sidebar-logo{padding:.75rem 0;text-align:center}
  .sidebar .brand-text,.sidebar .sub,
  .nav-section,.nav-season{display:none}
  .nav-item-link span.brand-text{display:none}
  .nav-item-link{
    justify-content:center;
    padding:.65rem 0;
    margin:2px 6px;
    border-radius:8px;
    position:relative;
  }
  .nav-item-link .icon{
    width:28px;height:28px;
    display:flex;align-items:center;justify-content:center;
    font-size:1.2rem;
    margin:0;
  }
  /* Tooltip przy hover */
  .nav-item-link::after{
    content:attr(data-label);
    position:absolute;left:62px;top:50%;transform:translateY(-50%);
    background:#1a2b4a;color:#fff;
    font-size:.75rem;font-weight:600;
    padding:4px 10px;border-radius:6px;
    white-space:nowrap;pointer-events:none;
    opacity:0;transition:opacity .15s;
    box-shadow:0 2px 8px rgba(0,0,0,.25);
    z-index:9999;
  }
  .nav-item-link:hover::after{opacity:1}
  .nav-item-link.active .icon{
    background:#EF9F2720;border-radius:6px;
  }
  .main-content{margin-left:56px;padding:1.25rem 1rem}
  /* Logo — tylko kropka */
  .sidebar-logo .brand{font-size:1.2rem;text-align:center;display:block}
}

/* Mobile (< 768px) — sidebar chowana, topbar widoczny */
@media(max-width:767px){
  .sidebar{transform:translateX(-100%);width:240px}
  .sidebar.mobile-open{transform:translateX(0)}
  .sidebar .brand-text,.sidebar .sub,
  .nav-section,.nav-item-link span,
  .nav-season{display:block}
  .nav-item-link{
    justify-content:flex-start;
    padding:.58rem .9rem;margin:1px .5rem;
  }
  .nav-item-link::after{display:none}
  .nav-item-link .icon{width:20px;height:auto;font-size:.95rem}
  .topbar{display:flex}
  .main-content{margin-left:0;padding:4.5rem 1rem 1.5rem}
  .hero{padding:1rem 1.1rem}
  .stat-val{font-size:1.2rem}
  .page-title{font-size:1.05rem}
  .table th,.table td{font-size:.72rem;padding:.3rem .4rem}
}

/* Bardzo małe (< 480px) */
@media(max-width:479px){
  .main-content{padding:4.5rem .75rem 1.5rem}
  .hero{padding:.85rem 1rem}
  .stat-card{padding:.65rem .5rem}
  .stat-val{font-size:1rem}
  .stat-val.sm{font-size:.9rem}
  .stat-lbl{font-size:.6rem}
  .card .card-body{padding:.75rem .65rem !important}
  .btn-sm{font-size:.72rem;padding:.25rem .5rem}
  .nav-tabs .nav-link{font-size:.75rem;padding:.35rem .5rem}
}
.nav-submenu{background:#ffffff08;border-radius:6px;margin:2px 6px 4px}
.nav-group>.nav-item-link{border-radius:8px;display:flex;align-items:center}
@media(max-width:991px){.nav-group .brand-text,.nav-submenu{display:none!important}}
</style>
"""

def nav(active="home"):
    gtk_name = "GTK"
    season = "2024/25"
    try:
        gtk_name = get_setting("gtk_name") or "GTK"
        season   = get_setting("current_season") or "2024/25"
    except: pass

    items = [
        ("home",     "/",          "🏠", "Strona główna"),
        ("history",  "/historia",  "📋", "Historia meczów"),
        ("season",   "/sezon",     "📊", "Statystyki drużyny"),
        ("settings", "/ustawienia","⚙️", "Ustawienia"),
    ]
    team_items = [
        ("roster",   "/roster",    "👥", "Skład drużyny"),
        ("players",  "/zawodnicy", "📈", "Statystyki indywidualne"),
    ]
    links = ""
    for key, href, icon, label in items:
        cls = "nav-item-link active" if active==key else "nav-item-link"
        links += (f'<a href="{href}" class="{cls}" data-label="{label}">'
                  f'<span class="icon">{icon}</span>'
                  f'<span class="brand-text">{label}</span></a>')

    # Drużyna — sekcja z podmenu
    team_open = active in ("players","roster")
    team_links = ""
    for key, href, icon, label in team_items:
        cls = "nav-item-link active" if active==key else "nav-item-link"
        team_links += (f'<a href="{href}" class="{cls}" data-label="{label}" '
                       f'style="padding-left:1.4rem;font-size:.8rem">'
                       f'<span class="icon" style="font-size:.8rem">{icon}</span>'
                       f'<span class="brand-text">{label}</span></a>')

    team_section = f"""
<div class="nav-group brand-text">
  <div class="nav-item-link {'active' if team_open else ''}"
       style="cursor:pointer" data-label="Drużyna"
       onclick="this.nextElementSibling.style.display=this.nextElementSibling.style.display=='none'?'block':'none'">
    <span class="icon">🏀</span>
    <span class="brand-text">Drużyna</span>
    <span class="brand-text ms-auto" style="font-size:.7rem;opacity:.6">{'▼' if team_open else '▶'}</span>
  </div>
  <div class="nav-submenu" style="display:{'block' if team_open else 'none'}">
    {team_links}
  </div>
</div>"""

    return f"""
<!-- TOPBAR mobile -->
<div class="topbar">
  <button class="hamburger" id="hamburger" aria-label="Menu" onclick="toggleSidebar()">
    <span></span><span></span><span></span>
  </button>
  <div class="t-brand"><span>●</span> Basket Kołcz</div>
</div>

<!-- OVERLAY mobile -->
<div class="sidebar-overlay" id="sidebarOverlay" onclick="toggleSidebar()"></div>

<!-- SIDEBAR -->
<div class="sidebar" id="sidebar">
  <div class="sidebar-logo">
    <div class="brand"><span>●</span> <span class="brand-text">Basket Kołcz</span></div>
    <div class="sub brand-text">Analytics Platform</div>
  </div>
  <div class="nav-season brand-text"><strong>{gtk_name}</strong>Sezon {season}</div>
  <div class="nav-section brand-text">Nawigacja</div>
  {links}
  {team_section}
</div>

<script>
function toggleSidebar(){{
  const s=document.getElementById('sidebar');
  const o=document.getElementById('sidebarOverlay');
  const h=document.getElementById('hamburger');
  s.classList.toggle('mobile-open');
  o.style.display=s.classList.contains('mobile-open')?'block':'none';
  h.classList.toggle('open');
}}
// Zamknij przy resize
window.addEventListener('resize',()=>{{
  if(window.innerWidth>767){{
    document.getElementById('sidebar').classList.remove('mobile-open');
    document.getElementById('sidebarOverlay').style.display='none';
    document.getElementById('hamburger').classList.remove('open');
  }}
}});
</script>"""

def base(content, scripts="", active="home"):
    # Flash messages
    flash_html = ""
    try:
        from flask import get_flashed_messages
        msgs = get_flashed_messages(with_categories=True)
        for cat, msg in msgs:
            css = "flash-success" if cat == "success" else "flash-error"
            flash_html += f'<div class="{css} flash-msg">{msg}</div>'
    except: pass

    return f"""<!DOCTYPE html>
<html lang="pl"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<meta name="apple-mobile-web-app-capable" content="yes">
<title>Basket Kołcz Analytics</title>
{CSS}
</head>
<body>
{nav(active)}
<div class="main-content">
{flash_html}
{content}
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
{scripts}
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — STRONA GŁÓWNA
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    try: init_db()
    except: pass

    gtk_name = "GTK"
    season = "2024/25"
    try:
        gtk_name = get_setting("gtk_name") or "GTK"
        season   = get_setting("current_season") or "2024/25"
    except: pass

    # Ostatnie 5 meczów
    recent = []
    try:
        db = get_db(); cur = db.cursor()
        try:
            cur.execute("""SELECT id,data_meczu,przeciwnik,
                           COALESCE(rozgrywki,'') as rozgrywki,
                           COALESCE(runda,'') as runda,
                           COALESCE(miejsce,'') as miejsce,
                           wynik_gtk,wynik_opp
                           FROM matches WHERE sezon=%s ORDER BY created_at DESC LIMIT 5""", (season,))
        except:
            cur.execute("""SELECT id,data_meczu,przeciwnik,
                           ''::text as rozgrywki,''::text as runda,''::text as miejsce,
                           wynik_gtk,wynik_opp
                           FROM matches WHERE sezon=%s ORDER BY created_at DESC LIMIT 5""", (season,))
        recent = cur.fetchall(); cur.close()
    except: pass

    recent_rows = ""
    for m in recent:
        wynik = f"{m['wynik_gtk']}:{m['wynik_opp']}"
        if m['wynik_gtk'] > m['wynik_opp']:   badge = '<span class="badge-win">W</span>'
        elif m['wynik_gtk'] < m['wynik_opp']: badge = '<span class="badge-loss">L</span>'
        else:                                  badge = '<span class="badge-draw">D</span>'
        dt       = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else '—'
        rozg     = m.get('rozgrywki','') or '—'
        runda    = m.get('runda','') or '—'
        miejsce  = m.get('miejsce','') or '—'
        recent_rows += f"""<tr>
            <td>{badge}</td>
            <td style="font-size:.8rem;color:#555">{dt}</td>
            <td style="font-size:.78rem;color:#888">{rozg}</td>
            <td style="font-size:.78rem;color:#888">{runda}</td>
            <td><a href="/mecz/{m['id']}" class="fw-bold text-decoration-none" style="color:#1a2b4a">{m['przeciwnik']}</a></td>
            <td style="font-size:.78rem;color:#888">{miejsce}</td>
            <td class="text-center fw-bold">{wynik}</td>
        </tr>"""

    content = f"""
<div class="hero mb-3">
  <div class="d-flex justify-content-between align-items-start flex-wrap gap-2">
    <div>
      <h1 style="font-size:1.5rem;font-weight:700;margin:0">{gtk_name} Analytics</h1>
      <p style="opacity:.75;margin:4px 0 0;font-size:.85rem">Sezon {season} · Wgraj nowy mecz</p>
    </div>
    <div class="d-flex gap-2 flex-wrap">
      <a href="/template/zapis" class="btn btn-outline-light btn-sm">📝 Zapis meczu</a>
      <a href="/template/szablon" class="btn btn-outline-light btn-sm">📋 Szablon</a>
    </div>
  </div>
</div>

<div class="row g-3">
<div class="col-lg-7">
  <div class="card p-3">
    <div class="section-hdr">Wgraj nowy mecz</div>
    <form method="POST" action="/upload" enctype="multipart/form-data">
      <div class="mb-2">
        <label class="form-label" style="font-size:.8rem;font-weight:600">Sezon</label>
        <input type="text" name="sezon" class="form-control form-control-sm" value="{season}" required>
        <div class="form-text" style="font-size:.72rem">Data meczu zostanie odczytana automatycznie z zakładki META w pliku.</div>
      </div>
      <div class="upload-zone" onclick="document.getElementById('fup').click()">
        <div style="font-size:2rem;margin-bottom:.5rem">📊</div>
        <h6 class="fw-bold mb-1" style="color:#1a2b4a">Wgraj plik zapis.xlsx</h6>
        <p class="text-muted mb-0" style="font-size:.8rem">Format: drużyna 1 = arkusz 1 ({gtk_name}), drużyna 2 = arkusz 2 (przeciwnik)</p>
        <input type="file" id="fup" name="file" accept=".xlsx" class="d-none" onchange="this.form.submit()">
      </div>
    </form>
  </div>
</div>

<div class="col-lg-5">
  <div class="card p-3">
    <div class="section-hdr">Ostatnie mecze — {season}</div>
    {'<p class="text-muted" style="font-size:.82rem">Brak meczów w tym sezonie</p>' if not recent_rows else f'<div class="table-responsive"><table class="table table-hover mb-0"><thead><tr><th style="width:36px"></th><th>Data</th><th>Rozgrywki</th><th>Runda</th><th>Przeciwnik</th><th>Miejsce</th><th class="text-center">Wynik</th></tr></thead><tbody>{recent_rows}</tbody></table></div>'}
    <div class="mt-2">
      <a href="/historia" class="btn btn-outline-primary btn-sm w-100">Zobacz wszystkie mecze →</a>
    </div>
  </div>
</div>
</div>"""

    return render_template_string(base(content, active="home"))

# ══════════════════════════════════════════════════════════════════════════════
# WALIDACJA
# ══════════════════════════════════════════════════════════════════════════════

VALID_CODES = {
    "2","0/2","3","0/3","BR","P","F",
    "2+1","2+0","3+1","3+0",
    "2D","0/2D","2D+1","2D+0/1W","2D+1/1W",
    "1/2W","2/2W","0/2W",
    "1/3W","2/3W","3/3W","0/3W",
    "1/2WL","2/2WL","0/2WL",
    "1/1WT","0/1WT",
    "0D+0/2W","0D+1/2W"
}

def read_meta(wb):
    """Odczytaj arkusz META i zwróć słownik danych"""
    meta = {}
    for sn in wb.sheetnames:
        if sn.upper() == "META":
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, values_only=False):
                if not row[0].value: continue
                key = str(row[0].value).strip().lower()
                # Obsłuż datetime z Excela
                raw_val = row[1].value if len(row) > 1 else None
                from datetime import datetime as _dt, date as _date
                if isinstance(raw_val, (_dt, _date)):
                    val = raw_val  # zachowaj jako obiekt — obsłużony w _do_save
                else:
                    val = str(raw_val).strip() if raw_val is not None else ""
                if "drużyna a" in key or "druzyna a" in key: meta["nazwa_a"] = val
                if "drużyna b" in key or "druzyna b" in key: meta["nazwa_b"] = val
                if "wynik a"   in key: meta["wynik_a"] = str(val) if val else ""
                if "wynik b"   in key: meta["wynik_b"] = str(val) if val else ""
                if "data"      in key: meta["data"] = raw_val  # surowa wartość
                if "rozgrywki" in key: meta["rozgrywki"] = val
                if "runda"     in key or "kolejka" in key: meta["runda"] = val
                if "miejsce"   in key: meta["miejsce"] = val
            break
    return meta

def validate_workbook(wb):
    """
    Zwraca dict:
      errors   — błędy krytyczne (lista str)
      warnings — ostrzeżenia (lista str)
      info     — informacje (lista str)
      meta     — dane z META
      names    — (name_a, name_b)
    """
    errors = []; warnings = []; info = []

    sheets_data = [s for s in wb.sheetnames if s.upper() not in ("META","KODY","LEGENDA")]

    if len(sheets_data) < 2:
        errors.append(f"Brakuje arkuszy danych — znaleziono {len(sheets_data)}, wymagane 2 (GTK + przeciwnik)")
        return {"errors":errors,"warnings":warnings,"info":info,"meta":{},"names":("","") }

    name_a, name_b = sheets_data[0], sheets_data[1]
    meta = read_meta(wb)

    # ── 1. Nazwy drużyn ────────────────────────────────────────────────────
    meta_a = meta.get("nazwa_a","")
    meta_b = meta.get("nazwa_b","")
    if meta_a and meta_a not in ("TWOJA_DRUZYNA","TWOJA_DRUŻYNA","-",""):
        if meta_a.upper() != name_a.upper():
            warnings.append(
                f"Nazwa drużyny A w META (<b>{meta_a}</b>) "
                f"różni się od nazwy arkusza (<b>{name_a}</b>). "
                f"Aplikacja użyje nazwy arkusza."
            )
        else:
            info.append(f"✓ Nazwa drużyny A: <b>{name_a}</b> — zgodna z META")
    if meta_b and meta_b not in ("RYWAL","-",""):
        if meta_b.upper() != name_b.upper():
            warnings.append(
                f"Nazwa drużyny B w META (<b>{meta_b}</b>) "
                f"różni się od nazwy arkusza (<b>{name_b}</b>). "
                f"Aplikacja użyje nazwy arkusza."
            )
        else:
            info.append(f"✓ Nazwa drużyny B: <b>{name_b}</b> — zgodna z META")

    # ── 2. Parsuj i sprawdź wyniki ─────────────────────────────────────────
    stats_a = parse_team_sheet(wb[name_a])
    stats_b = parse_team_sheet(wb[name_b])
    suma_a  = suma_quarters(stats_a)
    suma_b  = suma_quarters(stats_b)
    pts_a   = suma_a.get("pts",0)
    pts_b   = suma_b.get("pts",0)

    # Wynik końcowy z META
    try:
        meta_pts_a = int(meta.get("wynik_a","")) if meta.get("wynik_a") else None
        meta_pts_b = int(meta.get("wynik_b","")) if meta.get("wynik_b") else None
    except: meta_pts_a = meta_pts_b = None

    if meta_pts_a is not None:
        q_pts_a = [stats_a["quarter"].get(q,{}).get("pts",0) for q in [1,2,3,4]]
        if meta_pts_a != pts_a:
            errors.append({
                "msg": f"Wynik <b>{name_a}</b> w META: <b>{meta_pts_a}</b> pkt, suma z kodowania: <b>{pts_a}</b> pkt (różnica: {pts_a-meta_pts_a:+d})",
                "quarters": q_pts_a,
                "total": pts_a,
                "meta": meta_pts_a,
            })
        else:
            info.append(f"✓ Wynik {name_a}: <b>{pts_a}</b> pkt — zgodny z META")

    if meta_pts_b is not None:
        q_pts_b = [stats_b["quarter"].get(q,{}).get("pts",0) for q in [1,2,3,4]]
        if meta_pts_b != pts_b:
            errors.append({
                "msg": f"Wynik <b>{name_b}</b> w META: <b>{meta_pts_b}</b> pkt, suma z kodowania: <b>{pts_b}</b> pkt (różnica: {pts_b-meta_pts_b:+d})",
                "quarters": q_pts_b,
                "total": pts_b,
                "meta": meta_pts_b,
            })
        else:
            info.append(f"✓ Wynik {name_b}: <b>{pts_b}</b> pkt — zgodny z META")

    # ── 3. Wynik per kwarta ────────────────────────────────────────────────
    for sheet_name, stats in [(name_a, stats_a), (name_b, stats_b)]:
        total_check = sum(stats["quarter"].get(q,{}).get("pts",0) for q in [1,2,3,4])
        suma = suma_quarters(stats)
        if total_check != suma.get("pts",0):
            warnings.append(
                f"⚠️ {sheet_name}: suma kwart ({total_check}) "
                f"≠ łączna suma punktów ({suma.get('pts',0)})"
            )
        else:
            q_pts = [stats["quarter"].get(q,{}).get("pts",0) for q in [1,2,3,4]]
            info.append(f"✓ {sheet_name} per kwarta: {q_pts[0]}+{q_pts[1]}+{q_pts[2]}+{q_pts[3]} = <b>{total_check}</b>")

    # ── 4. Brakujące dane (puste kolumny A/B/C) ────────────────────────────
    for sheet_name in [name_a, name_b]:
        ws = wb[sheet_name]
        empty_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=200, values_only=True), 2):
            if not any(v is not None for v in row[:4]): break
            missing = []
            if row[0] is None: missing.append("Kwarta(A)")
            if row[2] is None: missing.append("Kod(C)")
            if missing:
                empty_rows.append(f"wiersz {i}: brak {', '.join(missing)}")
        if empty_rows:
            sample = empty_rows[:5]
            warnings.append(
                f"⚠️ {sheet_name} — brakujące dane w {len(empty_rows)} wierszach: "
                f"{'; '.join(sample)}"
                + (f" (i {len(empty_rows)-5} więcej...)" if len(empty_rows)>5 else "")
            )
        else:
            info.append(f"✓ {sheet_name}: wszystkie wiersze mają wymagane dane")

    # ── 5. Nieznane kody akcji ─────────────────────────────────────────────
    for sheet_name in [name_a, name_b]:
        ws = wb[sheet_name]
        unknown = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=500, values_only=True), 2):
            if not any(v is not None for v in row[:4]): break
            raw_c = str(row[2]).strip() if row[2] is not None else ""
            if not raw_c: continue
            for code in [c.strip().upper() for c in raw_c.split(";") if c.strip()]:
                if code not in VALID_CODES:
                    if code not in unknown: unknown[code] = []
                    unknown[code].append(i)
        if unknown:
            details = "; ".join(
                f"<b>{k}</b> (wiersze: {', '.join(map(str,v[:3]))}{'...' if len(v)>3 else ''})"
                for k,v in list(unknown.items())[:8]
            )
            errors.append(
                f"❌ {sheet_name} — nieznane kody akcji: {details}"
            )
        else:
            info.append(f"✓ {sheet_name}: wszystkie kody akcji są prawidłowe")

    return {
        "errors":   errors,
        "warnings": warnings,
        "info":     info,
        "meta":     meta,
        "names":    (name_a, name_b),
        "pts":      (pts_a, pts_b),
    }

# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD
# ══════════════════════════════════════════════════════════════════════════════

import tempfile, base64

def _do_save(wb, name_gtk, name_opp, sezon, data_meczu):
    """Właściwy zapis meczu do bazy"""
    try:
        init_db()
    except: pass

    meta = read_meta(wb)

    # Nazwy z META
    display_gtk = meta.get("nazwa_a","") or name_gtk
    display_opp = meta.get("nazwa_b","") or name_opp
    if display_gtk in ("TWOJA_DRUZYNA","TWOJA_DRUŻYNA","-",""): display_gtk = name_gtk
    if display_opp in ("RYWAL","-",""): display_opp = name_opp

    # Data — META ma priorytet nad formularzem
    meta_data = meta.get("data")
    if meta_data:
        from datetime import datetime as dt2, date
        # Jeśli Excel zwrócił obiekt datetime/date
        if isinstance(meta_data, (dt2, date)):
            data_meczu = meta_data.strftime('%Y-%m-%d') if isinstance(meta_data, dt2) else meta_data.isoformat()
        else:
            raw = str(meta_data).strip().replace(" ","").replace("\n","")
            # Spróbuj różnych formatów daty
            parsed = None
            for fmt in ('%d.%m.%Y','%d/%m/%Y','%Y-%m-%d','%d-%m-%Y',
                        '%Y.%m.%d','%d.%m.%y','%d/%m/%y'):
                try:
                    parsed = dt2.strptime(raw, fmt).strftime('%Y-%m-%d')
                    break
                except: pass
            if parsed:
                data_meczu = parsed
            # Jeśli nie udało się sparsować — zostaw datę z formularza
    
    # Sezon z META jeśli dostępny
    if meta.get("rozgrywki") and not sezon:
        sezon = meta["rozgrywki"]

    stats_gtk = parse_team_sheet(wb[name_gtk])
    stats_opp = parse_team_sheet(wb[name_opp])
    match_id  = save_match_to_db(
        display_opp, display_gtk, sezon, data_meczu,
        stats_gtk, stats_opp,
        rozgrywki=str(meta.get("rozgrywki","") or ""),
        runda=str(meta.get("runda","") or ""),
        miejsce=str(meta.get("miejsce","") or ""),
    )
    session.clear()
    session["last_match_id"] = match_id
    flash(f"✓ Mecz {display_gtk} vs {display_opp} zapisany pomyślnie!","success")
    return redirect(url_for("mecz", match_id=match_id))


# Tymczasowy storage dla plików oczekujących na potwierdzenie
# Używamy katalogu /tmp (dostępny na Render)
PENDING_DIR = "/tmp/basketkolcz_pending"
os.makedirs(PENDING_DIR, exist_ok=True)


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        flash("Nie wybrano pliku","error"); return redirect(url_for("index"))
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        flash("Plik musi być .xlsx","error"); return redirect(url_for("index"))

    sezon = request.form.get("sezon","2024/25")

    try:
        file_bytes = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        # ── Odczytaj datę z META ──────────────────────────────────────────
        meta = read_meta(wb)
        data_meczu = None
        if meta.get("data"):
            raw = str(meta["data"]).strip().replace(" ","")
            for fmt in ('%d.%m.%Y','%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%Y.%m.%d'):
                try:
                    from datetime import datetime as dt2
                    data_meczu = dt2.strptime(raw, fmt).strftime('%Y-%m-%d')
                    break
                except: pass
            if not data_meczu:
                # Fallback — dateutil
                try:
                    from dateutil import parser as dp
                    data_meczu = dp.parse(raw, dayfirst=True).strftime('%Y-%m-%d')
                except: pass

        # ── Walidacja ─────────────────────────────────────────────────────
        report = validate_workbook(wb)
        has_issues = bool(report["errors"] or report["warnings"])

        if has_issues:
            import uuid
            token = str(uuid.uuid4())
            tmp_path = os.path.join(PENDING_DIR, f"{token}.xlsx")
            with open(tmp_path, "wb") as fp:
                fp.write(file_bytes)

            import re as _re
            def clean(s):
                if isinstance(s, dict):
                    return {
                        "msg": _re.sub(r'<[^>]+>', '', s.get("msg",""))[:150],
                        "quarters": s.get("quarters",[]),
                        "total": s.get("total",0),
                        "meta": s.get("meta",None),
                    }
                return _re.sub(r'<[^>]+>', '', str(s))[:120]

            session.clear()
            session["pt"] = token
            session["ps"] = sezon
            session["pd"] = data_meczu or ""
            session["vr"] = {
                "e": [clean(e) for e in report["errors"][:8]],
                "w": [clean(w) for w in report["warnings"][:8]],
                "i": [clean(i) for i in report["info"][:6]],
                "n": list(report["names"]),
                "p": list(report["pts"]),
            }
            return redirect(url_for("validation_report"))

        # ── Brak problemów — zapisz od razu ──────────────────────────────
        return _do_save(wb, report["names"][0], report["names"][1], sezon, data_meczu)

    except Exception as e:
        flash(f"Błąd wgrywania: {str(e)}","error")
        return redirect(url_for("index"))        # Walidacja
        report = validate_workbook(wb)

        has_issues = bool(report["errors"] or report["warnings"])

        if has_issues:
            # Zapisz plik do /tmp z unikalnym tokenem
            import uuid
            token = str(uuid.uuid4())
            tmp_path = os.path.join(PENDING_DIR, f"{token}.xlsx")
            with open(tmp_path, "wb") as fp:
                fp.write(file_bytes)

            # Zapisz tylko lekkie dane w sesji (bez pliku!)
            import re as _re
            def clean(s):
                if isinstance(s, dict):
                    # Błąd z kwartami — zachowaj strukturę ale skróć msg
                    return {
                        "msg": _re.sub(r'<[^>]+>', '', s.get("msg",""))[:150],
                        "quarters": s.get("quarters",[]),
                        "total": s.get("total",0),
                        "meta": s.get("meta",None),
                    }
                return _re.sub(r'<[^>]+>', '', str(s))[:120]

            session.clear()
            session["pt"] = token
            session["ps"] = sezon
            session["pd"] = data_meczu or ""
            session["vr"] = {
                "e": [clean(e) for e in report["errors"][:8]],
                "w": [clean(w) for w in report["warnings"][:8]],
                "i": [clean(i) for i in report["info"][:6]],
                "n": list(report["names"]),
                "p": list(report["pts"]),
            }
            return redirect(url_for("validation_report"))

        # Brak problemów — zapisz od razu
        return _do_save(wb, report["names"][0], report["names"][1], sezon, data_meczu)

    except Exception as e:
        import traceback
        try:
            get_db().rollback()
        except: pass
        flash(f"Błąd wgrywania: {str(e)}","error")
        return redirect(url_for("index"))


@app.route("/walidacja/pobierz-z-bledami")
def download_with_errors():
    """Pobierz oryginalny plik z zaznaczonymi błędami na czerwono"""
    token = session.get("pt","")
    if not token:
        flash("Sesja wygasła — wgraj plik ponownie","error")
        return redirect(url_for("index"))

    tmp_path = os.path.join(PENDING_DIR, f"{token}.xlsx")
    if not os.path.exists(tmp_path):
        flash("Plik tymczasowy wygasł — wgraj ponownie","error")
        return redirect(url_for("index"))

    try:
        with open(tmp_path, "rb") as fp:
            file_bytes = fp.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

        RED_FILL   = PatternFill("solid", fgColor="FFCDD2")
        RED_FONT   = Font(color="B71C1C", bold=True)
        RED_BORDER = Border(
            left=Side(style="medium", color="E53935"),
            right=Side(style="medium", color="E53935"),
            top=Side(style="medium", color="E53935"),
            bottom=Side(style="medium", color="E53935"),
        )
        COMMENT_FILL = PatternFill("solid", fgColor="FFEBEE")

        report = validate_workbook(wb)

        # Zbierz błędy per arkusz per wiersz
        error_cells = {}  # {(sheet_name, row): [opisy]}

        sheets_data = [s for s in wb.sheetnames if s.upper() not in ("META","KODY","LEGENDA")]
        if len(sheets_data) >= 2:
            name_a, name_b = sheets_data[0], sheets_data[1]

            for sheet_name in [name_a, name_b]:
                ws = wb[sheet_name]

                # ── Wynik końcowy (cały arkusz)
                suma = suma_quarters(parse_team_sheet(ws))
                meta = read_meta(wb)
                key = "wynik_a" if sheet_name == name_a else "wynik_b"
                try:
                    meta_pts = int(meta.get(key, "")) if meta.get(key) else None
                    coded_pts = suma.get("pts", 0)
                    if meta_pts is not None and meta_pts != coded_pts:
                        # Zaznacz nagłówek arkusza (wiersz 1)
                        k = (sheet_name, 1)
                        if k not in error_cells: error_cells[k] = []
                        error_cells[k].append(
                            f"Wynik w META: {meta_pts}, kodowanie: {coded_pts} (różnica: {coded_pts-meta_pts:+d})"
                        )
                except: pass

                # ── Brakujące dane (kolumny A/B/C)
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=500, values_only=False), 2):
                    if not any(c.value is not None for c in row[:4]): break
                    missing = []
                    if row[0].value is None: missing.append("Kwarta(A)")
                    if row[2].value is None: missing.append("Kod(C)")
                    if missing:
                        k = (sheet_name, i)
                        if k not in error_cells: error_cells[k] = []
                        error_cells[k].append(f"Brak: {', '.join(missing)}")

                # ── Nieznane kody akcji
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=500, values_only=False), 2):
                    if not any(c.value is not None for c in row[:4]): break
                    raw_c = str(row[2].value).strip() if row[2].value is not None else ""
                    if not raw_c: continue
                    bad_codes = [c.strip().upper() for c in raw_c.split(";")
                                 if c.strip() and c.strip().upper() not in VALID_CODES]
                    if bad_codes:
                        k = (sheet_name, i)
                        if k not in error_cells: error_cells[k] = []
                        error_cells[k].append(f"Nieznany kod: {', '.join(bad_codes)}")

        # ── Zaznacz błędy w pliku ───────────────────────────────────────────
        for (sheet_name, row_idx), descs in error_cells.items():
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]

            if row_idx == 1:
                # Błąd wyniku — zaznacz wiersz nagłówka i dodaj komentarz w A1
                for col in range(1, 15):
                    c = ws.cell(row_idx, col)
                    c.fill = PatternFill("solid", fgColor="FFCCBC")
                # Wpisz info w wolnym miejscu
                ws.cell(1, 16).value = "⚠ " + " | ".join(descs)
                ws.cell(1, 16).font = Font(color="BF360C", bold=True, size=9)
                ws.cell(1, 16).fill = PatternFill("solid", fgColor="FFF3E0")
            else:
                # Błąd w wierszu danych — zaznacz cały wiersz
                for col in range(1, 12):
                    c = ws.cell(row_idx, col)
                    c.fill = RED_FILL
                    if col in [1, 3]:  # Kwarta i Kod — główne kolumny z błędem
                        c.font = RED_FONT
                        c.border = RED_BORDER
                # Opis błędu w kolumnie O
                desc_cell = ws.cell(row_idx, 15)
                desc_cell.value = "❌ " + " | ".join(descs)
                desc_cell.font = Font(color="B71C1C", bold=True, size=9)
                desc_cell.fill = COMMENT_FILL

        # ── Dodaj podsumowanie punktów per kwarta (prawa strona) ──────────────
        HDR_BLUE  = PatternFill("solid", fgColor="1A2B4A")
        HDR_WHITE = Font(color="FFFFFF", bold=True, size=9)
        SUM_FILL  = PatternFill("solid", fgColor="E3F2FD")
        SUM_FONT  = Font(bold=True, size=9, color="0C447C")
        OK_FILL   = PatternFill("solid", fgColor="E8F5E9")
        ERR_FILL  = PatternFill("solid", fgColor="FFEBEE")

        meta = read_meta(wb)

        for idx, sheet_name in enumerate(sheets_data):
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            stats = parse_team_sheet(ws)

            # Znajdź ostatni wiersz z danymi
            last_row = 1
            for r in ws.iter_rows(min_row=2, max_row=600, values_only=True):
                if not any(v is not None for v in r[:4]): break
                last_row += 1

            # Kolumny podsumowania — zaczynamy od P (16)
            COL_START = 17  # kolumna Q

            # Nagłówki
            for ci, lbl in enumerate(["KWARTA","PKT","2PM/A","3PM/A","FTM/A","BR","POSS"]):
                c = ws.cell(1, COL_START + ci, lbl)
                c.fill = HDR_BLUE; c.font = HDR_WHITE
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(COL_START + ci)].width = 9

            ws.column_dimensions[get_column_letter(COL_START)].width = 8

            # Dane per kwarta
            total_pts = 0
            for qi, qn in enumerate([1,2,3,4]):
                qd = stats["quarter"].get(qn, {})
                r = 2 + qi
                pts_q = qd.get("pts",0)
                total_pts += pts_q

                vals = [
                    f"{qn}Q",
                    pts_q,
                    f"{qd.get('p2m',0)}/{qd.get('p2a',0)}",
                    f"{qd.get('p3m',0)}/{qd.get('p3a',0)}",
                    f"{qd.get('ftm',0)}/{qd.get('fta',0)}",
                    qd.get("br",0),
                    qd.get("poss",0),
                ]
                for ci, v in enumerate(vals):
                    c = ws.cell(r, COL_START + ci, v)
                    c.fill = SUM_FILL; c.font = Font(size=9)
                    c.alignment = Alignment(horizontal="center")

            # Wiersz SUMA
            r_sum = 6
            ws.cell(r_sum, COL_START, "SUMA").fill = PatternFill("solid", fgColor="1A2B4A")
            ws.cell(r_sum, COL_START).font = Font(color="FFFFFF", bold=True, size=9)
            ws.cell(r_sum, COL_START).alignment = Alignment(horizontal="center")

            suma_all = suma_quarters(stats)
            suma_vals = [
                total_pts,
                f"{suma_all.get('p2m',0)}/{suma_all.get('p2a',0)}",
                f"{suma_all.get('p3m',0)}/{suma_all.get('p3a',0)}",
                f"{suma_all.get('ftm',0)}/{suma_all.get('fta',0)}",
                suma_all.get("br",0),
                suma_all.get("poss",0),
            ]
            for ci, v in enumerate(suma_vals):
                c = ws.cell(r_sum, COL_START + 1 + ci, v)
                c.font = Font(bold=True, size=9); c.alignment = Alignment(horizontal="center")
                c.fill = PatternFill("solid", fgColor="BBDEFB")

            # Porównanie z META
            meta_key = "wynik_a" if idx == 0 else "wynik_b"
            try:
                meta_pts = int(meta.get(meta_key,"")) if meta.get(meta_key) else None
            except: meta_pts = None

            r_meta = 8
            if meta_pts is not None:
                ws.cell(r_meta, COL_START, "META").fill = PatternFill("solid", fgColor="37474F")
                ws.cell(r_meta, COL_START).font = Font(color="FFFFFF", bold=True, size=9)
                ws.cell(r_meta, COL_START).alignment = Alignment(horizontal="center")

                match = meta_pts == total_pts
                fill = OK_FILL if match else ERR_FILL
                font_col = "1B5E20" if match else "B71C1C"

                ws.cell(r_meta, COL_START+1, meta_pts).fill = fill
                ws.cell(r_meta, COL_START+1).font = Font(bold=True, size=9, color=font_col)
                ws.cell(r_meta, COL_START+1).alignment = Alignment(horizontal="center")

                ws.cell(r_meta, COL_START+2, "✓ OK" if match else f"❌ RÓŻNICA: {total_pts-meta_pts:+d}")
                ws.cell(r_meta, COL_START+2).fill = fill
                ws.cell(r_meta, COL_START+2).font = Font(bold=True, size=9, color=font_col)
                ws.merge_cells(start_row=r_meta, start_column=COL_START+2,
                               end_row=r_meta, end_column=COL_START+5)

        # ── Dodaj legendę na początku każdego arkusza ──────────────────────
        for sheet_name in sheets_data:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            # Kolumna O nagłówek
            hdr = ws.cell(1, 15)
            hdr.value = "BŁĘDY WALIDACJI"
            hdr.font = Font(color="FFFFFF", bold=True, size=9)
            hdr.fill = PatternFill("solid", fgColor="C62828")
            ws.column_dimensions["O"].width = 40

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        names = session.get("vr", {}).get("n", ["plik","plik"])
        filename = f"BLEDY_{names[0]}_vs_{names[1]}.xlsx".replace(" ","_")
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        flash(f"Błąd generowania pliku: {str(e)}", "error")
        return redirect(url_for("validation_report"))


@app.route("/walidacja")
def validation_report():
    vr = session.get("vr")
    if not vr:
        return redirect(url_for("index"))

    errors   = vr.get("e",[])
    warnings = vr.get("w",[])
    info     = vr.get("i",[])
    names    = vr.get("n",["?","?"])
    pts      = vr.get("p",[0,0])
    has_errors = len(errors) > 0

    # Renderuj błędy — obsługa zarówno dict (z kwartami) jak i str
    def render_error(e, idx):
        if isinstance(e, dict):
            msg      = e.get("msg","")
            quarters = e.get("quarters",[])
            total    = e.get("total",0)
            meta_pts = e.get("meta",None)
            # Pasek per kwarta
            q_html = ""
            if quarters:
                q_items = " | ".join(
                    f'<span style="font-weight:700;color:#{'1a6b3c' if quarters[i]==max(quarters) else 'b71c1c' if quarters[i]==min(quarters) else '444'}">'
                    f'{i+1}Q — {quarters[i]}</span>'
                    for i in range(len(quarters))
                )
                q_html = f"""
                <div style="margin-top:.5rem;padding:.4rem .6rem;background:#fff;border-radius:6px;border:1px solid #ffcdd2;font-size:.82rem">
                  {q_items}
                  <span style="margin-left:.5rem;color:#888;font-size:.75rem">= {total} pkt (META: {meta_pts})</span>
                </div>"""
            return f"""
            <div class="val-item val-error" style="flex-direction:column;cursor:pointer" onclick="this.querySelector('.qdetail').style.display=this.querySelector('.qdetail').style.display=='none'?'block':'none'">
              <div style="display:flex;gap:.75rem;align-items:flex-start;width:100%">
                <span class="val-icon">🚫</span>
                <span style="flex:1">{msg}</span>
                <span style="font-size:.75rem;color:#c62828;flex-shrink:0">▼ rozwiń</span>
              </div>
              <div class="qdetail" style="display:none">{q_html}</div>
            </div>"""
        else:
            return f'<div class="val-item val-error"><span class="val-icon">🚫</span><span>{e}</span></div>'

    def render_warning(w):
        return f'<div class="val-item val-warning"><span class="val-icon">⚠️</span><span>{w}</span></div>'

    def render_info(i):
        return f'<div class="val-item val-info"><span class="val-icon">✓</span><span>{i}</span></div>'

    errors_html   = "".join(render_error(e, i) for i,e in enumerate(errors))
    warnings_html = "".join(render_warning(w) for w in warnings)
    info_html     = "".join(render_info(i) for i in info)

    content = f"""
<div class="page-title">🔍 Raport walidacji pliku</div>

<div class="card mb-3 p-3">
  <div class="d-flex gap-3 align-items-center flex-wrap">
    <div>
      <div style="font-size:.8rem;color:#888">Drużyny</div>
      <div class="fw-bold">{names[0]} vs {names[1]}</div>
    </div>
    <div>
      <div style="font-size:.8rem;color:#888">Wynik z kodowania</div>
      <div class="fw-bold">{pts[0]} : {pts[1]}</div>
    </div>
    <div class="ms-auto d-flex gap-2 flex-wrap">
      <a href="/" class="btn btn-outline-secondary btn-sm">← Anuluj</a>
      {'<a href="/walidacja/pobierz-z-bledami" class="btn btn-outline-danger btn-sm fw-bold">📥 Pobierz plik z błędami</a>' if has_errors else ''}
      {'<span class="btn btn-secondary btn-sm disabled">Zapisz (popraw błędy)</span>' if has_errors else
       '<form method="POST" action="/upload/force" style="display:inline"><button type="submit" class="btn btn-success btn-sm fw-bold">✓ Zapisz mimo ostrzeżeń</button></form>'}
    </div>
  </div>
</div>

<style>
.val-section{{margin-bottom:1rem}}
.val-section-title{{font-size:.72rem;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:.5rem;padding:.3rem .6rem;border-radius:6px}}
.val-item{{display:flex;gap:.75rem;align-items:flex-start;padding:.5rem .75rem;border-radius:8px;margin-bottom:.3rem;font-size:.85rem;line-height:1.4}}
.val-icon{{font-size:1rem;flex-shrink:0;margin-top:1px}}
.val-error{{background:#fff0f0;border-left:3px solid #e53935}}
.val-warning{{background:#fffde7;border-left:3px solid #f9a825}}
.val-info{{background:#f1f8e9;border-left:3px solid #43a047}}
</style>

<div class="card p-3">
  {'<div class="val-section"><div class="val-section-title" style="background:#ffebee;color:#c62828">🚫 Błędy krytyczne (' + str(len(errors)) + ') — kliknij aby rozwinąć</div>' + errors_html + '</div>' if errors else ''}
  {'<div class="val-section"><div class="val-section-title" style="background:#fff8e1;color:#f57f17">⚠️ Ostrzeżenia (' + str(len(warnings)) + ')</div>' + warnings_html + '</div>' if warnings else ''}
  {'<div class="val-section"><div class="val-section-title" style="background:#e8f5e9;color:#2e7d32">✓ Poprawne (' + str(len(info)) + ')</div>' + info_html + '</div>' if info else ''}
</div>

{'<div class="card p-3 mt-2" style="background:#fff0f0;border:1px solid #ffcdd2"><b>Plik zawiera błędy krytyczne.</b> Popraw plik i wgraj ponownie.</div>' if has_errors else '<div class="card p-3 mt-2" style="background:#fffde7;border:1px solid #fff176"><b>Plik zawiera ostrzeżenia.</b> Możesz zapisać mimo ostrzeżeń lub poprawić plik.</div>'}
"""
    return render_template_string(base(content, active="home"))


@app.route("/upload/force", methods=["POST"])
def upload_force():
    """Zapisz plik mimo ostrzeżeń"""
    token      = session.get("pt","")
    sezon      = session.get("ps","2024/25")
    data_meczu = session.get("pd")

    if not token:
        flash("Sesja wygasła — wgraj plik ponownie","error")
        return redirect(url_for("index"))

    tmp_path = os.path.join(PENDING_DIR, f"{token}.xlsx")
    if not os.path.exists(tmp_path):
        flash("Plik tymczasowy wygasł — wgraj ponownie","error")
        return redirect(url_for("index"))

    try:
        with open(tmp_path, "rb") as fp:
            file_bytes = fp.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        report = validate_workbook(wb)
        # Usuń plik tymczasowy
        try: os.remove(tmp_path)
        except: pass
        return _do_save(wb, report["names"][0], report["names"][1], sezon, data_meczu)
    except Exception as e:
        try: get_db().rollback()
        except: pass
        flash(f"Błąd zapisu: {str(e)}","error")
        return redirect(url_for("index"))

# ══════════════════════════════════════════════════════════════════════════════
# HISTORIA MECZÓW
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/historia")
def historia():
    try: init_db()
    except: pass
    sezon_filter    = request.args.get("sezon","")
    data_od         = request.args.get("data_od","")
    data_do         = request.args.get("data_do","")
    przeciwnik_filter = request.args.get("przeciwnik","").strip().lower()

    db = get_db(); cur = db.cursor()

    # Buduj WHERE
    conditions = []
    params = []
    if sezon_filter:
        conditions.append("sezon=%s"); params.append(sezon_filter)
    if data_od:
        conditions.append("data_meczu >= %s"); params.append(data_od)
    if data_do:
        conditions.append("data_meczu <= %s"); params.append(data_do)
    if przeciwnik_filter:
        conditions.append("LOWER(przeciwnik) LIKE %s"); params.append(f"%{przeciwnik_filter}%")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    try:
        cur.execute(f"""SELECT id,sezon,data_meczu,przeciwnik,
                       COALESCE(rozgrywki,'') as rozgrywki,
                       COALESCE(runda,'') as runda,
                       COALESCE(miejsce,'') as miejsce,
                       wynik_gtk,wynik_opp
                       FROM matches {where}
                       ORDER BY data_meczu DESC NULLS LAST, created_at DESC""", params)
    except:
        cur.execute(f"SELECT * FROM matches {where} ORDER BY created_at DESC", params)
    matches = cur.fetchall()

    cur.execute("SELECT DISTINCT sezon FROM matches ORDER BY sezon DESC")
    sezony = [r["sezon"] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT przeciwnik FROM matches ORDER BY przeciwnik")
    przeciwnicy = [r["przeciwnik"] for r in cur.fetchall()]
    cur.close()

    rows = ""
    for i, m in enumerate(matches):
        wynik = f"{m['wynik_gtk']} : {m['wynik_opp']}"
        if m['wynik_gtk'] > m['wynik_opp']:   badge = '<span class="badge-win">W</span>'
        elif m['wynik_gtk'] < m['wynik_opp']: badge = '<span class="badge-loss">P</span>'
        else:                                  badge = '<span class="badge-draw">R</span>'
        dt      = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else '—'
        rozg    = m.get('rozgrywki','') or '—'
        runda   = m.get('runda','') or '—'
        miejsce = m.get('miejsce','') or '—'
        bg = "background:#f8f9ff" if i%2==0 else ""
        rows += f"""<tr style="{bg}">
            <td style="width:44px">{badge}</td>
            <td style="font-size:.82rem;font-weight:600">{dt}</td>
            <td style="font-size:.78rem;color:#666">{rozg}</td>
            <td style="font-size:.78rem;color:#666">{runda}</td>
            <td><a href="/mecz/{m['id']}" class="fw-bold text-decoration-none" style="color:#1a2b4a">{m['przeciwnik']}</a></td>
            <td style="font-size:.78rem;color:#666">{miejsce}</td>
            <td class="text-center"><span style="font-size:.95rem;font-weight:700">{wynik}</span></td>
            <td class="text-center">
              <a href="/mecz/{m['id']}" class="btn btn-outline-primary btn-sm" style="font-size:.72rem">Raport</a>
              <a href="/mecz/{m['id']}/delete" class="btn btn-outline-danger btn-sm ms-1" style="font-size:.72rem"
                 onclick="return confirm('Usunąć ten mecz?')">✕</a>
            </td>
        </tr>"""

    season_opts = "".join([f'<option value="{s}" {"selected" if s==sezon_filter else ""}>{s}</option>' for s in sezony])
    opp_opts    = "".join([f'<option value="{p}" {"selected" if p.lower()==przeciwnik_filter else ""}>{p}</option>' for p in przeciwnicy])

    content = f"""
<div class="page-title">📋 Historia meczów</div>

<div class="card p-3 mb-3">
  <form method="GET">
    <div class="row g-2 align-items-end">
      <div class="col-auto">
        <label class="form-label mb-1" style="font-size:.75rem;font-weight:600">Sezon</label>
        <select name="sezon" class="form-select form-select-sm" style="width:110px">
          <option value="">Wszystkie</option>
          {season_opts}
        </select>
      </div>
      <div class="col-auto">
        <label class="form-label mb-1" style="font-size:.75rem;font-weight:600">Data od</label>
        <input type="date" name="data_od" class="form-control form-control-sm" value="{data_od}" style="width:140px">
      </div>
      <div class="col-auto">
        <label class="form-label mb-1" style="font-size:.75rem;font-weight:600">Data do</label>
        <input type="date" name="data_do" class="form-control form-control-sm" value="{data_do}" style="width:140px">
      </div>
      <div class="col-auto">
        <label class="form-label mb-1" style="font-size:.75rem;font-weight:600">Przeciwnik</label>
        <select name="przeciwnik" class="form-select form-select-sm" style="width:180px">
          <option value="">Wszyscy</option>
          {opp_opts}
        </select>
      </div>
      <div class="col-auto">
        <button type="submit" class="btn btn-primary btn-sm">Filtruj</button>
        <a href="/historia" class="btn btn-outline-secondary btn-sm ms-1">Wyczyść</a>
      </div>
      <div class="col-auto ms-auto">
        <span style="font-size:.82rem;color:#888;line-height:2.2">{len(matches)} meczów</span>
      </div>
    </div>
  </form>
</div>

<div class="card">
  <div class="card-body p-2">
    <div class="table-responsive">
      <table class="table table-hover mb-0">
        <thead><tr>
          <th style="width:44px"></th>
          <th style="cursor:pointer" onclick="sortTable(1)">Data ↕</th>
          <th>Rozgrywki</th>
          <th>Runda/Kolejka</th>
          <th style="cursor:pointer" onclick="sortTable(4)">Przeciwnik ↕</th>
          <th>Miejsce</th>
          <th class="text-center">Wynik</th>
          <th class="text-center">Akcje</th>
        </tr></thead>
        <tbody id="matchTable">
          {rows if rows else '<tr><td colspan="8" class="text-center text-muted py-4">Brak meczów spełniających kryteria</td></tr>'}
        </tbody>
      </table>
    </div>
  </div>
</div>"""

    scripts = """<script>
let sortDir = {};
function sortTable(col) {
    const tbody = document.getElementById('matchTable');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    sortDir[col] = !sortDir[col];
    rows.sort((a, b) => {
        const av = a.cells[col]?.textContent.trim() || '';
        const bv = b.cells[col]?.textContent.trim() || '';
        // Sortowanie daty DD.MM.YYYY
        if (col === 1) {
            const toDate = s => { const p=s.split('.'); return p.length===3?new Date(p[2],p[1]-1,p[0]):new Date(0); }
            return sortDir[col] ? toDate(bv)-toDate(av) : toDate(av)-toDate(bv);
        }
        return sortDir[col] ? bv.localeCompare(av,'pl') : av.localeCompare(bv,'pl');
    });
    rows.forEach(r => tbody.appendChild(r));
}
</script>"""

    return render_template_string(base(content, scripts, active="history"))

# ══════════════════════════════════════════════════════════════════════════════
# RAPORT MECZU
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/mecz/<int:match_id>")
def mecz(match_id):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m: flash("Mecz nie istnieje","error"); return redirect(url_for("historia"))

    gtk_name = (m.get("nazwa_gtk","") or "").strip() or get_setting("gtk_name") or "GTK"
    name_opp = m["przeciwnik"]

    cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (match_id,))
    all_stats = cur.fetchall()

    cur.execute("SELECT * FROM player_stats WHERE match_id=%s", (match_id,))
    all_players = cur.fetchall()

    cur.execute("SELECT * FROM timing_stats WHERE match_id=%s", (match_id,))
    all_timing = cur.fetchall()

    # Piątki
    try:
        cur.execute("""SELECT * FROM lineup_stats WHERE match_id=%s AND druzyna='gtk'
                       ORDER BY poss DESC""", (match_id,))
        all_lineups = list(cur.fetchall())
    except:
        all_lineups = []

    # Mapa roster_id → "Nazwisko I." dla GTK
    try:
        cur.execute("""SELECT ps.id as ps_id, r.imie, r.nazwisko
                       FROM player_stats ps
                       JOIN roster r ON ps.roster_id = r.id
                       WHERE ps.match_id=%s AND ps.druzyna='gtk'""", (match_id,))
        roster_map = {row["ps_id"]: f"{row['nazwisko']} {row['imie'][0]}." for row in cur.fetchall()}
    except:
        roster_map = {}

    # Mapa nr → nazwisko dla piątek
    try:
        cur.execute("""SELECT ps.nr, r.imie, r.nazwisko
                       FROM player_stats ps
                       JOIN roster r ON ps.roster_id = r.id
                       WHERE ps.match_id=%s AND ps.druzyna='gtk'""", (match_id,))
        nr_name_map = {str(row["nr"]): f"{row['nazwisko']} {row['imie'][0]}." for row in cur.fetchall()}
    except:
        nr_name_map = {}
    cur.close()

    def build_suma(druzyna):
        s = {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"fd":0}
        for row in all_stats:
            if row["druzyna"] == druzyna:
                for k in s: s[k] += row.get(k,0)
        return s

    suma_gtk = build_suma("gtk")
    suma_opp = build_suma("opp")
    kpi_gtk  = calc_kpi(suma_gtk)
    kpi_opp  = calc_kpi(suma_opp)

    dt = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else ""

    # KPI cards
    def kpi_cards(suma, kpi):
        cards = ""
        for val, lbl in [
            (suma.get("pts",0),"Punkty"),
            (kpi["efg"],"eFG%"),
            (kpi["ts"],"TS%"),
            (kpi["ortg"],"ORtg"),
            (kpi["ppp"],"PPP"),
            (kpi["p2_pct"],"2PT%"),
            (kpi["p3_pct"],"3PT%"),
            (kpi["ft_pct"],"FT%"),
        ]:
            cards += f'<div class="col-6 col-sm-3 col-md-2"><div class="stat-card"><div class="stat-val sm">{val}</div><div class="stat-lbl">{lbl}</div></div></div>'
        return cards

    # Per kwarta tabela
    def q_table(druzyna):
        rows = ""
        for qn in [1,2,3,4]:
            qd = next((r for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==qn), {})
            kq = calc_kpi(dict(qd) if qd else {})
            rows += f"""<tr>
                <td><span class="badge" style="background:{'#c8e6c9' if qn==1 else '#bbdefb' if qn==2 else '#fff9c4' if qn==3 else '#fce4ec'};color:#333">{qn}Q</span></td>
                <td><b>{qd.get('pts',0)}</b></td>
                <td>{qd.get('p2m',0)}/{qd.get('p2a',0)}</td><td>{kq['p2_pct']}</td>
                <td>{qd.get('p3m',0)}/{qd.get('p3a',0)}</td><td>{kq['p3_pct']}</td>
                <td>{qd.get('ftm',0)}/{qd.get('fta',0)}</td>
                <td>{qd.get('br',0)}</td><td>{qd.get('poss',0)}</td>
                <td>{kq['efg']}</td><td>{kq['ortg']}</td>
            </tr>"""
        sk = calc_kpi(suma_gtk if druzyna=="gtk" else suma_opp)
        sd = suma_gtk if druzyna=="gtk" else suma_opp
        rows += f"""<tr style="background:#f0f4ff;font-weight:700">
            <td>SUMA</td><td><b>{sd.get('pts',0)}</b></td>
            <td>{sd.get('p2m',0)}/{sd.get('p2a',0)}</td><td>{sk['p2_pct']}</td>
            <td>{sd.get('p3m',0)}/{sd.get('p3a',0)}</td><td>{sk['p3_pct']}</td>
            <td>{sd.get('ftm',0)}/{sd.get('fta',0)}</td>
            <td>{sd.get('br',0)}</td><td>{sd.get('poss',0)}</td>
            <td>{sk['efg']}</td><td>{sk['ortg']}</td>
        </tr>"""
        return f"""<div class="table-responsive"><table class="table table-hover mb-0">
            <thead><tr><th>Q</th><th>PKT</th><th>2PM/A</th><th>2P%</th><th>3PM/A</th><th>3P%</th><th>FTM/A</th><th>BR</th><th>POSS</th><th>eFG%</th><th>ORtg</th></tr></thead>
            <tbody>{rows}</tbody></table></div>"""

    # Zawodnicy tabela
    def p_table(druzyna):
        players = [r for r in all_players if r["druzyna"]==druzyna]
        if not players:
            return '<p class="text-muted p-2" style="font-size:.82rem">Brak danych zawodników</p>'
        rows = ""
        for pd in sorted(players, key=lambda x: x["pts"], reverse=True):
            fga = pd.get("p2a",0)+pd.get("p3a",0)
            fta = pd.get("fta",0); ftm = pd.get("ftm",0)
            pm2 = pd.get("p2m",0); pm3 = pd.get("p3m",0)
            efg = f"{(pm2+1.5*pm3)/fga:.0%}" if fga else "-"
            ts  = f"{pd.get('pts',0)/(2*(fga+0.44*fta)):.0%}" if (fga+fta) else "-"
            # Pokaż nazwisko jeśli przypisany (tylko GTK), inaczej numer
            if druzyna == "gtk" and pd["id"] in roster_map:
                id_cell = f'<td class="fw-bold">#{pd["nr"]} {roster_map[pd["id"]]}</td>'
            else:
                id_cell = f'<td class="fw-bold">#{pd["nr"]}</td>'
            rows += f"""<tr>
                {id_cell}
                <td class="fw-bold" style="color:#1a2b4a">{pd.get('pts',0)}</td>
                <td>{pm2}/{pd.get('p2a',0)}</td>
                <td>{pm3}/{pd.get('p3a',0)}</td>
                <td>{ftm}/{fta}</td>
                <td><b>{efg}</b></td><td>{ts}</td>
                <td>{pd.get('ast',0)}</td><td>{pd.get('oreb',0)}</td>
                <td>{pd.get('dreb',0)}</td><td>{pd.get('br',0)}</td>
                <td>{pd.get('finishes',0)}</td>
            </tr>"""
        hdr_id = "Zawodnik" if druzyna == "gtk" and roster_map else "#"
        return f"""<div class="table-responsive"><table class="table table-hover mb-0">
            <thead><tr><th>{hdr_id}</th><th>PTS</th><th>2PM/A</th><th>3PM/A</th><th>FTM/A</th><th>eFG%</th><th>TS%</th><th>AST</th><th>OREB</th><th>DREB</th><th>BR</th><th>FIN</th></tr></thead>
            <tbody>{rows}</tbody></table></div>"""

    # Piątki
    def lineup_table():
        if not all_lineups:
            return '<p class="text-muted p-3 mb-0" style="font-size:.82rem">Brak danych piątek — wgraj mecz ponownie aby wygenerować.</p>'
        rows = ""
        for i, lu in enumerate(all_lineups):
            fga = int(lu.get("p2a",0) or 0) + int(lu.get("p3a",0) or 0)
            pts  = int(lu.get("pts",0) or 0)
            poss = int(lu.get("poss",0) or 0)
            p2m  = int(lu.get("p2m",0) or 0); p2a = int(lu.get("p2a",0) or 0)
            p3m  = int(lu.get("p3m",0) or 0); p3a = int(lu.get("p3a",0) or 0)
            ftm  = int(lu.get("ftm",0) or 0); fta = int(lu.get("fta",0) or 0)
            br   = int(lu.get("br",0) or 0)
            efg  = f"{(p2m+1.5*p3m)/fga:.0%}" if fga else "—"
            ppp  = f"{pts/poss:.2f}" if poss else "—"
            ppp_color = "#1a6b3c" if poss and pts/poss>=0.9 else ("#8b1a1a" if poss and pts/poss<0.7 else "#444")
            bg = "#f8f9ff" if i%2==0 else "#fff"
            # Skład — zamień numery na nazwiska jeśli dostępne
            skladniki = " · ".join(nr_name_map.get(n, f"#{n}") for n in lu["lineup"].split("-"))
            rows += f"""<tr style="background:{bg}">
                <td style="font-size:.78rem">{skladniki}</td>
                <td class="text-center">{poss}</td>
                <td class="text-center fw-bold" style="color:#1a2b4a">{pts}</td>
                <td class="text-center fw-bold" style="color:{ppp_color}">{ppp}</td>
                <td class="text-center">{efg}</td>
                <td class="text-center">{p2m}/{p2a}</td>
                <td class="text-center">{p3m}/{p3a}</td>
                <td class="text-center">{ftm}/{fta}</td>
                <td class="text-center">{br}</td>
            </tr>"""
        return f"""<p style="font-size:.72rem;color:#aaa;margin-bottom:.5rem">Posiadania ≥ 1 · PPP: <span style="color:#1a6b3c">≥0.90 dobry</span> / <span style="color:#8b1a1a">&lt;0.70 słaby</span></p>
        <div class="table-responsive"><table class="table table-hover mb-0">
        <thead><tr>
          <th>Skład</th><th class="text-center">POSS</th><th class="text-center">PKT</th>
          <th class="text-center">PPP</th><th class="text-center">eFG%</th>
          <th class="text-center">2PM/A</th><th class="text-center">3PM/A</th>
          <th class="text-center">FTM/A</th><th class="text-center">BR</th>
        </tr></thead>
        <tbody>{rows}</tbody></table></div>"""

    # Shot timing
    def tim_table(druzyna):
        rows = ""
        for b in BUCKETS:
            td = next((r for r in all_timing if r["druzyna"]==druzyna and r["bucket"]==b), {})
            m2=td.get("made2",0); a2=td.get("att2",0)
            m3=td.get("made3",0); a3=td.get("att3",0)
            tot=m2+m3; att=a2+a3
            eff=f"{tot/att:.0%}" if att else "-"
            rows += f"<tr><td style=\"background:#fff\"><b>{b}</b></td><td>{m2}/{a2}</td><td>{m3}/{a3}</td><td><b>{eff}</b></td></tr>"
        return f"""<div class="table-responsive"><table class="table table-hover mb-0">
            <thead><tr><th>Czas</th><th>2PT</th><th>3PT</th><th>Eff%</th></tr></thead>
            <tbody>{rows}</tbody></table></div>"""

    pts_q_gtk = [next((r["pts"] for r in all_stats if r["druzyna"]=="gtk" and r["kwarta"]==q),0) for q in [1,2,3,4]]
    pts_q_opp = [next((r["pts"] for r in all_stats if r["druzyna"]=="opp" and r["kwarta"]==q),0) for q in [1,2,3,4]]

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
  <div>
    <div class="page-title mb-0">{gtk_name} vs {name_opp}</div>
    <div style="font-size:.8rem;color:#888">{dt} · Sezon {m['sezon']}</div>
  </div>
  <div class="d-flex gap-2">
    <a href="/historia" class="btn btn-outline-secondary btn-sm">← Historia</a>
    <a href="/mecz/{match_id}/edytuj" class="btn btn-outline-primary btn-sm">✏️ Przypisz zawodników</a>
    <a href="/mecz/{match_id}/export/xlsx" class="btn btn-warning btn-sm fw-bold">⬇ Excel</a>
    <a href="/mecz/{match_id}/export/pdf" class="btn btn-danger btn-sm fw-bold">⬇ PDF</a>
  </div>
</div>

<div class="hero mb-3">
  <div class="d-flex justify-content-between align-items-center flex-wrap gap-3">
    <div>
      <div style="font-size:.8rem;opacity:.7">{gtk_name}</div>
      <div style="font-size:2rem;font-weight:700">{m['wynik_gtk']}</div>
    </div>
    <div class="text-center">
      {'<span class="badge-win" style="font-size:.9rem;padding:6px 14px">WYGRANA</span>' if m['wynik_gtk']>m['wynik_opp'] else '<span class="badge-loss" style="font-size:.9rem;padding:6px 14px">PRZEGRANA</span>' if m['wynik_gtk']<m['wynik_opp'] else '<span class="badge-draw" style="font-size:.9rem;padding:6px 14px">REMIS</span>'}
    </div>
    <div class="text-end">
      <div style="font-size:.8rem;opacity:.7">{name_opp}</div>
      <div style="font-size:2rem;font-weight:700">{m['wynik_opp']}</div>
    </div>
  </div>
</div>

<ul class="nav nav-tabs mb-2" id="mainTabs">
  <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#tabGTK">{gtk_name}</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabOPP">{name_opp}</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabCMP">Porównanie</button></li>
</ul>

<div class="tab-content">

<div class="tab-pane fade show active" id="tabGTK">
  <div class="row g-2 mb-2">{kpi_cards(suma_gtk, kpi_gtk)}</div>
  <ul class="nav nav-tabs mt-2 mb-1" id="gtkTabs">
    <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#gtk_q">Per kwarta</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#gtk_p">Zawodnicy</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#gtk_l">Piątki</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#gtk_t">Timing rzutów</button></li>
  </ul>
  <div class="tab-content">
    <div class="tab-pane fade show active" id="gtk_q"><div class="card mt-1"><div class="card-body p-2">{q_table('gtk')}</div></div></div>
    <div class="tab-pane fade" id="gtk_p"><div class="card mt-1"><div class="card-body p-2">{p_table('gtk')}</div></div></div>
    <div class="tab-pane fade" id="gtk_l"><div class="card mt-1"><div class="card-body p-2">{lineup_table()}</div></div></div>
    <div class="tab-pane fade" id="gtk_t"><div class="card mt-1"><div class="card-body p-2">{tim_table('gtk')}</div></div></div>
  </div>
</div>

<div class="tab-pane fade" id="tabOPP">
  <div class="row g-2 mb-2">{kpi_cards(suma_opp, kpi_opp)}</div>
  <ul class="nav nav-tabs mt-2 mb-1" id="oppTabs">
    <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#opp_q">Per kwarta</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#opp_p">Zawodnicy</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#opp_t">Timing rzutów</button></li>
  </ul>
  <div class="tab-content">
    <div class="tab-pane fade show active" id="opp_q"><div class="card mt-1"><div class="card-body p-2">{q_table('opp')}</div></div></div>
    <div class="tab-pane fade" id="opp_p"><div class="card mt-1"><div class="card-body p-2">{p_table('opp')}</div></div></div>
    <div class="tab-pane fade" id="opp_t"><div class="card mt-1"><div class="card-body p-2">{tim_table('opp')}</div></div></div>
  </div>
</div>

<div class="tab-pane fade" id="tabCMP">
  <ul class="nav nav-tabs mt-2 mb-1" id="cmpTabs">
    <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#cmp_metrics">Kluczowe Metryki</button></li>
    <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#cmp_timing">Timing rzutów</button></li>
  </ul>
  <div class="tab-content">

  <div class="tab-pane fade show active" id="cmp_metrics">
    <div class="row g-3 mt-1">
      <div class="col-lg-5">
        <div class="card"><div class="card-body p-2">
          <p style="font-size:.72rem;color:#aaa;margin-bottom:.5rem">Kliknij metrykę aby zobaczyć breakdown per kwarta →</p>
          <table class="table table-sm mb-0" id="metricsTable">
            <thead><tr>
              <th>Metryka</th>
              <th class="text-center gtk-color">{gtk_name}</th>
              <th class="text-center opp-color">{name_opp}</th>
            </tr></thead>
            <tbody>
              {''.join(f'''<tr class="metric-row" data-metric="{key}" style="cursor:pointer;transition:.15s" title="Kliknij aby zobaczyć per kwarta">
                <td style="font-size:.82rem">
                  <b>{l}</b><br>
                  <span style="font-size:.7rem;color:#aaa">{desc}</span>
                  <div style="display:flex;height:4px;border-radius:3px;overflow:hidden;margin-top:4px">
                    <div style="flex:{fa};background:{'#1a6b3c' if fa>=fb else '#ddd'}"></div>
                    <div style="flex:{fb};background:{'#8b1a1a' if fb>fa else '#ddd'}"></div>
                  </div>
                </td>
                <td class="text-center fw-bold" style="{'color:#1a6b3c' if fa>fb else 'color:#555'}">{va}</td>
                <td class="text-center fw-bold" style="{'color:#8b1a1a' if fb>fa else 'color:#555'}">{vb}</td>
              </tr>''' for l,va,vb,desc,key,fa,fb in [
                ("Punkty",suma_gtk.get('pts',0),suma_opp.get('pts',0),"Łączna liczba punktów","pts",suma_gtk.get('pts',0),suma_opp.get('pts',0)),
                ("Posiadania",suma_gtk.get('poss',0),suma_opp.get('poss',0),"Liczba posiadań","poss",suma_gtk.get('poss',0),suma_opp.get('poss',0)),
                ("eFG%",kpi_gtk['efg'],kpi_opp['efg'],"Efektywny % rzutów z pola","efg",
                 float(kpi_gtk['efg'].replace('%','')) if kpi_gtk['efg']!='-' else 0,
                 float(kpi_opp['efg'].replace('%','')) if kpi_opp['efg']!='-' else 0),
                ("TS%",kpi_gtk['ts'],kpi_opp['ts'],"Prawdziwy % skuteczności","ts",
                 float(kpi_gtk['ts'].replace('%','')) if kpi_gtk['ts']!='-' else 0,
                 float(kpi_opp['ts'].replace('%','')) if kpi_opp['ts']!='-' else 0),
                ("ORtg",kpi_gtk['ortg'],kpi_opp['ortg'],"Punkty na 100 posiadań","ortg",
                 float(kpi_gtk['ortg']) if kpi_gtk['ortg']!='-' else 0,
                 float(kpi_opp['ortg']) if kpi_opp['ortg']!='-' else 0),
                ("PPP",kpi_gtk['ppp'],kpi_opp['ppp'],"Punkty na posiadanie","ppp",
                 float(kpi_gtk['ppp']) if kpi_gtk['ppp']!='-' else 0,
                 float(kpi_opp['ppp']) if kpi_opp['ppp']!='-' else 0),
                ("2PT%",kpi_gtk['p2_pct'],kpi_opp['p2_pct'],"Skuteczność za 2 pkt","p2_pct",
                 float(kpi_gtk['p2_pct'].replace('%','')) if kpi_gtk['p2_pct']!='-' else 0,
                 float(kpi_opp['p2_pct'].replace('%','')) if kpi_opp['p2_pct']!='-' else 0),
                ("3PT%",kpi_gtk['p3_pct'],kpi_opp['p3_pct'],"Skuteczność za 3 pkt","p3_pct",
                 float(kpi_gtk['p3_pct'].replace('%','')) if kpi_gtk['p3_pct']!='-' else 0,
                 float(kpi_opp['p3_pct'].replace('%','')) if kpi_opp['p3_pct']!='-' else 0),
                ("FT%",kpi_gtk['ft_pct'],kpi_opp['ft_pct'],"Skuteczność rzutów wolnych","ft_pct",
                 float(kpi_gtk['ft_pct'].replace('%','')) if kpi_gtk['ft_pct']!='-' else 0,
                 float(kpi_opp['ft_pct'].replace('%','')) if kpi_opp['ft_pct']!='-' else 0),
                ("Straty (BR)",suma_gtk.get('br',0),suma_opp.get('br',0),"Liczba strat — niższy = lepszy","br",
                 suma_opp.get('br',0),suma_gtk.get('br',0)),
                ("Faule wymuszone",suma_gtk.get('fd',0),suma_opp.get('fd',0),"Liczba wymuszonych fauli","fd",
                 suma_gtk.get('fd',0),suma_opp.get('fd',0)),
                ("2PM/A",f"{suma_gtk.get('p2m',0)}/{suma_gtk.get('p2a',0)}",f"{suma_opp.get('p2m',0)}/{suma_opp.get('p2a',0)}","Celne / próby za 2 pkt","p2m",suma_gtk.get('p2m',0),suma_opp.get('p2m',0)),
                ("3PM/A",f"{suma_gtk.get('p3m',0)}/{suma_gtk.get('p3a',0)}",f"{suma_opp.get('p3m',0)}/{suma_opp.get('p3a',0)}","Celne / próby za 3 pkt","p3m",suma_gtk.get('p3m',0),suma_opp.get('p3m',0)),
                ("FTM/A",f"{suma_gtk.get('ftm',0)}/{suma_gtk.get('fta',0)}",f"{suma_opp.get('ftm',0)}/{suma_opp.get('fta',0)}","Celne / próby rzutów wolnych","ftm",suma_gtk.get('ftm',0),suma_opp.get('ftm',0)),
              ])}
            </tbody>
          </table>
        </div></div>
      </div>

      <div class="col-lg-7">
        <div class="card" id="detailCard"><div class="card-body p-3">
          <div id="detailTitle" class="section-hdr">Punkty per kwarta</div>
          <canvas id="qChart"></canvas>
          <div id="qBreakdown" style="display:none">
            <canvas id="qDetailChart"></canvas>
            <div id="qDetailTable" class="mt-2"></div>
          </div>
        </div></div>
      </div>
    </div>
  </div>

  <div class="tab-pane fade" id="cmp_timing">
    <div class="card mt-2"><div class="card-body p-2">
      <p class="text-muted mb-2" style="font-size:.8rem">Porównanie skuteczności rzutów według czasu trwania posiadania (zegar 24s)</p>
      <div class="d-flex gap-3 mb-2" style="font-size:.78rem">
        <span><span style="display:inline-block;width:12px;height:8px;background:#1a6b3c;border-radius:2px;margin-right:4px"></span>{gtk_name}</span>
        <span><span style="display:inline-block;width:12px;height:8px;background:#8b1a1a;border-radius:2px;margin-right:4px"></span>{name_opp}</span>
      </div>
      <div class="table-responsive">
        <table class="table table-hover mb-0">
          <thead>
            <tr>
              <th rowspan="2" style="background:#f8f9fa;color:#1a2b4a;border-bottom:2px solid #dee2e6;vertical-align:middle">Czas</th>
              <th colspan="3" style="background:#e8f5e9;color:#1a6b3c;text-align:center;border-bottom:2px solid #1a6b3c">{gtk_name}</th>
              <th colspan="2" style="background:#f8f9fa;color:#555;text-align:center;border-bottom:2px solid #dee2e6"></th>
              <th colspan="3" style="background:#ffebee;color:#8b1a1a;text-align:center;border-bottom:2px solid #8b1a1a">{name_opp}</th>
            </tr>
            <tr>
              <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">Celne/Próby</th>
              <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">Eff%</th>
              <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">2PT | 3PT</th>
              <th style="background:#f8f9fa;color:#555;text-align:center;font-size:.72rem;width:90px">{gtk_name}</th>
              <th style="background:#f8f9fa;color:#555;text-align:center;font-size:.72rem;width:90px">{name_opp}</th>
              <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">2PT | 3PT</th>
              <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">Eff%</th>
              <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">Celne/Próby</th>
            </tr>
          </thead>
          <tbody>
            {''.join(f"""<tr style="background:{'#f9f9f9' if i%2==0 else '#fff'}">
              <td class="fw-bold" style="font-size:.82rem;background:#fff">{b}</td>
              <td class="text-center" style="font-size:.82rem;background:#f0fff4">{(lambda gd: f"{gd.get('made2',0)+gd.get('made3',0)}/{gd.get('att2',0)+gd.get('att3',0)}")(next((r for r in all_timing if r['druzyna']=='gtk' and r['bucket']==b),{}))}</td>
              <td class="text-center" style="font-size:.85rem;font-weight:700;color:#1a6b3c;background:#f0fff4">{(lambda gd: f"{(gd.get('made2',0)+gd.get('made3',0))/(gd.get('att2',0)+gd.get('att3',0)):.0%}" if (gd.get('att2',0)+gd.get('att3',0)) else "—")(next((r for r in all_timing if r['druzyna']=='gtk' and r['bucket']==b),{}))}</td>
              <td style="font-size:.75rem;color:#555;background:#f0fff4;text-align:center">{(lambda gd: f"{gd.get('made2',0)}/{gd.get('att2',0)} | {gd.get('made3',0)}/{gd.get('att3',0)}")(next((r for r in all_timing if r['druzyna']=='gtk' and r['bucket']==b),{}))}</td>
              <td style="padding:6px 8px"><div style="height:8px;width:{int((next((r for r in all_timing if r['druzyna']=='gtk' and r['bucket']==b),{}).get('att2',0)+next((r for r in all_timing if r['druzyna']=='gtk' and r['bucket']==b),{}).get('att3',0))/max(max((next((r for r in all_timing if r['druzyna']=='gtk' and r['bucket']==bb),{}).get('att2',0)+next((r for r in all_timing if r['druzyna']=='gtk' and r['bucket']==bb),{}).get('att3',0)) for bb in BUCKETS),1)*80)}px;background:#1a6b3c;border-radius:4px"></div></td>
              <td style="padding:6px 8px"><div style="height:8px;width:{int((next((r for r in all_timing if r['druzyna']=='opp' and r['bucket']==b),{}).get('att2',0)+next((r for r in all_timing if r['druzyna']=='opp' and r['bucket']==b),{}).get('att3',0))/max(max((next((r for r in all_timing if r['druzyna']=='opp' and r['bucket']==bb),{}).get('att2',0)+next((r for r in all_timing if r['druzyna']=='opp' and r['bucket']==bb),{}).get('att3',0)) for bb in BUCKETS),1)*80)}px;background:#8b1a1a;border-radius:4px"></div></td>
              <td style="font-size:.75rem;color:#555;background:#fff5f5;text-align:center">{(lambda od: f"{od.get('made2',0)}/{od.get('att2',0)} | {od.get('made3',0)}/{od.get('att3',0)}")(next((r for r in all_timing if r['druzyna']=='opp' and r['bucket']==b),{}))}</td>
              <td class="text-center" style="font-size:.85rem;font-weight:700;color:#8b1a1a;background:#fff5f5">{(lambda od: f"{(od.get('made2',0)+od.get('made3',0))/(od.get('att2',0)+od.get('att3',0)):.0%}" if (od.get('att2',0)+od.get('att3',0)) else "—")(next((r for r in all_timing if r['druzyna']=='opp' and r['bucket']==b),{}))}</td>
              <td class="text-center" style="font-size:.82rem;background:#fff5f5">{(lambda od: f"{od.get('made2',0)+od.get('made3',0)}/{od.get('att2',0)+od.get('att3',0)}")(next((r for r in all_timing if r['druzyna']=='opp' and r['bucket']==b),{}))}</td>
            </tr>""" for i,b in enumerate(BUCKETS))}
          </tbody>
        </table>
      </div>
    </div></div>
  </div>

  </div>
</div>

</div>"""

    scripts = f"""<script>
// ── Dane per kwarta ─────────────────────────────────────────────────────────
const qData = {{
  gtk: {{
    pts:  {[next((r['pts']  for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    poss: {[next((r['poss'] for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    p2m:  {[next((r['p2m']  for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    p2a:  {[next((r['p2a']  for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    p3m:  {[next((r['p3m']  for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    p3a:  {[next((r['p3a']  for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    ftm:  {[next((r['ftm']  for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    fta:  {[next((r['fta']  for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    br:   {[next((r['br']   for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    fd:   {[next((r['fd']   for r in all_stats if r['druzyna']=='gtk' and r['kwarta']==q),0) for q in [1,2,3,4]]},
  }},
  opp: {{
    pts:  {[next((r['pts']  for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    poss: {[next((r['poss'] for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    p2m:  {[next((r['p2m']  for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    p2a:  {[next((r['p2a']  for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    p3m:  {[next((r['p3m']  for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    p3a:  {[next((r['p3a']  for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    ftm:  {[next((r['ftm']  for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    fta:  {[next((r['fta']  for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    br:   {[next((r['br']   for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
    fd:   {[next((r['fd']   for r in all_stats if r['druzyna']=='opp' and r['kwarta']==q),0) for q in [1,2,3,4]]},
  }}
}};

const gtkName = '{gtk_name}';
const oppName = '{name_opp}';

// ── Funkcja obliczająca wartość metryki per kwarta ──────────────────────────
function getMetricValues(key) {{
  const d = qData;
  return [0,1,2,3].map(i => {{
    const g = d.gtk, o = d.opp;
    const pct = (n,dv) => dv>0 ? Math.round(n/dv*1000)/10 : null;
    switch(key) {{
      case 'pts':   return [g.pts[i],   o.pts[i]];
      case 'poss':  return [g.poss[i],  o.poss[i]];
      case 'p2m':   return [g.p2m[i],   o.p2m[i]];
      case 'p3m':   return [g.p3m[i],   o.p3m[i]];
      case 'ftm':   return [g.ftm[i],   o.ftm[i]];
      case 'br':    return [g.br[i],    o.br[i]];
      case 'fd':    return [g.fd[i],    o.fd[i]];
      case 'efg':   return [
        pct(g.p2m[i]+1.5*g.p3m[i], g.p2a[i]+g.p3a[i]),
        pct(o.p2m[i]+1.5*o.p3m[i], o.p2a[i]+o.p3a[i])
      ];
      case 'ts':    return [
        pct(g.pts[i], 2*(g.p2a[i]+g.p3a[i]+0.44*g.fta[i])),
        pct(o.pts[i], 2*(o.p2a[i]+o.p3a[i]+0.44*o.fta[i]))
      ];
      case 'ortg':  return [
        g.poss[i]>0 ? Math.round(g.pts[i]*100/g.poss[i]*10)/10 : null,
        o.poss[i]>0 ? Math.round(o.pts[i]*100/o.poss[i]*10)/10 : null
      ];
      case 'ppp':   return [
        g.poss[i]>0 ? Math.round(g.pts[i]/g.poss[i]*100)/100 : null,
        o.poss[i]>0 ? Math.round(o.pts[i]/o.poss[i]*100)/100 : null
      ];
      case 'p2_pct': return [pct(g.p2m[i],g.p2a[i]), pct(o.p2m[i],o.p2a[i])];
      case 'p3_pct': return [pct(g.p3m[i],g.p3a[i]), pct(o.p3m[i],o.p3a[i])];
      case 'ft_pct': return [pct(g.ftm[i],g.fta[i]), pct(o.ftm[i],o.fta[i])];
      default: return [0,0];
    }}
  }});
}}

// ── Domyślny wykres punktów ─────────────────────────────────────────────────
let activeChart = null;

function showDefaultChart() {{
  if(activeChart) {{ activeChart.destroy(); activeChart=null; }}
  document.getElementById('qChart').style.display = 'block';
  document.getElementById('qBreakdown').style.display = 'none';
  document.getElementById('detailTitle').textContent = 'Punkty per kwarta';
  activeChart = new Chart(document.getElementById('qChart'),{{
    type:'bar',
    data:{{labels:['1Q','2Q','3Q','4Q'],datasets:[
      {{label:gtkName, data:qData.gtk.pts, backgroundColor:'#1a6b3c88', borderColor:'#1a6b3c', borderWidth:2, borderRadius:6}},
      {{label:oppName, data:qData.opp.pts, backgroundColor:'#8b1a1a88', borderColor:'#8b1a1a', borderWidth:2, borderRadius:6}}
    ]}},
    options:{{responsive:true,plugins:{{legend:{{position:'top'}}}},scales:{{y:{{beginAtZero:true}}}}}}
  }});
}}
showDefaultChart();

// ── Kliknięcie metryki ──────────────────────────────────────────────────────
let detailChart = null;
let selectedRow = null;

document.querySelectorAll('.metric-row').forEach(row => {{
  row.addEventListener('click', function() {{
    const key = this.dataset.metric;
    const label = this.querySelector('b').textContent;

    // Podświetl wybrany wiersz
    document.querySelectorAll('.metric-row').forEach(r => r.style.background='');
    this.style.background = '#f0f4ff';
    selectedRow = this;

    const vals = getMetricValues(key);
    const gtkVals = vals.map(v => v[0]);
    const oppVals = vals.map(v => v[1]);

    // Ukryj główny wykres, pokaż breakdown
    document.getElementById('qChart').style.display = 'none';
    document.getElementById('qBreakdown').style.display = 'block';
    document.getElementById('detailTitle').textContent = label + ' — per kwarta';

    // Zniszcz stary wykres
    if(detailChart) {{ detailChart.destroy(); detailChart=null; }}
    if(activeChart) {{ activeChart.destroy(); activeChart=null; }}

    // Nowy wykres
    detailChart = new Chart(document.getElementById('qDetailChart'),{{
      type:'bar',
      data:{{
        labels:['1Q','2Q','3Q','4Q'],
        datasets:[
          {{label:gtkName, data:gtkVals, backgroundColor:'#1a6b3c88', borderColor:'#1a6b3c', borderWidth:2, borderRadius:6}},
          {{label:oppName, data:oppVals, backgroundColor:'#8b1a1a88', borderColor:'#8b1a1a', borderWidth:2, borderRadius:6}}
        ]
      }},
      options:{{
        responsive:true,
        plugins:{{
          legend:{{position:'top'}},
          tooltip:{{callbacks:{{label: ctx => ctx.dataset.label+': '+(ctx.raw!==null?ctx.raw:'-')}}}}
        }},
        scales:{{y:{{beginAtZero:true}}}}
      }}
    }});

    // Mini tabela pod wykresem
    const isPercent = ['efg','ts','p2_pct','p3_pct','ft_pct'].includes(key);
    let tableHtml = '<table class="table table-sm mb-0" style="font-size:.8rem"><thead><tr><th>Kwarta</th>';
    tableHtml += `<th class="text-center" style="color:#1a6b3c">${{gtkName}}</th>`;
    tableHtml += `<th class="text-center" style="color:#8b1a1a">${{oppName}}</th>`;
    tableHtml += '<th class="text-center" style="color:#888">Różnica</th></tr></thead><tbody>';
    const labels = ['1Q','2Q','3Q','4Q'];
    for(let i=0;i<4;i++) {{
      const gv = gtkVals[i]; const ov = oppVals[i];
      const fmt = v => v===null ? '-' : (isPercent ? v.toFixed(1)+'%' : v);
      let diff = '-'; let diffColor='#888';
      if(gv!==null && ov!==null) {{
        const d = Math.round((gv-ov)*100)/100;
        diff = (d>0?'+':'')+fmt(d);
        diffColor = d>0 ? '#1a6b3c' : (d<0 ? '#8b1a1a' : '#888');
      }}
      const bgG = gv!==null && ov!==null && gv>ov ? 'background:#f0fff4' : '';
      const bgO = gv!==null && ov!==null && ov>gv ? 'background:#fff0f0' : '';
      tableHtml += `<tr><td class="fw-bold">${{labels[i]}}</td>`;
      tableHtml += `<td class="text-center fw-bold" style="${{bgG}};color:#1a6b3c">${{fmt(gv)}}</td>`;
      tableHtml += `<td class="text-center fw-bold" style="${{bgO}};color:#8b1a1a">${{fmt(ov)}}</td>`;
      tableHtml += `<td class="text-center" style="color:${{diffColor}}">${{diff}}</td></tr>`;
    }}
    tableHtml += '</tbody></table>';
    document.getElementById('qDetailTable').innerHTML = tableHtml;
  }});
}});

// Kliknięcie poza tabelą → wróć do domyślnego wykresu
document.addEventListener('click', function(e) {{
  if(!e.target.closest('#metricsTable') && !e.target.closest('#detailCard')) {{
    if(selectedRow) {{
      selectedRow.style.background='';
      selectedRow=null;
      if(detailChart){{ detailChart.destroy(); detailChart=null; }}
      showDefaultChart();
    }}
  }}
}});
</script>"""

    return render_template_string(base(content, scripts, active="history"))

@app.route("/mecz/<int:match_id>/delete")
def mecz_delete(match_id):
    db = get_db(); cur = db.cursor()
    cur.execute("DELETE FROM matches WHERE id=%s", (match_id,))
    db.commit(); cur.close()
    flash("Mecz usunięty","success")
    return redirect(url_for("historia"))

# ══════════════════════════════════════════════════════════════════════════════
# STATYSTYKI SEZONU
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/sezon")
def sezon():
    sezon_filter = request.args.get("sezon", get_setting("current_season") or "2024/25")
    db = get_db(); cur = db.cursor()

    cur.execute("SELECT DISTINCT sezon FROM matches ORDER BY sezon DESC")
    sezony = [r["sezon"] for r in cur.fetchall()]

    cur.execute("SELECT COUNT(*) as cnt FROM matches WHERE sezon=%s", (sezon_filter,))
    n_matches = cur.fetchone()["cnt"]

    if n_matches == 0:
        cur.close()
        content = f"""
<div class="page-title">📊 Statystyki sezonu {sezon_filter}</div>
<div class="card p-4 text-center text-muted">Brak meczów w sezonie {sezon_filter}. Wgraj pliki zapis.xlsx.</div>"""
        return render_template_string(base(content, active="season"))

    # Agregaty GTK
    cur.execute("""
        SELECT
            SUM(pts) as pts, SUM(poss) as poss,
            SUM(p2m) as p2m, SUM(p2a) as p2a,
            SUM(p3m) as p3m, SUM(p3a) as p3a,
            SUM(ftm) as ftm, SUM(fta) as fta,
            SUM(br)  as br,  SUM(fd)  as fd
        FROM match_stats ms
        JOIN matches m ON ms.match_id = m.id
        WHERE m.sezon=%s AND ms.druzyna='gtk'
    """, (sezon_filter,))
    gtk_tot = dict(cur.fetchone())

    # Agregaty OPP
    cur.execute("""
        SELECT
            SUM(pts) as pts, SUM(poss) as poss,
            SUM(p2m) as p2m, SUM(p2a) as p2a,
            SUM(p3m) as p3m, SUM(p3a) as p3a,
            SUM(ftm) as ftm, SUM(fta) as fta,
            SUM(br)  as br,  SUM(fd)  as fd
        FROM match_stats ms
        JOIN matches m ON ms.match_id = m.id
        WHERE m.sezon=%s AND ms.druzyna='opp'
    """, (sezon_filter,))
    opp_tot = dict(cur.fetchone())

    # Wyniki meczów
    cur.execute("""SELECT wynik_gtk, wynik_opp FROM matches WHERE sezon=%s""", (sezon_filter,))
    results = cur.fetchall()
    wins   = sum(1 for r in results if r["wynik_gtk"] > r["wynik_opp"])
    losses = sum(1 for r in results if r["wynik_gtk"] < r["wynik_opp"])
    draws  = n_matches - wins - losses

    # Shot timing per sezon
    cur.execute("""
        SELECT ts.druzyna, ts.bucket,
               SUM(ts.made2) as made2, SUM(ts.att2) as att2,
               SUM(ts.made3) as made3, SUM(ts.att3) as att3
        FROM timing_stats ts
        JOIN matches m ON ts.match_id=m.id
        WHERE m.sezon=%s
        GROUP BY ts.druzyna, ts.bucket
    """, (sezon_filter,))
    timing_rows = cur.fetchall()
    cur.close()

    gtk_name = get_setting("gtk_name") or "GTK"

    def avg(d, k):
        v = d.get(k) or 0
        return round(v / n_matches, 1)

    def avg_kpi(d):
        k = calc_kpi(d)
        return k

    gtk_kpi = calc_kpi(gtk_tot)
    opp_kpi = calc_kpi(opp_tot)

    def cmp_row(lbl, vg, vo, higher_is_better=True):
        try:
            fg = float(str(vg).replace('%','').replace('-','0'))
            fo = float(str(vo).replace('%','').replace('-','0'))
            sg = "font-weight:700;color:#1a6b3c" if (higher_is_better and fg>fo) or (not higher_is_better and fg<fo) else ""
            so = "font-weight:700;color:#8b1a1a" if (higher_is_better and fo>fg) or (not higher_is_better and fo<fg) else ""
        except: sg=so=""
        return f"<tr><td><b>{lbl}</b></td><td class='text-center' style='{sg}'>{vg}</td><td class='text-center' style='{so}'>{vo}</td></tr>"

    kpi_rows = (
        cmp_row("Pkt / mecz",       avg(gtk_tot,"pts"),  avg(opp_tot,"pts")) +
        cmp_row("Posiadania / mecz", avg(gtk_tot,"poss"), avg(opp_tot,"poss")) +
        cmp_row("eFG%",             gtk_kpi["efg"],      opp_kpi["efg"]) +
        cmp_row("TS%",              gtk_kpi["ts"],       opp_kpi["ts"]) +
        cmp_row("ORtg",             gtk_kpi["ortg"],     opp_kpi["ortg"]) +
        cmp_row("PPP",              gtk_kpi["ppp"],      opp_kpi["ppp"]) +
        cmp_row("2PT%",             gtk_kpi["p2_pct"],   opp_kpi["p2_pct"]) +
        cmp_row("3PT%",             gtk_kpi["p3_pct"],   opp_kpi["p3_pct"]) +
        cmp_row("FT%",              gtk_kpi["ft_pct"],   opp_kpi["ft_pct"]) +
        cmp_row("Straty / mecz",    avg(gtk_tot,"br"),   avg(opp_tot,"br"), False) +
        cmp_row("FT Rate",          gtk_kpi["ftr"],      opp_kpi["ftr"])
    )

    # Timing tabela
    def timing_row(bucket):
        gd = next((r for r in timing_rows if r["druzyna"]=="gtk" and r["bucket"]==bucket), {})
        od = next((r for r in timing_rows if r["druzyna"]=="opp" and r["bucket"]==bucket), {})
        gm2=int(gd.get("made2",0) or 0); ga2=int(gd.get("att2",0) or 0)
        gm3=int(gd.get("made3",0) or 0); ga3=int(gd.get("att3",0) or 0)
        om2=int(od.get("made2",0) or 0); oa2=int(od.get("att2",0) or 0)
        om3=int(od.get("made3",0) or 0); oa3=int(od.get("att3",0) or 0)
        gm=gm2+gm3; ga=ga2+ga3
        om=om2+om3; oa=oa2+oa3
        ge=f"{gm/ga:.0%}" if ga else "—"
        oe=f"{om/oa:.0%}" if oa else "—"
        # Paski
        max_att = max(
            max((int(r.get("att2",0) or 0)+int(r.get("att3",0) or 0)) for r in timing_rows if r["druzyna"]=="gtk") if timing_rows else 1,
            1
        )
        gbar = int(ga/max_att*80) if ga else 0
        obar = int(oa/max_att*80) if oa else 0
        return f"""<tr>
            <td class="fw-bold" style="font-size:.85rem;background:#fff">{bucket}</td>
            <td style="font-size:.8rem;background:#f0fff4;text-align:center">{gm}/{ga}</td>
            <td style="font-weight:700;color:#1a6b3c;background:#f0fff4;text-align:center">{ge}</td>
            <td style="font-size:.75rem;color:#555;background:#f0fff4;text-align:center">{gm2}/{ga2} | {gm3}/{ga3}</td>
            <td style="padding:6px 8px;background:#fff"><div style="height:8px;width:{gbar}px;background:#1a6b3c;border-radius:4px"></div></td>
            <td style="padding:6px 8px;background:#fff"><div style="height:8px;width:{obar}px;background:#8b1a1a;border-radius:4px"></div></td>
            <td style="font-size:.75rem;color:#555;background:#fff5f5;text-align:center">{om2}/{oa2} | {om3}/{oa3}</td>
            <td style="font-weight:700;color:#8b1a1a;background:#fff5f5;text-align:center">{oe}</td>
            <td style="font-size:.8rem;background:#fff5f5;text-align:center">{om}/{oa}</td>
        </tr>"""

    tim_rows = "".join(timing_row(b) for b in BUCKETS)

    pts_per_match_gtk = [0,0,0,0]
    pts_per_match_opp = [0,0,0,0]
    season_opts = "".join([f'<option value="{s}" {"selected" if s==sezon_filter else ""}>{s}</option>' for s in sezony])

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
  <div class="page-title mb-0">📊 Statystyki drużyny</div>
  <form method="GET" class="d-flex gap-2 align-items-center">
    <label style="font-size:.82rem;font-weight:600">Sezon:</label>
    <select name="sezon" class="form-select form-select-sm" style="width:120px" onchange="this.form.submit()">
      {season_opts}
    </select>
  </form>
</div>

<div class="row g-2 mb-3">
  <div class="col-6 col-md-3"><div class="stat-card"><div class="stat-val">{n_matches}</div><div class="stat-lbl">Mecze</div></div></div>
  <div class="col-6 col-md-3"><div class="stat-card"><div class="stat-val" style="color:#1a6b3c">{wins}</div><div class="stat-lbl">Wygrane</div></div></div>
  <div class="col-6 col-md-3"><div class="stat-card"><div class="stat-val" style="color:#8b1a1a">{losses}</div><div class="stat-lbl">Przegrane</div></div></div>
  <div class="col-6 col-md-3"><div class="stat-card"><div class="stat-val">{wins/n_matches:.0%}</div><div class="stat-lbl">Win%</div></div></div>
</div>

<ul class="nav nav-tabs mb-2">
  <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#sMetrics">Metryki średnie</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#sTiming">Timing rzutów</button></li>
</ul>

<div class="tab-content">

<div class="tab-pane fade show active" id="sMetrics">
  <div class="row g-3">
    <div class="col-lg-6">
      <div class="card"><div class="card-body p-2">
        <div class="section-hdr">Średnie na mecz — {gtk_name} vs Przeciwnicy</div>
        <p style="font-size:.75rem;color:#aaa">Suma wszystkich danych ÷ liczba meczów ({n_matches})</p>
        <table class="table table-sm table-hover mb-0">
          <thead><tr>
            <th>Metryka</th>
            <th class="text-center" style="color:#1a6b3c">{gtk_name}</th>
            <th class="text-center" style="color:#8b1a1a">Przeciwnicy</th>
          </tr></thead>
          <tbody>{kpi_rows}</tbody>
        </table>
      </div></div>
    </div>
    <div class="col-lg-6">
      <div class="card"><div class="card-body p-3">
        <div class="section-hdr">Bilans meczów</div>
        <canvas id="winChart" style="max-height:200px"></canvas>
      </div></div>
    </div>
  </div>
</div>

<div class="tab-pane fade" id="sTiming">
  <div class="card mt-1"><div class="card-body p-2">
    <div class="section-hdr">Timing rzutów — skuteczność według czasu posiadania (zegar 24s)</div>
    <div class="d-flex gap-3 mb-2" style="font-size:.78rem">
      <span><span style="display:inline-block;width:12px;height:8px;background:#1a6b3c;border-radius:2px;margin-right:4px"></span>{gtk_name}</span>
      <span><span style="display:inline-block;width:12px;height:8px;background:#8b1a1a;border-radius:2px;margin-right:4px"></span>Przeciwnicy</span>
    </div>
    <div class="table-responsive">
    <table class="table table-hover mb-0">
      <thead>
        <tr>
          <th rowspan="2" style="vertical-align:middle;background:#1a2b4a;color:#fff">Czas</th>
          <th colspan="3" style="background:#e8f5e9;color:#1a6b3c;text-align:center;border-bottom:2px solid #1a6b3c">{gtk_name}</th>
          <th colspan="2" style="background:#fff;border-bottom:2px solid #dee2e6"></th>
          <th colspan="3" style="background:#ffebee;color:#8b1a1a;text-align:center;border-bottom:2px solid #8b1a1a">Przeciwnicy</th>
        </tr>
        <tr>
          <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">Celne/Próby</th>
          <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">Eff%</th>
          <th style="background:#e8f5e9;color:#1a6b3c;text-align:center;font-size:.72rem">2PT | 3PT</th>
          <th style="background:#fff;text-align:center;font-size:.72rem;width:90px;color:#555">{gtk_name}</th>
          <th style="background:#fff;text-align:center;font-size:.72rem;width:90px;color:#555">Przeciwnicy</th>
          <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">2PT | 3PT</th>
          <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">Eff%</th>
          <th style="background:#ffebee;color:#8b1a1a;text-align:center;font-size:.72rem">Celne/Próby</th>
        </tr>
      </thead>
      <tbody>{tim_rows}</tbody>
    </table>
    </div>
  </div></div>
</div>

</div>"""

    scripts = f"""<script>
new Chart(document.getElementById('winChart'),{{
  type:'doughnut',
  data:{{
    labels:['Wygrane','Przegrane'],
    datasets:[{{data:[{wins},{losses}],
      backgroundColor:['#1a6b3c','#8b1a1a'],
      borderWidth:0}}]
  }},
  options:{{responsive:true,plugins:{{legend:{{position:'bottom'}}}}}}
}});
</script>"""

    return render_template_string(base(content, scripts, active="season"))

# ══════════════════════════════════════════════════════════════════════════════
# ZAWODNICY SEZONU
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/zawodnicy")
def zawodnicy():
    sezon_filter = request.args.get("sezon", get_setting("current_season") or "2024/25")
    db = get_db(); cur = db.cursor()

    cur.execute("SELECT DISTINCT sezon FROM matches ORDER BY sezon DESC")
    sezony = [r["sezon"] for r in cur.fetchall()]

    cur.execute("SELECT COUNT(*) as cnt FROM matches WHERE sezon=%s", (sezon_filter,))
    n_matches = cur.fetchone()["cnt"]

    # Agreguj TYLKO po roster_id — nieprzypisani osobno per numer
    try:
        cur.execute("""
            SELECT
                grp_id, nazwa,
                SUM(pts) as pts, SUM(p2m) as p2m, SUM(p2a) as p2a,
                SUM(p3m) as p3m, SUM(p3a) as p3a,
                SUM(ftm) as ftm, SUM(fta) as fta,
                SUM(ast) as ast, SUM(oreb) as oreb, SUM(dreb) as dreb,
                SUM(br) as br, SUM(fd) as fd, SUM(finishes) as finishes,
                COUNT(DISTINCT match_id) as mecze,
                BOOL_OR(ma_nieprzypisane) as ma_nieprzypisane
            FROM (
                SELECT
                    CASE WHEN r.id IS NOT NULL THEN r.id::text
                         ELSE 'nr_'||ps.nr::text END as grp_id,
                    CASE WHEN r.id IS NOT NULL
                         THEN r.nazwisko || ' ' || r.imie
                         ELSE '— nieprzypisany #' || ps.nr::text
                    END as nazwa,
                    ps.match_id,
                    SUM(ps.pts) as pts, SUM(ps.p2m) as p2m, SUM(ps.p2a) as p2a,
                    SUM(ps.p3m) as p3m, SUM(ps.p3a) as p3a,
                    SUM(ps.ftm) as ftm, SUM(ps.fta) as fta,
                    SUM(ps.ast) as ast, SUM(ps.oreb) as oreb, SUM(ps.dreb) as dreb,
                    SUM(ps.br) as br, SUM(ps.fd) as fd, SUM(ps.finishes) as finishes,
                    (r.id IS NULL) as ma_nieprzypisane
                FROM player_stats ps
                JOIN matches m ON ps.match_id=m.id
                LEFT JOIN roster r ON ps.roster_id=r.id
                WHERE m.sezon=%s AND ps.druzyna='gtk'
                GROUP BY r.id, r.imie, r.nazwisko, ps.nr, ps.match_id
            ) sub
            GROUP BY grp_id, nazwa
            ORDER BY BOOL_OR(ma_nieprzypisane) ASC, SUM(pts) DESC
        """, (sezon_filter,))
    except Exception:
        try: get_db().rollback()
        except: pass
        cur = get_db().cursor()
        cur.execute("""
            SELECT ps.nr::text as grp_id,
                   '— nieprzypisany #'||ps.nr::text as nazwa,
                   SUM(ps.pts) as pts, SUM(ps.p2m) as p2m, SUM(ps.p2a) as p2a,
                   SUM(ps.p3m) as p3m, SUM(ps.p3a) as p3a,
                   SUM(ps.ftm) as ftm, SUM(ps.fta) as fta,
                   SUM(ps.ast) as ast, SUM(ps.oreb) as oreb, SUM(ps.dreb) as dreb,
                   SUM(ps.br) as br, SUM(ps.fd) as fd, SUM(ps.finishes) as finishes,
                   COUNT(DISTINCT ps.match_id) as mecze,
                   TRUE as ma_nieprzypisane
            FROM player_stats ps
            JOIN matches m ON ps.match_id=m.id
            WHERE m.sezon=%s AND ps.druzyna='gtk'
            GROUP BY ps.nr ORDER BY SUM(ps.pts) DESC
        """, (sezon_filter,))
    players = cur.fetchall()
    cur.close()

    rows = ""
    for i, p in enumerate(players):
        fga = int(p.get("p2a",0) or 0) + int(p.get("p3a",0) or 0)
        fta = int(p.get("fta",0) or 0)
        ftm = int(p.get("ftm",0) or 0)
        pm2 = int(p.get("p2m",0) or 0)
        pm3 = int(p.get("p3m",0) or 0)
        pts = int(p.get("pts",0) or 0)
        efg = f"{(pm2+1.5*pm3)/fga:.1%}" if fga else "-"
        ts  = f"{pts/(2*(fga+0.44*fta)):.1%}" if (fga+fta) else "-"
        n   = int(p.get("mecze",1) or 1)
        ppg = f"{pts/n:.1f}"
        nie = p.get('ma_nieprzypisane')
        nazwa = p.get('nazwa','?')
        bg = "background:#fff8e1" if nie else ("background:#f8f9ff" if i%2==0 else "")
        warn = ' <span title="Przypisz zawodnika w raporcie meczu" style="color:#f9a825;font-size:.75rem">⚠ nieprzypisany</span>' if nie else ''
        rows += f"""<tr style="{bg}">
            <td class="fw-bold">{nazwa}{warn}</td>
            <td>{ppg}</td>
            <td>{pm2}/{int(p.get('p2a',0) or 0)}</td>
            <td>{pm3}/{int(p.get('p3a',0) or 0)}</td>
            <td>{ftm}/{fta}</td>
            <td><b>{efg}</b></td>
            <td>{ts}</td>
            <td>{int(p.get('ast',0) or 0)}</td>
            <td>{int(p.get('oreb',0) or 0)}</td>
            <td>{int(p.get('dreb',0) or 0)}</td>
            <td>{int(p.get('br',0) or 0)}</td>
            <td>{int(p.get('finishes',0) or 0)}</td>
            <td class="fw-bold" style="color:#1a2b4a">{pts}</td>
            <td style="font-size:.78rem;color:#888">{n}</td>
        </tr>"""

    season_opts = "".join([f'<option value="{s}" {"selected" if s==sezon_filter else ""}>{s}</option>' for s in sezony])
    gtk_name = get_setting("gtk_name") or "GTK"

    def th(label, col):
        return f'<th style="cursor:pointer;white-space:nowrap;user-select:none" onclick="sortZaw({col})"><span id="thz_{col}">{label}</span></th>'

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
  <div class="page-title mb-0">📈 Statystyki indywidualne</div>
  <form method="GET" class="d-flex gap-2 align-items-center">
    <select name="sezon" class="form-select form-select-sm" style="width:120px" onchange="this.form.submit()">
      {season_opts}
    </select>
  </form>
</div>
<div class="card">
  <div class="card-body p-2">
    <p style="font-size:.78rem;color:#aaa">Suma wszystkich meczów sezonu {sezon_filter} ({n_matches} meczów)</p>
    <div class="table-responsive">
      <table class="table table-hover mb-0" id="zawTable">
        <thead><tr>
          {th('Zawodnik',0)}
          {th('PPG',1)}
          {th('2PM/A',2)}
          {th('3PM/A',3)}
          {th('FTM/A',4)}
          {th('eFG%',5)}
          {th('TS%',6)}
          {th('AST',7)}
          {th('OREB',8)}
          {th('DREB',9)}
          {th('BR',10)}
          {th('FIN',11)}
          {th('PTS',12)}
          {th('Mecze',13)}
        </tr></thead>
        <tbody id="zawBody">
          {rows if rows else '<tr><td colspan="14" class="text-center text-muted py-4">Brak danych zawodników</td></tr>'}
        </tbody>
      </table>
    </div>
  </div>
</div>"""

    scripts = """<script>
let _zawDir = {};
function sortZaw(col) {
    const tbody = document.getElementById('zawBody');
    if(!tbody) return;
    const rows = Array.from(tbody.querySelectorAll('tr'));
    _zawDir[col] = !_zawDir[col];

    // Aktualizuj strzałki
    document.querySelectorAll('[id^="thz_"]').forEach(el => {
        el.textContent = el.textContent.replace(/ [↑↓]$/,'');
    });
    const thEl = document.getElementById('thz_' + col);
    if(thEl) thEl.textContent += _zawDir[col] ? ' ↓' : ' ↑';

    rows.sort((a, b) => {
        const av = a.cells[col]?.textContent.trim() || '';
        const bv = b.cells[col]?.textContent.trim() || '';
        // Liczba lub tekst
        const an = parseFloat(av.replace('%','').replace('/','.')); 
        const bn = parseFloat(bv.replace('%','').replace('/','.')); 
        if(!isNaN(an) && !isNaN(bn)) {
            return _zawDir[col] ? bn - an : an - bn;
        }
        return _zawDir[col] ? bv.localeCompare(av,'pl') : av.localeCompare(bv,'pl');
    });
    rows.forEach(r => tbody.appendChild(r));
}
// Domyślnie sortuj po PPG malejąco
window.addEventListener('DOMContentLoaded', () => { _zawDir[1]=true; sortZaw(1); });
</script>"""

    return render_template_string(base(content, scripts, active="players"))

# ══════════════════════════════════════════════════════════════════════════════
# USTAWIENIA
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/ustawienia", methods=["GET","POST"])
def ustawienia():
    if request.method == "POST":
        set_setting("gtk_name", request.form.get("gtk_name","GTK"))
        set_setting("current_season", request.form.get("current_season","2024/25"))
        flash("Ustawienia zapisane!","success")
        return redirect(url_for("ustawienia"))

    gtk_name = get_setting("gtk_name") or "GTK"
    season   = get_setting("current_season") or "2024/25"

    content = f"""
<div class="page-title">⚙️ Ustawienia</div>
<div class="row justify-content-center">
<div class="col-lg-6">
  <div class="card p-3">
    <form method="POST">
      <div class="mb-3">
        <label class="form-label fw-bold">Nazwa drużyny głównej</label>
        <input type="text" name="gtk_name" class="form-control" value="{gtk_name}" placeholder="np. GTK, Kotwica, AZS...">
        <div class="form-text">Ta nazwa pojawi się we wszystkich raportach i statystykach.</div>
      </div>
      <div class="mb-3">
        <label class="form-label fw-bold">Aktualny sezon</label>
        <input type="text" name="current_season" class="form-control" value="{season}" placeholder="np. 2024/25">
        <div class="form-text">Domyślny sezon przy wgrywaniu nowych meczów.</div>
      </div>
      <button type="submit" class="btn btn-primary w-100">Zapisz ustawienia</button>
    </form>
  </div>

  <div class="card mt-3 p-3">
    <div class="section-hdr">Pobierz szablony</div>
    <div class="row g-2">
      <div class="col-6">
        <a href="/template/zapis" class="btn btn-outline-primary w-100 btn-sm">📝 Zapis meczu</a>
      </div>
      <div class="col-6">
        <a href="/template/szablon" class="btn btn-outline-success w-100 btn-sm">📋 Szablon raportu</a>
      </div>
    </div>
  </div>
</div></div>"""

    flash_html = ""
    return render_template_string(base(content, active="settings"))

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT (z match_id)
# ══════════════════════════════════════════════════════════════════════════════

def _get_szablon_b64():
    return (
    "UEsDBBQABgAIAAAAIQDFxOklggEAALgHAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADMlc9OAjEQxu8mvsOmV8MWUIkxLBxQj0oiPkBtZ9mG"
    "btu0BeHtnS1/YswKIWziXrbZdub7fp2kM8PxulTJCpyXRmekl3ZJApobIfU8Ix+zl84DSXxgWjBlNGRk"
    "A56MR9dXw9nGgk8wW/uMFCHYR0o9L6BkPjUWNJ7kxpUs4K+bU8v4gs2B9rvdAeVGB9ChEyoNMho+Qc6W"
    "KiTPa9zekjhQniSTbWDllRFmrZKcBSSlKy1+uXR2DilmxhhfSOtvEIPQWofq5G+DXd4blsZJAcmUufDK"
    "SsSga0W/jFt8GrNIj4vUUJo8lxyE4csSK5B664AJXwCEUqVxTUsm9Z77iH8M9jQuvYZBqvtF4TM5+i3h"
    "uG0Jx11LOO5bwjH4J46AfQlo/F7+VKLMiYfhw0aBb7o9RNFTzgVzIN6Dww7eOMBP7RMcnCk+KbCVNVyE"
    "g+4xf+yvU2esx0nj4HyA/SipsjsWhcAFCYdhUteUD444pS6+MVRzUICo8aZx7o6+AQAA//8DAFBLAwQU"
    "AAYACAAAACEAtVUwI/QAAABMAgAACwAIAl9yZWxzLy5yZWxzIKIEAiigAAIAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAKySTU/DMAyG70j8h8j31d2QEEJLd0FIuyFUfoBJ3A+1jaMkG92/JxwQVBqD"
    "A0d/vX78ytvdPI3qyCH24jSsixIUOyO2d62Gl/pxdQcqJnKWRnGs4cQRdtX11faZR0p5KHa9jyqruKih"
    "S8nfI0bT8USxEM8uVxoJE6UchhY9mYFaxk1Z3mL4rgHVQlPtrYawtzeg6pPPm3/XlqbpDT+IOUzs0pkV"
    "yHNiZ9mufMhsIfX5GlVTaDlpsGKecjoieV9kbMDzRJu/E/18LU6cyFIiNBL4Ms9HxyWg9X9atDTxy515"
    "xDcJw6vI8MmCix+o3gEAAP//AwBQSwMEFAAGAAgAAAAhAD3UBfWMAwAA/wgAAA8AAAB4bC93b3JrYm9v"
    "ay54bWykVutu4jgU/r/SvoOV/2niQNIQlY4SknTQAEVA2+0IaeQmpnibxIxjCtVonmofYV9sj8OlAVYr"
    "toPAju3j79y+c8LVp3WeoVcqSsaLtoYvTA3RIuEpK57b2t0k1l0NlZIUKcl4QdvaGy21T9e//3a14uLl"
    "ifMXBABF2dbmUi48wyiTOc1JecEXtICTGRc5kbAUz0a5EJSk5ZxSmWeGZZqOkRNWaBsET5yDwWczltCQ"
    "J8ucFnIDImhGJJhfztmi3KHlyTlwOREvy4We8HwBEE8sY/KtAtVQnnjd54IL8pSB22tso7WArwM/bMJg"
    "7TTB0YmqnCWCl3wmLwDa2Bh94j82DYwPQrA+jcF5SE1D0Femcri3SjgftMrZYznvYNj8ZTQM1Kq44kHw"
    "Pohm722ztOurGcvo/Ya6iCwWA5KrTGUaykgpo5RJmra1S1jyFT3YEMtFsGQZnFqtS8vWjOs9nYcCFpB7"
    "P5NUFETSDi8kUG1r+q/SqsLuzDmQGI3o9yUTFGoHKATuwEgSjzyVQyLnaCmytnbjTft///UnSt/Kl2ng"
    "j79Ek6n/xQ+jftefdnq3d2GE/IHfe/waTWt8JKfk/x+MJIkKiAFB2Bi6eT4OCNgrvB3rhlIgeO6GPYj8"
    "mLxCHiDb6bZMuxBo99uPVux2HMfCOg7dQG9GLUcPYtfWg1YYRX7sW66Lf4IXwvESTpZyvs2twmxrTUjk"
    "yVGfrHcn2PSWLH3X/8PcfnQ1Hw27s5/KU9XF7hldle8sUEu0fmBFyldtTccmdMG3w+WqOnxgqZwDjRqW"
    "DdWy2ftM2fMcLMZWU20C25Vlbe3AonBjUQwfXQ0HFhk1k6p+CaZVMyoqjk8iv49uokE08nvQnlVHVTHG"
    "GhKe0iS6KVae1e8Me/5jNBrXxKGD7cWtY/FedxDdDevijZp441h8/Pl2gibdfndwU9PQrF1pHl/p3N6N"
    "Jujr7SCqa7FrV6rCrPvQi8Dp0K9pgPjufXAq2u7ilZAsGQqkpio2LWxaLSVB17JXymqGImOQl8B2A7PR"
    "svRmjGO9iVumHgROU7fDuGFf4rAT2bEipnrReWuFOPtg/3KN6jYlcgmFr2q+WntqjLe7+83ZZmOb84MC"
    "9kahcmV7+78Ex/Aiz+iZwvH9mYKdQX/SP1O2F02+PcTnCvv9IPTPl/dHI/9xEv2xU2H8a0ANyDl0sV3m"
    "jd1/l+t/AAAA//8DAFBLAwQUAAYACAAAACEAZwxbpSkBAAABBgAAGgAIAXhsL19yZWxzL3dvcmtib29r"
    "LnhtbC5yZWxzIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAvJTLboMwEEX3lfoPlvfFQBL6UEwWrSpl29IP"
    "GJkBo4CNbPfB39eiakikyNkgNpZmRp57PLbvdvfTteQLjW204jSJYkpQCV02qub0o3i9e6DEOlAltFoh"
    "pwNaustvb7Zv2ILzm6xsekt8F2U5lc71T4xZIbEDG+kela9U2nTgfGhq1oM4QI0sjeOMmdMeND/rSfYl"
    "p2Zfev1i6L3y9d66qhqBL1p8dqjcBQlm3dD6A5ACTI2O07848oyUXZZfzSn/rc3BSkQ3ERxTlo2VVQjm"
    "fk4Y5+8IJ5AxZOOahBjShQeShmCShWGCk8kWhslCk9ksDLMJXpN3lfk+sYBWPEto1PR4j6kQxXrhkaxD"
    "MI+z2poEg+W7M961T93tNP0Pw86MO/8FAAD//wMAUEsDBBQABgAIAAAAIQBYz9TTERQAALJ2AAAYAAAA"
    "eGwvd29ya3NoZWV0cy9zaGVldDEueG1srF1rUyI7t/5+qs5/sPg0e/odEYFWKPWtAAICykW56DdGcYba"
    "Kr7AXPY5df77WelOoDtPVoyMuzaO85B1y+1JJyvTJ//+/fy093O2XM0XL6eZ3P5BZm/2cr94mL98O80M"
    "b+pfjjN7q/X05WH6tHiZnWb+ma0y/z777/86+bVY/r36Pput90jDy+o08329fi1ns6v777Pn6Wp/8Tp7"
    "oW8eF8vn6Zr+uvyWXb0uZ9OHSOj5KXt4cBBmn6fzl0ysobz00bF4fJzfz2qL+x/Ps5d1rGQ5e5quyf/V"
    "9/nrSmt7vvdR9zxd/v3j9cv94vmVVHydP83X/0RKM3vP9+WLby+L5fTrE8X9O1eY3u/9XtL/h/TJazMR"
    "Dpae5/fLxWrxuN4nzdnYZwy/lC1lp/cbTRi/l5pcIbuc/ZzLBtyqOtzNpVxxo+twqyy/o7Jwo0xW17L8"
    "Y/5wmvnfA/XfF/ozJ38cbH/o7/4vc3byMKcWllHtLWePpxmRK4turpTJnp1EPWg0n/1aJX7fW31f/Gos"
    "5w+d+cuMuiN15PX06/XsaXa/npHdXGZPdtmvi8XfUvSCoAOy8jp9me39vn6lho/K/KN+LZD44rUze1xX"
    "Z09Pp5lKMbM3vV/Pf856JHGa+bpYrxfPg/m37+togKwJe1wu/mf2EnkYmZW+S/2nGVIVF5XeG9/FmqQl"
    "m2Tazn82VXGVk6Wzm6pI/q6rpR6Nvd5y72H2OP3xtB4sfjVn0mOKtLBPEUadt/zwT222uqfRRDWyn5dq"
    "7xdPpIN+7j3P5axAg2H6O67C+cP6O/1Gs8L9jxXVwFgBSiwWoL4TCeRDqvS4wJFTIE9fRxIF2UyxhNVE"
    "NnYtCrs2XU/PTpaLX3s0ckhuRXVNDZ8rSyUyyMPifomaLXZ0EzkTNsUr9Qip6DQTkkOkYUU95OfZwUn2"
    "J9XvvSpSiYsUC1HMUqgKSA2Qc0DqgDQAaQJyAUgLkDYgHUAuAaEuJUNPxNUFpAdIH5ABINeA3AAyBGQE"
    "yBiQCSC3gNwBIlQrJ0IV2KoCm1VguwpsWIEtK7BpBbatwMYV2LoCm1dg+wpsYKFbuBjNHTRyNsOHRiwM"
    "n1x+h+EjFVEfSg6fnDF8VJHE8AGkBsg5IHVAGoA0AbkApAVIG5AOIJeAXAHSBaQHSB+QASDXgNwAMgRk"
    "BMgYkAkgt4DcASIEQhWEqghhuwpsWIEtK7BpBbatwMYV2LoCm1dg+wpsYKFbGIcPrW9w+IT7B+9mH6ko"
    "IrMN+RwaoycuERKzbork00WqcZHE5FYD5FwhUSgR8dUVQrS7UVxIK26AmiYgF6C4pTyWCzLNqMW04jao"
    "6QByCYqvFJKsijCtuAtCPeUNccLGm6O0UB9sDwC5BsU3SrFcOOowj9OKh6BmBGrGMUI/N1pKaS0TZShZ"
    "JGcsUW7B0h0gQiBUQUj3pm1PEao7hdS7N07mjFle6B62neaF7mIJVbpLJTUZPV6oTpY2Z/R5oftdwpzq"
    "eMWEOd3RUnVn9HKBfU/ozpfQrntfAkr1tmySXqlTwPxQ2CxwfZemUot7cohLpMjXGGpVSxFj0NQsRYwh"
    "cm4pYnT2elwkOZ3kjJ7cwCKHRkduWooYPe3CUsToQq24SCnRyw6NDtS2FDE6RsdSxKjdS0sRo3avVCsm"
    "p8NDo3q7tjJG/fZUmeREdmhUcN9SJm/U8MBWxqjia1sZo45vVBki/S0xGZU8tJUxanlkK2NU89hWxqjn"
    "iSqTHOp5o55vbWWMer6zlTHqWQhLoYJR0aJiK2ROm2p85pJuF8wZUY3QZNMXzAlRDdFUGXOmU2M0Vcao"
    "aqEGaaqMUdVCjdJUGaOqhRqmqTJGVQs1TlNlzKpWAzVHlbldTphVrYZqupBZ1WqwpguZVa2Ha8qcWdd6"
    "vKYKbSs7RQVEmrhULO3wpCUV0UZFcrwVjcaTm1bpJ/oqIDVAzgGpA9IApAnIBSAtQNqAdAC5BOQKkC4g"
    "PUD6gAwAuQbkBpAhICNAxoBMALkF5A4QIRDCVhXYrALbVWDDCmxZgU0rsG0FNq7A1hXYvALbV2ADC2xh"
    "oZsYn77oAeBjnr6kIlpgUZVvpxZjrquoMuF28w+QmkKOaPfwMVL1fbqcPWTijexaWK7l5JbUPNqBvqif"
    "DwbdwadK+KUa/uvgr5Pso2XP8VypVMup9dKq+jwsn2vVucyZVl0Ns5XwX5kvGa38S/qJog4RNABpOmNq"
    "huWmNny4NVwPgwYf04VXTBdh+UKrzidVZ5uumFoQQRuQjjOmTljuaMOFreFWGLT5mC69YroMy5dadTGp"
    "OttxxXQFEXQB6QHSB2TgjHsQlgfauXDrXC8M+nzc115xX4fla62aBofun70wO3DFfQMRDAEZOWMaheWR"
    "Nny8NXwTBkM+pjEYmQByC8gdIEIghNOIqGIpPZFsJxuhJ4IEpIevfboR9bBMG9Jqwiltg2+GAfU2bsIR"
    "egpgtDZIa2Mz19BJ2nayCerh58OgFX7O89r1dKIeyZj5TNC8Qtvkynf6c2OFaisraq5OI/T08oYJml9o"
    "212bSExeVANvmdCTzBsmWmSitTGRmMQiE+fOKPSs5Z74RZtMtDcmEtPVJ5qDc/vFz63wr+ynqMn/cjCB"
    "0FPiG+ZobqQjB11piSlMRvTp8PMn6nXBwX6h8LkS/uU06DdhCpox6UBDG0xMS5M3G0nPmm800hVZuNpY"
    "SMxOFbJQd7aRmoUP1SM915m7ZKG7sZCYhqjSPucODvjelnqAoN2dj1ntSEXmasd4aquoMonVDiA1hcBE"
    "ES9w6DgZTlDPlQjby6RodMD988xcq4D9BiBNl0eHcuEGHl34eBSdjYNHLbDfBqTj8ijaQQSPLn08itbC"
    "4NEV2O8C0gOkD8jA5XXUJcDrax+vj+ImMFr2BuwPARm5PDq2tuwYlEwAuQXkDhAhEMLhIapYSg+QJItj"
    "qbortCjtBSpb6K5vHXq5A2t9CD062PkwGnz20Sf0OHELqzFmNLDQI8UtbB9lQg8q95yhxpJpWY8/t7B9"
    "NAmvkZhT48G0rMeiO2b7iBBq2PLEEjWV6vdbyym6kPl7ZmJMbpejSanIpAtjA6+iyiToApCaQt5DF0pk"
    "F7oA+w1Ami6PGLrw8YihC7DfBqTj8oihCx+PGLoA+11AeoD0ARm4vGbowsdrhi7A/hCQkcsjhi5AyQSQ"
    "W0DuABECIRweooql9ABJ0gWWqrtC4+jCJcPShRLajS68hDm68BLm6MKnZ9HWf7QYBLrwEubowkuYowuv"
    "mDm6iIX/iC6omT+GLqQiky6Ms5yKKpOgC0BqCnkPXSgRqg25jYuPY46nC7DfAKTp8oihCx+PGLoA+21A"
    "Oi6PGLrw8YihC7DfBaQHSB+Qgctrhi58vGboAuwPARm5PGLoApRMALkF5A4QIRDC4SGqWEoPkCRdYKm6"
    "KzSOLlwyLF0oIfrDMfi4pwsvYY4uvIQ5ulDC7jmDowsvYY4uvIQ5uvCKmaOLWPiP6IK6wQfxRaTJIIwQ"
    "Mu+lOSqUYAwtt4VqGnoPZ2iZXUgDXWgg1HR6xfCGl1cMcaALbYQ6Tq8Y7vDyiiEPdKGLUA+hPkIDp+8M"
    "g3j5zlAIujBEaOT0imERVDNB6BahO4TobgMMErrcgFjVgm2GTpJNLOXqziA5PnEK0Ze2jVk6/lGu78Yo"
    "ftIcpfhJc5yipXcjFT9pjlX8pDla8Yub4xUl/WfEIs96PmTfKjpxoctgyayy0LyTogvR+d/J4+aUsRIG"
    "laOgchxUSgENIfY4s2oXr4ZB9SioHgfVUkCjjRWP00HIxch6MmHEcYSiTMqbeO9+yrG7S0d29aOgfhzU"
    "SwENcNbd+NhXu7uprAYlexwFjeOgUQqI/1jx+DyXi5YjQtkd4nuHfLQcEcayce1u3KU8jtZR0DoOWqWg"
    "5XA3PlWFaNuUBnIUtI+DdikgXmWjjU9JuWg5gvWJliNYa7RXYXB1FFwdB1el4Mrhbnw+CdF2w6B7FHSP"
    "g24pIL5mo+3ZexbljvSOgt5x0CsFxO2seN8u3qfUk6Ogfxz0SwGtA1jxOG2Fq2tuQeBT19yCwFrXlFVy"
    "cxTcHAc3peDG4e7QHu2QklKOguFxMCwFtL5go40TWrhouYVG0mNITRvn6JL1NoMskSwyDoPxUTA+Dsal"
    "YOxwapKKyZzK4uvbcI51u4vQ3S5CtDjiKiDaFGJWLpWdpNKs4FkXdMN0Fw/jRuM6A1PvKi2Ik2KXcU4H"
    "2WWcGmc7LuO8pNllnJc0u4zzmSHYrWQ/nmaXcV622WWcV9zsMk61M5vW4nH6KO8xfEy6e6SJbhar/hNd"
    "zA/Ny5G6UPJqPlxpq2Gpc4TqCDUQaiJ0gVALoTZCHYQuEbpCqItQD6E+QgOErhG6QWiI0AihMUIThG4R"
    "ukOIZm7LdUULVrVglhYnmkN9ljanJEcsZ2l1yiHEcpZ2pyw9LGdpecp2w3KWtqeMMlUO8+LlzZYPeoaK"
    "ry4dJm8lFc3M+MicsTmnrjwlN+cUxCTH5wqUHS83JikDjB4gt49iucKXaq7AJ8hr6+zzUPRPp5yT/vON"
    "/mSOPP0zOhXS78iSx/gaCDU1ZI+PvqZMeR1fMlU+V6CnJj6+C7/4qBily2v9yXx5io+MO+JrYTBthDru"
    "+OhryprX9pNp8xQfqWPb79IvPipGqfNafzJ3nuIj4474rjCYLkI9hPoIDdy1QF9TDr32MplET7VA6tha"
    "uParBSpGifRafzKTnmqBjDtq4QaDGSI0csdHX1M+vbafTKin+EgdG98YTU0QukXoDiFiBJhdaKMTMZo2"
    "zFMDWmUjRnMDlKu7K4JW0AWZXa+rIpleT1UhOySfYP+GapoNZIq9Up1PptiTajIsk+zppyvLXptgF93x"
    "vyclZyXaaNWmUnn21KNkbbmSxjdz0xtm5OREPKnNpHLtPcxsJqg3zFA5mW+vzaTy7aUZamdXNJtJz80k"
    "RORkpr0xk8q5pwaKku5zBZl1H/cEd96938CnpHsyuplg86nMe4otTr0ne3HuPZl3J997WpWzLi1IdI0m"
    "0+892m0z8b7RblROpuBrK8kUfGmFeryr2fRc/mYWvrQi/yG5aImRT6Xh02iK8/A5U6nMSnlP8IOWV/El"
    "z/TyykzFj8wZy6tYLnEcWtOlrGefVDfWdHwt5NxLpg10W7JUHf1qINR0+8Xt+qr43H5x275QOW30q+P2"
    "i9uf9fKL26AFv7roVw+hPkIDt/fcjqeX99yWJ3g/RL9Gbr+4vUlQPUHVtwjdIURrA1BGawPEqhZsM4ZS"
    "h6Aou+n59rHGbp8pTVYponnmFFRJsXOofGDKc3k1Okq3NLt95mWb3T7z6W55LrVGe+6cAogHrXmcl37S"
    "7PaZV9zs9lks7T4FJf4xPE9TjLwT+jGnoPHt0jTFmOn7RCSQXgNQTZd6H8W8dUE4euDnKAacaKCrTbdf"
    "HMV4+cVRDPjVRr86br84ivHyi6MY8KuLfvUQ6iM0cHvPUYyX9xzFgPdD9Gvk9oujGFA9QdW3CN0hRBRj"
    "udZtwaoWbDOGUhSD+uruKFmKUZreSTFKakeK8ZJmKcZLmqUYn+7GU4yXNEsxXtIsxXjFzVJMLP1nFCPv"
    "HX4MxcQ3GNMUY6b85+DmYxWhmobeRzE+d0/Zpxjwq4F+Nd1+cRTj5RdHMeBXG/3quP3iKMbLL45iwK8u"
    "+tVDqI/QwO09RzFe3nMUA94P0a+R2y+OYkD1BFXfInSHEFGM5c6xBbMMIdqzQ9lzC1Z3R8lSjOu6Mv8U"
    "o6R2pBgvaZZivKRZivHpbjzFeEmzFOMlzVKMV9wsxcTSf0Yx8l7ax1CMuuGWPIfESwLqDl3ykgBANZk4"
    "KTMK30cxSmi3jTJwoqGd2LradPvFUYyXXxzFgF9t9Kvj9oujGC+/OIoBv7roVw+hPkIDt/ccxXh5z1EM"
    "eD9Ev0ZuvziKAdUTVH2L0B1CRDGgjDbKEKtasM0YSj3FoGzdHSVLMa4RylOMktqRYrykWYrxkmYpxqe7"
    "8RTjJc1SjJc0SzFecbMUE0v/GcXI85yPoRipiXJPkv+WK14XUIXo/C15XYAO3miTmT50byBHFwdyx44r"
    "A1YVdFZNVwWK9KG7A6SChp3j2kBShZlfyp71xEJySPKZ9OxTktVpOpmj8+gifegGATlN491xecCqgg67"
    "Kf2lSB+6RUAqiBkdFwiccXMU6RU3R5FWp+mgmU7gi/ShuwTkdMvhdHxQLN9JlOoydIhMaTFF+tB9AlJB"
    "zOu4SuCMm6Ngr7g5CrbGTQe1dImgSB+6VUBOXzmcjk9aIW46raWLBEX60M0CUkHM7rhUYPWCMnXoMkGR"
    "PnS7gFTQSsBxscCqglJxKB2nSB+6YUAqaOXguFzgrH1uCeFV+9wSwuo0ZfDQtYIifeieATl943B6GJ1z"
    "Q+1Tig6l6RTpQ3cNSAWtTBzXDJxxc0sUTihKOBnnSvKegSwjT62S9wzIszF5NibPxuTZ2OHZJBWcOf+R"
    "VttJ2u0uQne7CNHSylVzjH+0+NpFqrqTVJxpqbuHZw2qhnuvVJwkxUkdsotAZ22wp6Vq2NEfjuti7Gmp"
    "lzS7CPSSZheBPhMGvwj0kmYXgV7S7CLQK252Eaja2XnZgDstjV+DGL8P8Hm2/DaT72xc7d0vfsiXGlKS"
    "09nJBt6+TTJ+haL5RTtP6UFR6xjfXOXLVlwcUtZS1BkMgXq+TG+loUkIbEsTlLaNX7XyZXrVDOK9fJne"
    "OGRRVSRNUXMaNiZkgt7cgxI3+TK9dMiiqUYS9IIe/KaSL9O7fCwSTZKgd+zQN9ltpcs3an6bXU6X3+Yv"
    "q70nequlfKvkEZ2FL+M3T8Z/oRdiRm+yiV9vGf36nd4NO6OXKR7sU+HHxWKt/yINbN42e/b/AAAA//8D"
    "AFBLAwQUAAYACAAAACEAHtNDiO0PAAD5aAAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbKyda1Pi"
    "2BaGv5+q8x8oPvW003LRaJtSp0YwkJArCvKVRmypUbGAvp1T57+ftZO9Icmbldlk6Jpp7ce93n19k5C1"
    "Yy7/+Pn6Uvs+X60Xy7ereuu4Wa/N32bLx8Xb16v66N769LleW2+mb4/Tl+Xb/Kr+a76u/3H9739d/liu"
    "/lo/z+ebGim8ra/qz5vNu9lorGfP89fp+nj5Pn+jnzwtV6/TDf1z9bWxfl/Np49x0OtLo91snjVep4u3"
    "eqJgrnQ0lk9Pi9m8u5x9e52/bRKR1fxluqH2r58X72ul9jrTkXudrv769v5ptnx9J4kvi5fF5lcsWq+9"
    "zkz769tyNf3yQv3+2Tqdzmo/V/Rfm/4/UdXEHGp6XcxWy/XyaXNMyo2kzdj9i8ZFYzrbKmH/tWRap43V"
    "/PtCTOBOql2tSS1jq9XeiZ1UFDvbionhWpnfFo9X9f825Z9P9LUl/mp+arbEX6k//6tfXz4uaIZFr2qr"
    "+dNV/c+WOTlp1xvXl/ECGi/mP9ap72vr5+WP3mrx6C7e5rQaaR2LFfplufxLFLWp5iaJvk/f5rWfd+80"
    "z1d16uIv+S01cLN8d+dPm8785eWq3jmt16azzeL7PKSIq/qX5WazfB0uvj5vYj9siD2tlv+Zv8Utmr/M"
    "qTA1Vehf1UkqKSpam/tZoiRqKopM10PBjW1f09+rfluxt8JV7XH+NP32shkuf/Tnoolk5NNj6kG8OM3H"
    "X935ekZuoSE4PhG1zpYvpEF/114XwvW02Kc/468/Fo+bZxFer82+ranLDxLIsCSABi4OoK8y4Ky0PI1u"
    "XJ5KyfLnpeXpp3H5C039Fs120oNdi8praKkmtXZtKu9DSzWqTcOl1Yv2dpioMhlBR9OycaVhTwaWvimr"
    "o5HMYLw6utPN9PpytfxRowMItW1Na5AM0DLbJCLWgviaVLpdHczSoDUhRP4UKlf1M+oxha/JNt+vz04u"
    "G99pEc5kmZukjHEaLwwR1QHSBXILxALSA9IHYgNxgAyAuEA8ID6QAEgIJAIyBHIH5B7ICMgYyAOQiSSG"
    "mIsGrYTtcqAVCMuhRXC/5SBUaDnQ2t0th9PcckjKGHET4kXUSci5OLJuF5GRjerKqN0iugViAekB6QOx"
    "oT2ObA8dKrbt+ZxtzgCC3ISk+3CRjfGkLH3ZdfMsW8aH5gVAQqg7kiQtfJ4VHmLzznJ9upPto2PArn25"
    "PtxDa0bQmnFCMovgvJltzgPoTNI6mZVJQwor83R7dNQ9UAmV+KC37dt5K7cwC4q0s0U6SZH00e48d7Tr"
    "FhTJOeC2oEhuuVsFRXJLpVdQJDfp/YIiuUm3C4rk5tyR45Jet59z8zkoKpMbX7egTN4iskjaRZ9zc+AX"
    "lclNQlBUJjcLYVGZ3DREuCROc9MwLOp5rsydLJN21ufcXN0XlclN1qioTG4Ix9jki7z5Cork5mpSUGQ3"
    "Dxl7UqfwxLG3PYUKnTjoMm7rTyM3GTdJmfR1BJAukFsgFpAekD4QG4gDZADEBeIB8YEEQEIgEZAhkDsg"
    "90BGQMZAHoBMJMHrCKNoOZwdN4nvdykhhOiK9Gx71XiTELoMF59L4utIIF1JzulK9CleSs/T1fyxnnww"
    "7BpmV1zFrxfxRzzbuh0Og+GHjnHUN35v/nbZeBLXrzm33EIlVmkllmFaqpJW/VpVcltaSQ8q6QOxgThA"
    "BkBcIB4QH0gAJAQSlQ5DZJiRGob2bhh6xlHH+Ng+ujU+nrBDPkyU2+fbmb6TdckLy82qcH7vDPNO1Xmy"
    "q1NMcOvY+Hhr/Nb40DWOLOO33+uf6mq+P2XP8fdaVd0b5r2q6nRXVWQ0PrQ/JpUcNY9PTz/2jd/Kahtp"
    "1TYyzJGqzdjV5lFtnhjQI+oVLbDyjo21qhob5lhVdVa/pnGStsgN0wPM0USSz/DZQlgu/1GzVeWYIISy"
    "x4SEpI8JQLqSwDEhOQzQwQQ9DyJWmUhL9BhEeiDSB2IDcYAMgLhAPCA+kABICCQq62Z8Sw26OUxC0maV"
    "IqxZxcDHd5e+X+fNpxMafwaE0JFOaHy+gtCxTmh82oHQB+j9RBK0AX1sOIwNhFDWBglJ2wBIV5J9bAAi"
    "VpkIYwMQ6QOxgThABkBcIB4QH0gAJAQSlXWTsUESkraBFKliA51QxgY6oYwNdEIZG0DvJ5KgDUSa6CBn"
    "AyGUtUFC0jYA0pVkHxuAiFUmwtgARPpAbCAOkAEQF4gHxAcSAAmBRGXdZGyQhKRtIEXoi/hoh5duJWcD"
    "nVDGBjqhjA10QhkbQO8nkqAN6KPuYWwghLI2SEjaBkC6kuxjAxCxykQYG4BIH4gNxAEyAOIC8YD4QAIg"
    "IZCorJuMDZKQtA2kSBUb6IQyNtAJZWygE8rYAHo/kQRtIBJyh/FBrJQ1gkRpJyDqKrSPF1DGKpVh3IAy"
    "fUQ2IgfRAJGLyEPkIwoQhYii0h4zxpAxaWcomSrW0IplvKEVy5hDK5ZxBw7BRKECf4iM3UEul1pJ7i99"
    "R02ijD+SUinUVaX28gfIWKUynD9Apo9tthE5iAaIXEQeIh9RgChEFJX2mPOHnKXUvS8lU8kfcgBLYzl/"
    "6MRy/tCJ5fwBQzBRa7fAHyJveBh/JBnIjD8SlPEHoK5Ik4vc4l7+ABmrVIbzB8j0lczuNrmNyEE0QOQi"
    "8hD5iAJEIaKotMecP+Qspf0hB6GSP3RiOX/oxHL+0Inl/AFDMJEj2S7wh8jkHcYfSU4w4w+ZJkylZMR2"
    "KWGGHeoqtJc/QMYqleH8ATJ9bKCNyEE0QOQi8hD5iAJEIaKotMecP+Qspf0hB6GSP3RiOX/oxHL+0Inl"
    "/AFDMJEjWeQPkck8jD+SnGjGHwnKnD8AdcUWyr3PHyBjlcpw/gCZvpJJnz+glIOlBohcRB4iH1GAKEQU"
    "lfaY84ecpbQ/ZPcq+UMnlvOHTiznD51Yzh8wBBM5kkX+EFm9w/hD5gdTGf0WJJU7iLoK7XX+AGWrVIbz"
    "B8j0sYE2IgfRAJGLyEPkIwoQhYii0h5z/sDcu5Kp5I+/SzqLu7+cP3RiOX/oxHL+wNS2HIIif4h032H8"
    "IROHaX9AgrYT5+Zz11dlSdtmYW76FmWSTSvcxxjOH9DAPirbiBxEA0QuIg+RjyhAFCJK9qdwPeb8gelu"
    "pVzJH/8g4a1VL+ePf5DylvWmbuFNFCr4/CHygIfxh8wo0q3k3f7k/B5e8ewFXUudiCd4nq7vRt6HG8O8"
    "aZ0xm7s6WL5jmGQypnxXlY/9EG9Cu0WJW8MkgzESFkr0MhKwbY1+bIZxETpW0klQ9KpnmD22ij6rJ462"
    "8ZUPbOOwK8Q4FWIGFWLcCjFehRi/QkxQISaZTLVK0/PNz0+Ey2aIK29omEN2WdyhxD2iEaIxoges+8Ew"
    "H9i6JxmJzAbilsgNwRGCDqZ7PomUpJjO6ebh7vCQ2zt+E9d1Vc88i5TEpVAXS90ishD1EPUR2YgcRANE"
    "LiIPkY8oQBQiihANEd0hukc0QjRG9IBoolDBE0oHyxS2hVI2UyhR+pM6oq5CzPbidtPsnoib13SU/bzb"
    "ptlpN4/67Sa/wxhrssproh+blqrpIrXJ+G9q6mFN1LD8DTsbkYNogMhF5CHyEQWIQkRR+aDQj81IDQol"
    "1LZbr6nXRzQHYtcx/V2y7Vjqp3OTqkr22i5+4viOqr7bVp3a9f1BzH2897jdFJuP6V80d6Xbj/VqvKca"
    "77c1pjZY0ygkW5CTquQmZKqzdBeyXqUjqnS0rTS1w5pmmXYiJ8N8JDpJA/13HR3r1UnFzPG2Ttpqze9I"
    "xumbKITXpeLZ3MNcl8ZKuaMJpC07slTqANNVqPC+Rtxi3JiMMlapzEXx1mSU6SOyETmIBohcRB4iH1GA"
    "KEQUlfaYnF+0G3uopil140/plH5wo9RjrJffqawXLD9F5oJHesHFe6THesHFN1UecBgmChWY5GDJVfGE"
    "e/6Ui8lVWSpjkrLkKmcSTK4q5UKvcSbB5Co20EbkIBogchF5iHxEAaIQUVTaY9YkmF1VOtVMopPmpORl"
    "kcNGejVzJtGqmTMJZljVAi4wycEyrG3MsEqUuS7FDKsqtdeZBDOspTKcSTDDim22ETmIBohcRB4iH1GA"
    "KEQUlfaYNQmmWJVONZPo5DpZk2gFcybRCuZMgmlWtYALTHKwNKv4HSX5MwmmWWWpzJmkLM3KnUkwzaqU"
    "9zqTYJoVG2gjchANELmIPEQ+ogBRiCgq7TFrEsyzKp1qJtFJeLIm0QrmTKIVzJkEc61qAReY5GC51jbm"
    "WiXKnEkgu9lVpfY6k2CutVSGO5NgrhXbbCNyEA0QuYg8RD6iAFGIKCrtMWsSTLYqnWom0cl6sibRCuZM"
    "ohXMmQQTrmoBF5jkYAlX2qkDZxJMuMpSmTNJWcKVO5OAsqWU9zqTYMIVG2gjchANELmIPEQ+ogBRiCgq"
    "7TFrEsy4Kp1qJtFJfbIm0QrmTKIVzJkEhmGiFnCBSQ6WdaWboWASeF62I0tlTFL2DC1nElC2lPJeJgGZ"
    "PjbQRuQgGiByEXmIfEQBohBRVNpj1iT4+LHSqWYSnWeBWZNoBXMm0QrmTILPIasFXGASGpkD3QIWSrlb"
    "wAnKXG4B6tKGIn7rJ2cSkLFKZbjLLZDpK5nU1k9EDqIBIheRh8hHFCAKEUWlPWZNIqcpfQtYjkI1k2gF"
    "c3e3tII5k2gFcyaBYZjI4SzY39YuzM5X+eUtsVLOJPCsbUeWypxJyp6/5UwCypZS3utMAjJ9bKCNyEE0"
    "QOQi8hD5iAJEIaKotMesSfDRZaVTzSQ6zxGzZxKtYM4kWsGcSfAZZrWA8UxC280OdCaJlbImkSh9JkHU"
    "VWifD+4oY5XKMGcSlOkjshE5iAaIXEQeIh9RgChEFJX2mDOJDErvFFA6lUyiF8ycSfSCGZPoBTMmwWGY"
    "KFRgkoNl3OPtADmTYMZdlkqfSRTayyT4JHOpDGcSfJIZG2gjchANELmIPEQ+ogBRiCjZVsPuDWcy7mqa"
    "UpdbSrqaSXSeKebOJHo1cybRqpkzCT7OrEamwCQHy7jTGxDindD0hd8urQrRXh61XZr229yctLj90hhA"
    "O5zMDhvQVQHx4CQbplGDtguZt6yGhRq9jAbumD5pm2FchHbi0bMbyZZpqqTHVtJnFcVuPtIo2lRiVwly"
    "qgQNqgS5VYK8KkF+laCgSlAyrbR3Ol6y+b3TzDxFuIKGuAqHtECG7AK5Q417RCNEY0QPWPkDVf7AVi7e"
    "ZxK7ObZRsn06eeNH8k6H1/nq61y8j2Rdmy2/ifd30K7Qyy3dvRcl3syU43dtk36LO63vHB+3Tfqt7Mj/"
    "bJs3hfzUpN8LjOWdtkm/L79Ap3Vhiv2++BOP7FsU0Wmb9Ov6xasLdh0Wb2r5Ovemq6+Lt3Xthd6WIl5e"
    "ck4phVXygpPkH/SilfjXvyevTYm/faZXDM3pdRTNYyr8tFxu1D9EBduXFl3/HwAA//8DAFBLAwQUAAYA"
    "CAAAACEAPL8da30kAABkMQEAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQzLnhtbKydW1PcyLKF30/E+Q8E"
    "T7PHGzcNNLY7bHaAaJBAwrpwcfPG4LZNjAEHMPbMOXH++0mpqhpJS5lkl3rHbF8WlVWrqqSvJXUq/f4/"
    "f99+X/k5e3i8ub/7sDp8vb66Mru7vv98c/f1w+rZ6cHa29WVx6eru89X3+/vZh9W/5k9rv5n57//6/2v"
    "+4c/H7/NZk8r1MPd44fVb09PP8aDweP1t9nt1ePr+x+zO/rJl/uH26sn+uvD18Hjj4fZ1ecq6Pb7YGN9"
    "fXtwe3Vzt2p6GD9o+rj/8uXmerZ/f/3X7ezuyXTyMPt+9UT+H7/d/Hh0vd1ea7q7vXr4868fa9f3tz+o"
    "iz9uvt88/VN1urpyez2Ovt7dP1z98Z3m/fdw6+p65e8H+m+D/r/phql0GOn25vrh/vH+y9Nr6nlgPOP0"
    "3w3eDa6u5z3h/FXdDLcGD7OfN+UGPne14WdpOJr3tfHc2aZnZ9vzzsrlehj/dfP5w+r/rtv/rdHvw/KX"
    "9bV1OhaqP7mf/d/qzvvPN7TD5axWHmZfPqzuDse7e2/WVwc776sj6Pxm9uux9ueVx2/3vw4fbj7HN3cz"
    "OhzpQC4P0T/u7/8sm0Y09Dr1+uPqbrbyd/GDNpoO99WVf+wfyeHT/Y949uUpmH3//mF1b2t15er66ebn"
    "LKWID6t/3D893d/mN1+/PVUnxBNpXx7u/2d2VzmafZ9RY/Ja9v9hlboyTUu3rZ+ZnsqRuiLr41DwYD7X"
    "+p/dvA+qkyt9WPk8+3L11/en/P5XOCst0tS2XtMMqqNz/Pmf/dnjNZ0utASvN8tRr++/Ux/068rtTXna"
    "09F+9Xf1+6+bz0/fPqxu0PZf//VIU74wwtCGmQD6aRVAv9sAwoTQnlbXDLA9D6A/CQHDNy7i3TyCJMkS"
    "bbjxRLOxpuQxyjmaiJF2DOrQRNB0pYkPzAJXm7d/9XS18/7h/tcKneDk7ZEOETo+h+MN6qTcqg3aKDOx"
    "+eYxO0dbVnayW/byYXWbFoTCH+mo/rnzbvP94CcdI9e2zZ5pM9qq9q2MCkDZB2UCygEoh6CEoESgHIFy"
    "DEoMSgLKCSgfQUlByUDJQSlAOQXlDJRzUC5A+QTKFJRLUHbtNte2cNft6qjc1QEdU/MDi45lOLCG85NY"
    "e2CVvdCBRQfm84G11TqwTJtRZaE6HAOjvCkR6g7H1tG4b4Oej8YJKAfQ8aHtmE7PZzujpp0Q+olAOYKe"
    "j23PdD4/97zd7DmGfhJQTqDnj0ahU/q54zfNjtOOJm+bTbKOjRgOm21ycFOAcgr+zqxS8zfcaHZ83jV4"
    "a0MvYKhPMNTULnJjLd41x7qEfnZ3Udqr99047OmYg8N+a/5ZpD3sy14qNs/3bLi+3jrsu9q0diQwbepY"
    "ftPa+/2OJq29n2CT4Xpriw662rS26NC0qZ/Lb1pNwo4mrdM9wibD9Vabo642rfP02LShT/P5Gr9pNYk7"
    "mrTOyQSbDNdbbU662rS24aPdTfqNP0072rxt7UPWdVS09jPHNlutBSy6ummdKacdbYatg/QM22y3Zn7e"
    "4aZl+KJrpNah/qmrTWtxprZN/fx/27Jz2dFmOGwdpru7na1ai7i719nq+SBrgINM4eflwuAoe6HPy/pR"
    "PRy2jsfyToIa1a/EQNkHZQLKASiHoISgRKAcgXIMSgxKAsoJKB9BSUHJQMlBKUA5BeUMlHNQLkD5BMoU"
    "lEtQdndRcruKV2J0tYJHFomLXeKXvdCtwfb88n0PlMAow+c2+6BMrGI/BZ4e6GbhS8W8b1cPs8+r5t56"
    "MhpPhnQQP95UN8nRwSTPP+a/BaPB/ujfq2ur/3o/+FLeZKw1PxMPbOdvOjs9GI0PXKfD1R3X6f5oLRj9"
    "e9112SLYIcwgBCVSzSkajSM3/Mbz8IejQSjN6Uic09FofOQ63XzuNBytHfJzOoYZxKAkqjklo3Hiht96"
    "Hv54NIilOZ2IczoZjU9cp6PnTuPR2jE/p48wgxSUDJQclMIqlqDM8VmMxoWzuP1sMR8NMmnep+K8T0fj"
    "U9cpHcHu+AxHr2gxuePzTOzybDQ+c12+fe4yGL06HP2+8ep49Psm2/O5aiXOR+NzN8K75xHOXliJC1Xn"
    "F6PxxfyMpYdlbkmo91xa50+wrVPV4Twdjafz8WqI+O1w9Gr4evT78ehfg9+q/fiXQKFLM9YGPf8pr+WY"
    "Q+hyNL6cj1Xjwdnod7rWl4+j3V07xJs5jXctjjfpJ8Ko1IyeWs7HrSHjcrRG3XZPq3GhQreoS/g4KXtp"
    "fpyAEhil/nECysQq7MeJ+QShhxP4YWFD4cOiDKmeLv7caX8YwPghKJHG0Ua5b+DoSHJUPSYFR8cwfgxK"
    "onFUPQgBRyeSo+pKAxx9hPFTUDJQclAKq7AgLreputIA16eS6+qUAddnUshbs1utg+FcY/Bd50ZfaEKH"
    "1VcLMLlPsFJTzf4O7RHdumK6NLE8q6qToftw3d21wXUKGYmnUNWfPZafvTQAQ48rlgCYspcmYEAJjFIH"
    "DCgTq/gAxoYuAhgYPwQl0jhiACM5YgAD48egJBpHDGAkRwxgYPwUlAyUHJTCKj6AkVwzgJFCGMBoDDKA"
    "0YRygIGVmmr2lwOMifUEjA2uA8ZIvQBTZhC0v/Ma0nXTYjfEZS9NwIASGKUOGFAmVvEBjA1dBDAwfghK"
    "pHHEAEZyxAAGxo9BSTSOGMBIjhjAwPgpKBkoOSiFVXwAI7lmACOFMIDRGGQAownlAAMrNdXsLwcYE+sJ"
    "GBtcB4yRegGmfHLVHzBlL03AgBIYpQ4YUCZWoXl13xgKt0g2dBHAwPghKJHGEQMYyREDGBg/BiXROGIA"
    "IzliAAPjp6BkoOSgFFah3/idZW6RJNcMYKQQBjAagwxgNKEcYGClppr95QBjYj0BY4PrgDFSL8DQxJdB"
    "mKqbJmJQCqxUhwxKEyf5YMbFLsIZtBCiFKlcMagRXTGsQQsxSonKFYMb0RXDG7SQopShlKNUOMkHOqJ3"
    "hjpiDIMdlUeGO6pYatT1AO8TrtdUtdccemywJ3tcdB0+VutHnzI1rP/1DU26fYGDUmClBn1MYE2auFZe"
    "9LHdLUQfsBCi0UjliqOP5IqjD7iK0VWicsXRR3LF0QdcpegqQylHqXCSF30k7xx9pBiOPjZG9MjRRxPL"
    "0gcWeqraa5Y+9gRlv8gSHw67c7lBH9NjP/qUuXdLoI9J4asnNJTpqc07rsBKDfqYVg36WMmLPjZ2IfqA"
    "hRCNRk4SXXH0kVxx9AFXMbpKVK44+kiuOPqAqxRdZSjlKBVO8qKP5J2jjxTD0cfGeNFHE8vSBxZ6qtpr"
    "lj72bPSkj41u0Mdo/ehT5vQtgT4mNbBBH5CCoU0grCVUoTRxkhd97AgL0QdchegqUrni6CO54ugDrmJ0"
    "lahccfSRXHH0AVcpuspQylEqnORFH8k7Rx8phqOPjfGijyaWpQ8s9FS11yx97NnoSR8b3aCP0frRp8z7"
    "XAJ9TPpogz4gBUMjNa59QJq4Vl70sd0tRB+wEKLRSOWKo4/kiqMPuIrRVaJyxdFHcsXRB1yl6CpDKUep"
    "cJIXfSTvHH2kGI4+NsaLPppYlj6w0FPVXrP0sWejJ31sdIM+RutHnzLrcAn0scmLtVTy8tv39p0X5HDu"
    "21aNO6+X8jqFL7dcdwvRB1yF6CpyktedF5fQW86Eow+4itFVonLF0UdyxdEHXKXoKkMpR6lwkhd9JO8c"
    "faQYjj42xos+mliWPrDQU9Ves/SxZ6MnfWx0gz5G60efMrVwCfSxSYt1+oAUEGOqNzDrd14gTVwrr2sf"
    "KeOUST9GVyFKkcoVd+0jueLoAwsTo6tE5Yqjj+SKow+4StFVhlKOUuEkL/pI3jn6SDEcfWyMF300sSx9"
    "YKGnqr1m6WPPRk/62OgGfYzWjz5l3uES6GMzGuv0ASkoi42U73/X6QPSxLXyoo+UjsrRByyEaDRSueLo"
    "I7ni6AOuYnSVqFxx9JFccfQBVym6ylDKUSqc5EUfyTtHHymGo4+N8aKPJpalDyz0VLXXLH3s2ehJHxvd"
    "oI/R+tGnTEpcAn1sumOdPiAFQyM16APSxLXyoo/tbqE7L7AQotFI5Yqjj+SKow+4itFVonLF0UdyxdEH"
    "XKXoKkMpR6lwkhd9JO8cfaQYjj42xos+mliWPrDQU9Ves/SxZ6MnfWx0gz5G60cfWtdl0KfsppVtCFJQ"
    "vZvauvYxrRrPfazkRR8buxB9wEKIRs3r/eRddMXRR3LF0QdcxejKvKD/kiuOPpIrjj7gKkVXGUo5SuYl"
    "e/LuRR/JO0cfKYajj43xoo8mlqUPLLR5e/2lvWbpY89GT/rY6AZ9jNaLPhud2YZ0ii32ulbVDZTGaVX/"
    "2bOt6rVxUNpHaYLSAUqHKIUoRSgdoXSMUoxSgtIJSh9RSlHKUMpRKlA6RekMpXOULlD6hNIUpUuUqIyb"
    "ST5rlC90GlbNKYtxLuFDr+qm+aGHUmCl+iU3ShMnsR8vVV1aajaebJaPPMo6BrWqLDTKYH9jQ6qf40Zg"
    "CuhQzwfznmulWajXNeqdL6KDkwlRinTzo2bjaO6iVn3lkOZH3QrzO5LnRz8eH817rpdgoflR7+z8jnEy"
    "MUqJbn7UbJzMXdSqttAoA+pWmN+JPD/68fhk3nOtWgv1uka9s/P7iJNJUcpQylEqnMR+XFZHMTUbF84r"
    "5b/Pa79QlwMaSViFU3kV6Mfj03nPtSovdOy8KteXLbUj93tG/Z7N+61VdKEz4xUdPWW9HfpVKLijW5lz"
    "Gud8Pk6tggsZeGFlLnQjULPxxXyEGkHKEWj9hbX/hBs+1R321Gw8nQ9ag8tvtHam+s7GRll+x+ySWIDH"
    "jvhSBR4a8XI+Yg0kNE1Tg0ecKn2gdOQYWu3FMjw09u7efPAaay7pXCy7VlTi2Sgzivo/Faq6aX1E2fyl"
    "5wdFgW3V+IiCnKuJayXeAdGHUmdBHhfcfWPWfbdxiL5ClCKdr+53JY9kX913EsdoIkYp0fmyNx6t2i0n"
    "si/7hkGras1HNJGilKGUo1Q4SbzzobOwa7dPRfd0XVYFteuJykH21rpdqEdnE4rSVLWkL3TB3Uf0J1yz"
    "qWrHN+zB3q7WY4Pl98FMWh8UDiKodGQlWk2+Q9uwxzhTsKesF78MDtm0pdrT6arnBpoCKzU4BNlXE9fK"
    "j0NS6hllVnQdmYfoK0Qp0vniOCT64jiEmYnoK9H54jgk+uI4hLmJ6CtDKUepcJIfhyT3LIfEII5DmhRD"
    "OiG7cHmhmyPHIcxPVO04yyFNfiLPoY78RGunH4fKLKMlXA/ZBKY6h0AKaGna39GjNHGSH4ekJDSWQ+Ar"
    "RF+RzhfHIdEXxyHwFaOvROeL45Doi+MQ+ErRV4ZSjlLhJD8OSe5ZDolBHIdskGyT45AqmOMQLPZUteMs"
    "h+x5KT6x5jlko+tPrK2dfhwqH74tgUM2lanOIZACmh1wCKSJa+XHISkdjeUQmAjRaqTzxXFI9MVxCHzF"
    "6CvR+eI4JPriOAS+UvSVoZSjVDjJj0OSe5ZDYhDHIRvkxyFVMMchWOypasdZDtnz0pNDNrrBIaP141CZ"
    "ebQEDtmkpjqHQAroHhE4BNLEtfLjkJSYxnIITIRoNdL54jgk+uI4BL5i9JXofHEcEn1xHAJfKfrKUMpR"
    "KpzkxyHJPcshMYjjkA3y45AqmOMQLPZUteMsh+x56ckhG93gkNH6cajMQVoCh2x6U51DIAXlP2LYyp1G"
    "aeIkPw7ZIRZ7Tg2+QvQV6XxxHBJ9cRwCXzH6SnS+OA6JvjgOga8UfWUo5SgVTvLjkOSe5ZAYxHHIBvlx"
    "SBXMcQgWe6racZZD9rz05JCNbnDIaP04VGYjLYFDNtGpziGQgg0jNZ5TgzRxrfw4ZPtbjENgIkSrkc4X"
    "xyHRF8ch8BWjr0Tni+OQ6IvjEPhK0VeGUo5S4SQ/DknuWQ6JQRyHbJAfh1TBHIdgsaeqHWc5ZM9LTw7Z"
    "6AaHjNaLQ5vLqd5addP83h6lwEp1DqE0cZIXh1zwQhxCEyFKkc4XwyHZF8MhNBGjlOh8MRySfTEcQhMp"
    "ShlKOUqFk7w4JLrnOCQHMRzS2WSeU+uCGQ7hmk1VO85xyAZ7fm/vouscslo/DnVmVtND8cUyqzdtmcja"
    "9RBKgZUaHIJ6lRPXyo9DUrFO7vkQ+gpRinS+OA6JvjgOweLE6CvR+eI4JPriOAS+UvSVoZSjVDjJj0OS"
    "e5ZDYhDHIU1RVu57e90cOQ5hRVfVjrMc0lR0Zb8vc2d1g0Oaiq4v5A9tLifVvuqmdT1kMzSf0RTYVg0O"
    "QeXKiWvlxyGpbCfLITARotVI54vjkOiL4xD4itFXovPFcUj0xXEIfKXoK0MpR6lwkh+HJPcsh8QgjkM2"
    "yOu+TDdHjkOw2FPVjrMcsuel332ZO9EbHDI99rseWk4+9SYkT++hFFipwSHMp3at/DgkFfBkOQQmQrQa"
    "6XxxHBJ9cRwCXzH6SnS+OA6JvjgOga8UfWUo5SgVTvLjkOSe5ZAYxHHIBvlxSBXMcQgWe6racZZD9lT1"
    "5FBHPrW1049Dy8mn3sR8apQCKzU4hPnUrpUfh3zyqdFXiFKk88VxyCefGk3EKCU6XxyHfPKp0USKUoZS"
    "jlLhJD8O+eRTuxE7HyHSO1VdKffnOpvc8yFVMjbHIcynVu04y6Fe+dR2aHOXVb21Qq+WmR77cajMj+z/"
    "fVnpoVX3A6XASg0OQbboxLXy45CYH8y814G+QpQinS+OQz751GgiRinR+eI45JNPjSZSlDKUcpQKJ/lx"
    "yCef2o24GIf65FPr5shxCPOpVTvOcqhXPrUduskh02P5I/7fqHzp+VCZC7kEDtk0y/pzapCC6p3cZv0h"
    "lEx9hRfrrDDvubr+Fvu+DPOp0Zepi/CiL45DPvnUaCJGydQzeNEXxyGffGo0kaKUoZSjZOoQvFSFiHvP"
    "Vdxt9r7MJ5/ajeR3X9YnnxrXzLzK/2ItLOY9V9uf7/dlHfnUtsd+HCpzkIBDJC74fZlJZXpDT85LJK7u"
    "vP+5Mxy+fT/4ufN+cL3zvrqC29s0reqViFDaR2mC0gFKhyiFKEUoHaF0jFKMUoLSCUofUUpRylDKUSpQ"
    "OkXpDKVzlC5Q+oTSFKVLlHZ3O7T5fmMlos0y5WQJH382m6X+8QdSUA3W+viDpJyJa8VehptKRJvvxpNR"
    "+ZCEKhFRPuG8hguNMtjffCdVInIjMJWIqOeDec+1SjbU6xr1zlciwvmFKEW6+VGzceRc0Bfy8/kd0vyo"
    "W6kSkTy/I+r5aN5zvUYNzY965ysR4WRilBLd/KjZOJm7qFW0Oab5UbdSJSJ5fifU88m851oNG+p1jXrn"
    "KxHhZFKUMpRylAonsZ+XphIReS3mXmvVcKjLAY0kVSKSV+GUej6d91wreUPHzqtyfdlKRHK/Z9Tv2bzf"
    "WmEbOjNe0dFTViKiX4VKRLqVOadxzufj1GrYkIEXVuZCNwI1G1/MR6gRpByB1l+qRIQbPtUd9tRsPJ0P"
    "WoPLb7R2phLR5ruyEpHZJbESkR3xpUpENOKlG3GrBhKapqlEJE6VPlA6MhWtxl/xVEc3PSl6R5WI5oPX"
    "WHNJ52LZtaISEXlexkdU1U3zG3yUAivVnxShNHGS+KSIPpQ6KxG54O7HAN3fBB2iiRClSOWLPky6fB2J"
    "vig9o+sx6TGaiFFKdL66i9eeyL66q9d+RBMpShlKOUqFk8RbH3om2lmJSHbf/VzwTA6yd9jtSkQ6m/Yb"
    "0FbxnwtdcPcR/QnXbKrb8e5/qPrSBst3aNSoa8F3d110/Rt8q8l3aFv2GGcqEdGPl8IhzGisem5WIrJS"
    "g0OY0eha+XFITFfjOAQmQrQaqXyxHJJ8sRzCjEb0leh8cRwSfXEcwoxG9JWhlKNUOMmPQ6J7jkNiEMch"
    "TUbjJschVTDHIVjsqW7HOQ7ZU1X8Bp/nkI1ucMho/Ti0nIzGLUhf3EMpsFKDQ5jR6Fr5cUhMV+M4hBmN"
    "aDVS+WI5JPliOYQZjegr0fniOCT64jiEGY3oK0MpR6lwkh+HRPcch8QgjkOajEaWQ6pgjkOY0ajbcY5D"
    "moxGnkM2usEho/XjUPnwrf+jwy2b6FR7dIhSYKUGhyBfa+Ja+XFITFfjOAQmQrQaqXyxHJJ8sRwCXzH6"
    "SnS+OA6JvjgOga8UfWUo5SgVTvLjkOie45AYxHFIk5TIckgVzHEIFnuq23GOQ/ZU9bwestENDhmtH4eW"
    "k9G4hRmNKAVWanAIMxpdKz8OiTluHIfARIhWI5UvlkOSL5ZD4CtGX4nOF8ch0RfHIfCVoq8MpRylwkl+"
    "HBLdcxwSgzgOaZISWQ6pgjkOYUajbsc5DmkyGvnrIRvd4JDR+nFoORmNW5jRiFJgpQaHMKPRtfLjkJjj"
    "xnEITIRoNVL5Yjkk+WI5BL5i9JXofHEcEn1xHAJfKfrKUMpRKpzkxyHRPcchMYjjkCajkeWQKpjjECz2"
    "VLfjHIfsqep5PWSjGxwyWj8OLSejcQszGlEKrNTgECQTTlwrPw6J6WochzCjEa1GKl8shyRfLIfAV4y+"
    "Ep0vjkOiL45D4CtFXxlKOUqFk/w4JLrnOCQGcRzSJCWyHFIFcxyCxZ7qdpzjkD1VPTnUkdFo7fTjUFnd"
    "cQnPh2zhyPrzIZCCLSh2uY/SxEl+HBLLf3IcAl8h+opUvlgOSb5YDoGvGH0lOl8ch0RfHIfAV4q+MpRy"
    "lAon+XFIdM9xSAziOGSDZJvc92WqYI5DsNhT3Y5zHLLnpSeHbHTjesho/TjUmVlNF1qLZVZvmZxpU8DE"
    "pFGjFFipcT1kAmvSxLXy45Dtb7H8ITARotVI5YvlkOSL5RD4itFXovPFcUj0xXEIfKXoK0MpR6lwkh+H"
    "RPcch8QgjkM2yI9DqmCOQ7DYU92Ocxyyp6onh2x0g0NG68ehMo9yCddDkFe/twVSYKUGh0yrBoes5Mch"
    "G7wYh8BEiFYjJ4m+WA5JvlgOga8YfSU6XxyHRF8ch8BXir4ylHKUCif5cUh0z3FIDOI4ZIP8OKQK5jgE"
    "iz3V7TjHIXteenLIRjc4ZLReHBotJ5+66qaZT41SYKU6h1CaOMmLQy54IQ6hiRClSOWL45Doi+MQmohR"
    "SnS+GA7JvhgOoYkUpQylHKXCSV4ckt0zHJKDGA7pbDL3ZbpghkO4ZlPdjjMcssGe+dQuus4hq/XjUJkL"
    "2f96aGTTLGvPh1AKrNTgEGSLTlwrPw5JybJUabzzX3ZFXyFKkcoXyyGffGo0EaOU6HxxHPLJp0YTKUoZ"
    "SjlKhZP8OOSTT+1G7PyoorfIOysR6WxyHOqTT41rNtXtOMehXvnUduhGBRCr9eNQmQu5BA7ZNMs6h0AK"
    "RpAauo/SxEl+HJKSZVkOga8QfUUqXyyHxCRe5v0yNBGjlOh8cRwSfXHXQ7BeKfrKUMpRKpzkxyHRPXc9"
    "JAZxHFKlRHMcUgVz10Ow2FPdjnMcsuel332ZHbrJoSXkU1evyS6BQ5hPXfXcfL/MSo3rIcyndq38OOST"
    "T42+QpRMXYSX6r+wHPLJp0YTMUqmnsGLvjgO+eRTo4kUpQylHCVTh+ClSkTce66uv+5LG45DPvnUbiSv"
    "50O6YI5DmE+tOkPobfmuV1XNW/l0XnpyqCOf2trpdz1U5iD1r0Q0MqlMrUpE71qViGyreiUilPZRmqB0"
    "gNIhSiFKEUpHKB2jFKOUoHSC0keUUpQylHKUCpROUTpD6RylC5Q+oTRF6RKl3d0Obc9pWIlotJy0taqb"
    "1mNJyGQLbKvGxx+mrblW7MefqUQ02h5P3pQPVakSEb3n9VyJaLQ92B9tS5WI3AhMJSLq+WDec62SDfW6"
    "RnPgKxHh/EKUIt38qNk4mruo1ag5pPlRt1IlInl+R9Tz0bzneo0amh/1zlciwsnEKCW6+VGzcTJ3Uato"
    "c0zzo26lSkTy/E6o55N5z7UaNtTrGvXOVyLCyaQoZSjlKBVOYj8vTSUi8lrMvdaq4VCXAxpJqkQkr8Ip"
    "9Xw677lW8oaOnVfl+rKViOR+z6jfM9cvPZCsn3ev6OgpKxHRr0IlIt3KnNM45/NxajVsyMALK3OhG4Ga"
    "jS/mI9QIUo5A6y9VIsINn+oOe2o2ns4HrcHlN1o7U4lotF1WIjK7JFYisiO+VImIRrycj1gDCU3TVCIS"
    "p0ofKB2ZilZ7sRIRjb27Nx+8xppLOhfLrhWViEZlNtIS7tBsolP9SRFIQTVYs1geShMniXdo9KHUWYnI"
    "BXdes9Obgt1PrDGjEX1FOl/dNXOOZF/d9xLHaCJGKdH5sk9AWkVyTmRf9slHqyzPRzSRopShlKNUOEm8"
    "9aHUmc5KRLL77u8nzsQggm73E2tNUiI9uu2yeaGaIz2D6Qr+hGs2Ve043Td136FpMhrNzcvPndbhQlDp"
    "yGi0mnyHRlfJLTuDh/tfO+/pl5WHD6ujMhtpCRyyiU51DoEUVIO1OAT5WhPXyo9Dtr/FOAQmQrQa6Xxx"
    "HBJ9cRwCXzH6SnS+OA6JvjgOga8UfWUo5SgVTvLjkOie45AUxHLIBok2WQ6pgjkOwWJPVTvOcsiel+KT"
    "Ip5DNrrxDb7R+nGozEZaAodsolOdQyAFI0jO2kdp4iQ/DtkhFuMQ+ArRV6TzxXFI9MVxCHzF6CvR+eI4"
    "JPriOAS+UvSVoZSjVDjJj0Oie45DUhDLIRvkxyFVMMchWOypasdZDtnz0pNDNrrBIaP14tD2cjIaq26a"
    "jw5RCqxUf3SI0sRJXhxywQtxCE2EKEU6XwyHZF8Mh9BEjFKi88VwSPbFcAhNpChlKOUoFU7y4pDsnuGQ"
    "GMRxSGWTux7SBTMcwjWbqnac45ANlr85Y6+HXHSdQ1brx6HlZDRuY0YjSoGVGhzCjEbXyo9DUo4b93wI"
    "fYUoRTpfHIdEXxyHYHFi9JXofHEcEn1xHAJfKfrKUMpRKpzkxyHRPcchKYjlkCYpkeWQKpjjECz2VLXj"
    "LIc0GY08h2x0g0NG68eh5WQ0bmNGI0qBlRocgnytiWvlxyEpXY3lEGY0otVI54vjkOiL4xD4itFXovPF"
    "cUj0xXEIMxrRV4ZSjlLhJD8Oie45DklBLIc0SYksh1TBHIcwo1G14yyHNBmNPIdsdINDRuvHoTJDqf/z"
    "oW3MaEQpsFKDQ5jR6Fr5cUhKV2M5BCZCtBrpfHEcEn1xHAJfMfpKdL44Dom+OA6BrxR9ZSjlKBVO8uOQ"
    "6J7jkBTEcsgGeT0fUs2R+74M12yq2nGWQ/ZU9Xs+5M7qBodMj/04VFZ3XAKHbOHI2nPqbZACKzU4BPUv"
    "J66VH4ek8p8sh8BEiFYjnS+OQ6IvjkPgK0Zfic4XxyHRF8ch8JWirwylHKXCSX4cEt1zHJKCWA7ZID8O"
    "qYK56yFY7Klqx1kO2fPSk0M2usEho/XjUGdm9cKViLZt4cg6h0AKbKsGh6D+5cS18uOQVP6T5RCYCNFq"
    "pPPFcUj0xXEIfMXoK9H54jgk+uI4BL5S9JWhlKNUOMmPQ6J7jkNSEMshG+THIVUwxyFY7Klqx1kO2fPS"
    "k0M2usEho/Xj0HJS7bchr34PpcBKDQ5hqr1r5cchqfwnyyEwEaLVSOeL45Doi+MQ+IrRV6LzxXFI9MVx"
    "CHyl6CtDKUepcJIfh0T3HIekIJZDNsiPQ6pgjkOw2FPVjrMcsqeqJ4c68q6tnX4cWk4+9TbmU6MUWKnB"
    "IUhlnrhWfhyy/S32vT3mU6PVSOeL45Doi+MQ+IrRV6LzxXFI9MVxCHyl6CtDKUepcJIfh0T3HIekIJZD"
    "ffKpVXNknw/BYk9VO85yqFc+tTurG9dDpsd+HCpzIZfwfMimWdbvy0AKtiE1dB+liZP8OGSHWIxD4CtE"
    "X5HOF8ch0RfHIfAVo69E54vjkOiL4xD4StFXhlKOUuEkPw6J7jkOSUEsh2yQ3/WQKpi7HoLFnqp2nOWQ"
    "PS89r4dsdINDRuvHoTIXcgkcMimV9UrV2yAFVmpcD5lW9QqxrpUfh2x/i3EITIRoNdL54jgk+uI4BL5i"
    "9JXofHEcEn1xHAJfKfrKUMpRKpzkxyHRPcchKYjlkA3y45AqmOMQLPZUteMsh+x56ckhG93gkNF6cah6"
    "TbY/h6pumvnUKAVWqnMIJVNf4aXKOtx7rq6/hTiEJkKUTF2EF30xHJJ9MRxCEzFKpp7Bi74YDsm+GA6h"
    "iRSlDKUcJVOH4KVKRNx7rrJ7hkNiEMchF+TFIV0wwyFcM/Mq/0s7znHI9ueZT+2i6xyymh+HBo/fZrOn"
    "/aunq533t7OHr7Ng9v3748r1/V93TzRFqpDwLK9U5Suo2Ay932/em51HmB9lG+PTahlb+u4WRVRnZusH"
    "5xvjT50BG0OK2KjeW273tfmWfrRZvXnd/tHGeK+zt7Kzrr4ON8ZHXQHHG+OTLn26UboqXxhujRxsjA8q"
    "ffC8hjvvf1x9nSVXD19v7h5Xvs++0Hquv35DXyo83Hz9Nv/L0/0PWufVlT/un57ub6s/fptdfZ7Ry8fr"
    "r6nxl/v7J/cXGnjw6/7hz2rPdv4fAAD//wMAUEsDBBQABgAIAAAAIQA28DM9JBIAAJJ5AAAYAAAAeGwv"
    "d29ya3NoZWV0cy9zaGVldDQueG1srJ1bU+NIEoXfN2L/A+GnmfF2Y124RsPEDLR1v5phn91gGscAJmz3"
    "bTf2v2+WqgpUOlVl5HbHTGO+zkqp5JNSouNCH37//viw93W2XM0XT2cD5/1osDd7ulnczp8+nw3+uhq/"
    "Ox7srdbTp9vpw+Jpdjb4MVsNfj//5z8+fFss/17dz2brPcrwtDob3K/Xz6f7+6ub+9njdPV+8Tx7on+5"
    "Wywfp2v6dvl5f/W8nE1vm0GPD/vuaHS4/zidPw14htPlW3Is7u7mN7PLxc2Xx9nTmidZzh6ma9r/1f38"
    "eSWzPd68Jd3jdPn3l+d3N4vHZ0rxaf4wX/9okg72Hm9Oo89Pi+X00wPN+7vjT2/2vi/pP5f+9+RmGg5b"
    "epzfLBerxd36PWXe5/uM0z/ZP9mf3rxkwvm/KY3j7y9nX+fsDXxN5W63S87BSy73NZm3ZbLDl2TscC1P"
    "v8xvzwb/HYk/7+irw/4avRt57K/Wn/8Nzj/czukdZrPaW87uzgZ/OKfX/miwf/6hEdD1fPZt1Xq9t7pf"
    "fAuW89t0/jQjNZKOmUI/LRZ/s9CItjyipM/Tp9ne98kzvc9nA9qrH68v14vndHa3vpg9PJwNLv3B3vRm"
    "Pf86K2nE2eDTYr1ePNbzz/frph7WxO6Wi//Mnpo9mj3MKJh2leU/G1AqHsr2tvNvPBPbkm5kezs0eP9l"
    "ru3Xct7jprbK5d7t7G765WFdL76FM7aLVMj+e5pBI87T2x+Xs9UNVQsdgvce2+rN4oFy0N97j3NW9ST2"
    "6ffm67f57fqeXh0O9m6+rGjK/xZADOMDSBvNAPoqBtBZwhJPB5rHvw440g3Y5zvWTPpyup6ef1guvu1R"
    "XdAerujQ0vvqnLqUhE3RpQnybb5M2jBjmipL8gfLcjY4pE3T8BWp4eu5444+7H+lg3sjgv7kQQd+M2E2"
    "7ALIJZCPQMZAAiAhkAhIDCQBkgLJgORACiAlkApIDWQC5ArIX0CuBTlgx3mf3uaX95reXnivHYL93muW"
    "5WxwoLzVTuet1sW4asyFLsZTYy55zBGpvKUrXw36KBK96moMJAASAomAxEASICmQDEgOpABSAqkEad7H"
    "psxqTg7plPByOA47h34Co67EKOUYHqjH8C9dzKEac62LOXqJUZRGbxgozR+9H/knrT8vJ6q3nmVY1uaM"
    "1RJDV3m6mK7ydDFd5fGY9pF23OOO8HjMiXJgT9SYsSaP1zkxBpo8XmdeoS5PZ16RLk9nXrEuT6egEl2e"
    "jmBSXZ6OYDJdnlfBNJLOdXk6x7nQ5ekc51KTx+8c50qTx+8c51qjDb9znCe6mM5xvhIxdCVtibVbddqg"
    "btlpgwx1R5vDMzydqvud4VkW6uGoU3ndd78jkj9FUPuSf9DZ9QtNjON3DgLrB1lVH760BR+BjIEEQEIg"
    "EZAYSAIkBZIByYEUQEogFZBakCPqmO6ag30/Xc5uB7wvr/3T2mUX83nTYV/6w7E/DP1h7A9Tf5j7w5Le"
    "kzvWa3WEPrGmnfinE5nWGZx/9IeBP4z8YeIPM39Y+MPKkPZKpBUnxfVSu9dX/umVTO8OzqPxx7ou6l9q"
    "f5/+H078X/81eDf4Vez4u85lSLeFfimuRQpxVuY72Urxm2vfD+UyRpWzg3JiWTaWkwiylpMmxvE7NXfJ"
    "g9rlBGQMJAASAomAxEASICmQDEgOpABSAqmA1IJAOfEKonYYS8U2xGGnJBhyJYYYy4BtzeVDu/rWDX0V"
    "5wFJ82A4ObCVyLVIYdD3wW+ePYmibzrD70DfLMtGfYsgq741MY7faRgueVBb30DGQAIgIZAISAwkAZIC"
    "yYDkQAogJZAKSC1IH33bhhj0LYZso2/d0Bd9/1IfDp33B7/Vh7+SRg+Hk0Or0EUug9APmxO5JYkidJLd"
    "DoTOsmwUughShN7R8IUmBvsiHtQWOpAxkABICCQCEgNJgKRAMiA5kAJICaQCUgvSR+i2IQahiyHbCF03"
    "9PVEfkT6PhpOjqz6FikM+j5q9G1JouibWQndu3gOXSn69f0sy0Z9iyDa+5cfDg66+tbEYKPCg9r6BjIG"
    "EgAJgURAYiAJkBRIBiQHUgApgVRAakH66Ns2xKBvMYS+sLcK+3VLo6Ib+qrvY9L38XBybNW3SGHQ93HT"
    "qFiSKPqmJDvQN8uyUd8iyKpvTQw2KjyorW8gYyABkBBIBCQGkgBJgWRAciAFkBJIBaQWpI++bUMM+hZD"
    "ttG3bmirUTnhjcoJa1ROhpMTq9BFLvryWmavtXLSnMgtSRShO+Su7UDpTZqNUpdRitY7N98udEHYrIio"
    "ttoRjREFiEJEEaIYUYIoRZQhyhEViEpEFaJaoj7at44xiF+O2Ub92rGvknVGJHtnNJw4I6vwZRqD8p1R"
    "I31rIlX87GbXz7cxDjfANty/lFF28YtU7SDsZEQqRfx8YAuNMSpAFCKKEMWIEkQpogxRjqhAVCKqENUS"
    "9RK/OFTaMSbxizFbiV83tiV+h4nfIfE7dvGLNCbxO02DY02kip/ZazsQP3fpNolfRNnFrwnCNocZyJ37"
    "94jGiAJEIaIIUYwoQZQiyhDliApEJaIKUS1RL/GLo9dL/GLMVuLXjW01Po7LOx/HZa0PfTehV5Y79tdy"
    "0qYycPk1wJZILQNmd+2gDLhrtqkMRJRSBh1n8cLRBGkaIGHTtWwsMVC5BkBUgFEhoghRjChBlCLKEOWI"
    "CkQlogpRLVGvMhDHpVcZiDFblYFubOsawG6iOx6J37OLX6Qxid/j4rclUsXPzKsdiJ97YJvEL6Ls4tcE"
    "aRogcDY/OoDGiAJEIaIIUYwoQZQiyhDliApEJaIKUS1RL/Gb7Fp2f8bUAG3yYi33duQ+KoXTEj9zah2y"
    "ah2rV3st05jE7/MGyJZIFT9ztnYgfm6QbRK/iLKLXxOkaYDAh/zIbr2qPdEYUYAoRBQhihEliFJEGaIc"
    "UYGoRFQhqiXqJf4tDFi5na3O/FYLloR/IBqgg6YBIjvWsfuxcmdMZXDArwG2RGoZMN9rB2XA7bNNZSCi"
    "2mVw2Pk0yQX7yDK7b6r+ENz9GI+IUn4IBgdzjFEBohBRhChGlCBKEWWIckQFohJRhaiWqFcZbOHTyu1s"
    "VQZWp5amwK4B5K46do9W7oJJ/NyltSZSxc9MsR2In3trm8Qvouzi1wRpGiAepYgf0NgBFCAKEUWIYkQJ"
    "ohRRhihHVCAqEVWIaol6iV8cl17dvxizlfh1Y1sNELNvHbJeHbuBK6dqEv8Rb4BsiVTxM8dsB+Lnxtsm"
    "8Ysou/g1QZoGSDh97R99AY0dQAGiEFGEKEaUIEoRZYhyRAWiElGFqJaol/jFceklfjFmK/HrxrbvAB2L"
    "Bui4aYDIoXXsPq+ctKkMjnkDZEuklgHz03ZQBtyW21QGIkopg85nwi8cTZDmDpDwAdtlAGgscrWuFAGi"
    "EFGEKEaUIEoRZYhyRAWiElGFqJaoVxmIQ9WrDMSYrcpAN7Z1DThh1wBybR279yunahI/d3+tiRTxu7vx"
    "f5s0G/1fGWUVvy4IGyAR1W6AEI0RBYhCRBGiGFGCKEWUIcoRFYhKRBWiWqI+4reOMdwBkmO2Eb927Kv4"
    "Xeb/uuT/unb/V6YxiN8dNQ2QNZEq/t34v83qg83i11i73cV0FzKV+qNv9yPJIkoRP/q/GBUgChFFiGJE"
    "CaIUUYYoR1QgKhFViPhyFbpp2Uv8W/i/ctNbid/q/5LwHd4AuQ5rgOg7Wi1jdQHkzpjKwGkaIGsitQx2"
    "Y4G5wrcSx4gvyvY7H+/5U0aJnW+iYB2XLggbIBGllAH4XWOMChCFiCJEMaIEUYooQ5QjKhCViCpEtURQ"
    "Bs3vW6B/Pq3pFy40C7u8wfml6w3psAxp0kOa0pB2eEi7M6SNGdZ32fNPKP9E5vdphRflomM8pCM4pOMz"
    "pNkPaW5D2nN9/iuZ31hezTwo7PRKbuegtdTLZfYZbYB2xLraS7uZ1mXobWmuZRpT/XEXzro/av3txoVj"
    "v0WB3TbdUH8iyl5/miBND4YunNiHVkmOEQWIQkQRohhRgihFlCHKERWISkQVoloi7WWo+T0duBDMOqb5"
    "3QK4EkyOsV6GmvX7X8+7S8G0Y1viZy6cS+aZa3fhZBqT+LkLZ02kin83LpwrrLMN4hdRdvFrgvAmlNig"
    "cvFBFw6jAkQhoghRjChBlCLKEOWICkQlogpRLVEv8dtcOJP4dU5aexEzs6BN4t/gwrnChXMbF46+m9Ar"
    "28eQ5KRNZcBdOGsitQx248K5wjrbUAYiSimD7poamaodpOnBwHL7KAYq1wCICjAqRBQhihEliFJEGaIc"
    "UYGoRFQhqiXqVQY2F85UBm9ZLmkqA7sL5zIXziUXzrW7cHKqJvFzF86aSBX/blw4V1hnG8Qvouzi1wRp"
    "GiB04cQ+KOJHFw6jQkQRohhRgihFlCHKERWISkQVolqiXuK3uXAm8b/FhTOJ3+7CucyFc8k8c+0unJyq"
    "SfzchbMmUsXP/JGftx9cbrNs6v5FlF38miBNAyR8nZb9IPZBET9EBRgVIooQxYgSRCmiDFGOqEBUIqoQ"
    "1RL1Er84LtoxJvGLMVt1/7qxLRfOFS6c27hw9N2EXlkbIJHQVAbchbMmUsuAEu2iDFiazT8EiyilDLoL"
    "0VxNkKYB4lHKzwGAxiJX24VDFCKKEMWIEkQpogxRjqhAVCKqENUS9SoDcah6lYEYs1UZ6Ma2fghmLpxL"
    "Lpxrd+HkVE3i5y6cNZEifm83LlyTZqP4ZZRV/LogbIBEVFv8iMaIAkQhoghRjChBlCLKEOWICkQlogpR"
    "LVEf8VvHGK4Bcsw24teOfRW/x1w4j1w4z+7CyTQG8XvchbMmUsW/GxfOE/6avfuXUXbxi1TtIGyARCpF"
    "/OjCYVSAKEQUIYoRJYhSRBmiHFGBqERUIaol6iV+mwtnEv9bVmEaun+5j0rhtBogT7hwXuPC0XcTemVr"
    "gGRCUxlwF86aSC2D3azH9MQiyg1lIKKUMuguRJOp1DLofg5bRCllAEs0xxgVIAoRRYhiRAmiFFGGKEdU"
    "ICoRVYhqiXqVgW09pqkM3rIe01QG1vWYNAV2DaDFk559Faacqkn8fBWmNZEq/t1Y0N6bLGgZZRe/SKWK"
    "v/vLD0UqRfxoQWNUgChEFCGKESWIUkQZohxRgahEVCGqJeolftsqTJP437IK0yR++ypMj/m/HtnInn0V"
    "ppyqSfxe8zEkayJV/Lvxf703+b8yyi5+rf/b/RiSSKWIH1dhYlSAKEQUIYoRJYhSRBmiHFGBqERUIaol"
    "6iV+2ypMk/jfsgrTJH7d2HYD5POPIXk++xiSR04wvbI2QCKhqQz85mNI1kRqGezGCfbe5ATLqHYZwEI0"
    "XRDeARJRShmgE4xRAaIQUYQoRpQgShFliHJEBaISUYWolqhXGWzhBMvtbPVDsNUJpikw8ZP/69n9X7kL"
    "JvFz/9eaSBX/bvxftvpg8+1PGWUXv8Yk1twBQv9XZG+7AIgCRCGiCFGMKEGUIsoQ5YgKRCWiClEtUS/x"
    "b+H/yu1sJX67/+sx/9cj/9ez+79yF0ziP+QNkC2RKv7d+L9swfAbxK+xdvHMr/V/oQFC/1fsgyJ+9H8x"
    "KkQUIYoRJYhSRBmiHFGBqERUIaol6iX+LfxfuZ2txG/1f0n4R6IBOmoaIHKCPbsTLHfGVAb8V+pSWnMi"
    "tQyYpfbzTjCzgN9QBiJKuQZ0F6LJVBvuAAkrsOUEi4FKGUBUgFEhoghRjChBlCLKEOWICkQlogpRLVGv"
    "MhDHpY8FJrezVRmI7elvhNIU2DWA/F/P7v/KXTCJn/u/1kSq+CnRLsTP0my2wESUXfyaIE0DxKOU7h/Q"
    "2AMUIAoRRYhiRAmiFFGGKEdUICoRVYhqiXqJXxyXXuIXY7YSv25sywJj/q9H/q9n93/lVE3iP+ENkC2R"
    "Iv5mUcPPn/mbNBvFL6Os4tcFoQUmotriRzRGFCAKEUWIYkQJohRRhihHVCAqEVWI+PIa00I0wwoAmaaP"
    "+OWYbcSvHdu6A+SPeAPkj1gDRN/Roh7rHSCZ0FAGPv99vNZEvAz4w1H5c0IfZ8vPM/bo1tXezeILe9Qp"
    "LVn68EL5o6ou3dOqeeRNh7NHyzbrVzu8dukZVM3zKV/Ts0fIfp5l0+Xn+dNq74Ee48qeqnpE9weW/Mmr"
    "/Bt6AmzzDED+PNfm5T09+3hGzxMdvafgu8ViLb9hD8B8eZry+f8BAAD//wMAUEsDBBQABgAIAAAAIQAr"
    "ixNtTAoAAD4+AAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDUueG1srJtbU+M6EoDft2r/Q8pPu+ccSGxM"
    "gBTh1JBgiB3YU+zt2ZM44JokTtnmMru1/31bUssXdSQcO9QMCe3ullr6JMtq6/r3z8269x6lWZxsx5Z9"
    "OrB60XaRLOPty9j65z+8k0url+Xhdhmuk200tn5GmfX7zZ//dP2RpD+y1yjKe+Bhm42t1zzfjfr9bPEa"
    "bcLsNNlFW7iyStJNmMOf6Us/26VRuORGm3XfGQyG/U0Yby3hYZQ28ZGsVvEimiaLt020zYWTNFqHOdQ/"
    "e413mfS2WTRxtwnTH2+7k0Wy2YGL7/E6zn9yp1ZvsxjNXrZJGn5fQ9yfthsuep8p/HPg/5kshstJSZt4"
    "kSZZsspPwXNf1JmGf9W/6oeLwhONv5Eb2+2n0XvMOrB05bSrkn1e+HJKZ2ctnQ0LZ6y50tFbvBxb/x3g"
    "zwl82uzX4GTgsl+Vn/9ZN9fLGHqYRdVLo9XY+maPnhzH6t9cc4D+FUcfWeV7L3tNPu7TeDmPtxHQCBwz"
    "Qr8nyQ+mOoOSB+B0F26j3uffd9DPYwtq9bP8mie7ebTKJ9F6PbamrtULF3n8Hv0BFmPre5LnyeY5fnnN"
    "+XjIQbZKk/9EW16jaB2BMlSV+R9b4Eqostoq14QnVtI+y2o5YNwvYq1+l3F7fGz9kfaW0Sp8W+fPycdD"
    "xKoIA9k9hQg4nKPlz2mULWC0QBOcnrFSF8kafMDv3iZmox5gDz/550e8zF/H1rnVW7xlEPG/xd82Wgl9"
    "QIPrwyfqQ8eYDKChuQF8ooFtNoC6c4NhYXBhLACucn34xALA0hABTGpcHz5R3+z/CvXhE/XB0uDfBvxE"
    "kzIORRuCrcmi6ASI3VSnvug8DsY0zMOb6zT56MHcAQ4ywA/Yt0c2OGEYOPApCi3A0FABODAn35iXsTWE"
    "5gDzDEbM+43tXl333wHABSrdCqVzl1PBzCZEMiWSOyLxiOSeSB6IZEYkPpEERDInkkcieULJOYurD81a"
    "tC3QStq2RLhp2zIv0LaATtm25wOlbYVStW1RwmvFO2kqJBcwSiud5NYd3RFHHpHcE8kDKWyGhcGAqhR2"
    "WS/MJ44CIpkTySORPFWLr/UATB2kB9zh6aCYqZp2AnPEB0ulD2ylD/bpOHWdyT6ds7rOVOjU+1vtpn06"
    "53U/3j6dYV3nfo+Oo/h52Ofnou5nJnRgsqq0j9rb+3SUOSLYozNUWJ/v01H64nGfjtIXT9gXMNmVlR6W"
    "nVEDCZToUG4DEnM0tly2qoA5UiEIL8KEUqmR0vWTwkNVSekzthBhvA6LufaOSDyUXEBVVtzXa5hGS0ss"
    "nDx35LH7cxbzJdDMu3t+/tvzX6bur3fub4O/XvdXLACla+7RJbKbp3td37uje+natm5K133P/c06saTz"
    "k3rrPKBzhEzj/MEdPdhwH2L1dqrOf3HM7mekgXwiCYxNFrijQMZ1VhY9c3/19U02b9Rkc3c0l67dqut+"
    "YGqyx0ZN9uiOHmWTnVedQ5MZ3T8hioAJ41XTI08uX4LzHhmW7u/dk7mm7rWRB1P1cUYec1SMPGUuuMWL"
    "9ZGnTJaTwoNp5Aml6sgjEg8lZOSJwQbLCDqy0EQ7spgpX2+/36jjBk2140YMFV6qYjojVfeJJDAFw58b"
    "SDDzJsHwZSIJ5rFJMHzNQ0yfsPu0tLJ24BNm1bTGIqxmjsMicwQswiwBXa3c/2/xIgyuyl1AueFOCg9V"
    "JeWOOxVKVRaJxEPJISyiSRsW0bQNi6TqPpEEpmA0LDYJRsNik2A0LGL3dWCR3ePUBze7zYqEOSpYVBYb"
    "t3ixzqKyaJsUHkwsCqUqi0TioeQQFtGkDYto2oZFUnWfSAJTMBoWmwSjYbFJMBoWsfs6sMh2e4/CInNU"
    "sKisaW/xYo3FC2UFOik8mFgUSlUWicRDySEsokkbFtG0DYuk6j6RBKZgNCw2CUbDYpNgNCxi93VgEZrw"
    "OCwyR8AirIjhHq0sBW/xYm1740J5nJsUHiosXigrz6lQqrJIJB5KDmERTdqwiKZtWCRV94kkMAWjYbFJ"
    "MBoWmwSjYRG7rwOLbEP3ODByTwWNymLwVl6t46isKyelDxOPqFUFkoo8KToESWnThklp2wZKWn2figJj"
    "RBouG0WkAbNRRBoyZU92QZPtWh/lnm2L/W+cKJVHkFt5tY6muqdV+jCiKUqqoUlEHvqCHQ1lX8vwdC1t"
    "WqGJVWiFJqm+L6tSbt0Fxoh0aKJnY0Q6NJtEpEMTaeiCJtt2PA6aIjGAaJJ8FF6to6msOicsYSPWAUY0"
    "hVYNTSLy0NdhaKKbVmiibSs0SfV9Wf0qmqi1d7Dp0GwSkQ7NJhHp0MSe7ILmvoRSq6duW2QfEE1bzefJ"
    "y3U21Q3J0omRTUx0VJIBaFjB1ZOig6ZN9NyKTbRtxSaJyKcRBcaIdGw2iUjHZpOIdGwiDl3YZLvxx5k2"
    "cV9fPPrYapaKvSbAZsQ6m+oGpdSCXddyG5M8/aBWbd4kaRdPah3E5leJFcOOuSyvFZs0j0SDDIwR6dhs"
    "EpGOza8yQaw1dGx+leb5euPcPloWh3sqHoRsNY8jL9fZVBalk9KJkU2aykHD2rzZIpkj3bSaNzukc2j1"
    "fSoKpOige3qHlI4szzjadGx2T+pAZx5r3sRdfZw31bwOL4jMm+pmutT6Yt4kuY87NKyxacqHDNi7ATTR"
    "KN20YrNJRoS/jUmyczNafZ+KAik6iM0OKR5ZXis2uyd5eFr+OPd03OVHNtU8Dy9IZfNS3VyXWl+wSVM9"
    "aFhjs0WyR7ppxWaHdA+tvk9FgRQdxGaHlI8srxWb3ZM+NturPw6buOvPGg5eilLzPrwgYBN7XShdqpvt"
    "UgsCK9ebl+puO2rV1puYrigfjzypddB6s0nCRPOGhiyv1pPli1H25S9nfaiU4eWoGQ3Mp6LAGJhu2dkk"
    "MN2yE233BwaVhsCgUobAnmS/dnkyYjv5xyEVcwJIqpoVsvFynVSyD184MZKKmYvqUzsReVjiYTtK6KbV"
    "LIq2OlKvOKlXRlJJFL6MorqxhFoHTaZNAtORagxsZrPAAtsU2JPs/Q6kOkdLGHFPMGkiqWrGSF6uk6pu"
    "y5dOTKSiVnVOpSJPig6ZU6VNG1KlrYZUZ8BIdQYmUmkUPhUFxsA0c2qjwDSkmgOb8cCgUqY5VfZrF1KP"
    "lj9yMGOApKoJJHm5Tqq6S186MZJKE0hoWF2ZStFBpDZJt2ju/rI8Hak2J9U2kkrzSDSwwBiYjtQOeSRz"
    "YDOHBQaVMpLaPZ3E35s+yt2feyrnVDWfJC/XSVX37EsnRlJpPgkNa6Sasi+a53vpptWcui/7Uq5THYeT"
    "6hhJpWklGph4jR5eaD7k7t8oMN2cagxsxgODShlJ7ZBdEgc2xbm8TZS+ROw4adZbJG/s+CW88H9dSMXp"
    "jJkD79Sz/RtFPnVGcBaLyr85o8leORyP5bNSvyyWHXd9iR7D9CXeZr01HDllJ0DZsbFUnBIVf8BpVX4g"
    "Spw95V9f4Zx2BOf6BqegvEqSXP7BDsYVJ79v/g8AAP//AwBQSwMEFAAGAAgAAAAhAIPmjfHhBQAAnyEA"
    "ABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0Ni54bWysWtuO4kYQfY+Uf7D8PtjdGBgQsFqwN1kpkaIol2eP"
    "MWAtxsj2XFZR/j3VDYNx1fFoMvJol0vP6dPt6lPVx+6Zf3rJD85TWlZZcVy4auC7TnpMik123C3cP//4"
    "cnfvOlUdHzfxoTimC/d7Wrmflj/+MH8uym/VPk1rhxiO1cLd1/Vp5nlVsk/zuBoUp/RIv9kWZR7X9LXc"
    "edWpTOON7ZQfPO37Yy+Ps6N7ZpiV7+EottssScMieczTY30mKdNDXNP8q312ql7Z8uQ9dHlcfns83SVF"
    "fiKKh+yQ1d8tqevkyezr7liU8cOBrvtFBXHivJT0T9P/4eswtl2MlGdJWVTFth4Qs3ees7z8qTf14uTK"
    "JK//XTQq8Mr0KTML2FDpj01Jja5cuiEbfpBsfCUz4Spnj9lm4f7jX37u6F2ZF//OH5mXm59/3eV8k9EK"
    "m6tyynS7cD+rWaSnrrecWwH9laXP1c1np9oXzz+V2eaX7JiSGknHRqEPRfHNQL/SyL7p7IneX6xCfyud"
    "TbqNHw/178Xzz2m229eUDsEgoJmbpZ9tvodplZDmiGgwNFRJcaAZ0KuTZyZ3SDLxi31/zjb13nR3neSx"
    "qov870vDpdu5A0XYdqD31w6UbG90oCjaDvT+2mH8Zgca33ag90uHgAZ7Y4TRpQO9XzoMYQfvfO02lmFc"
    "x8t5WTw7JGAKQnWKTTlQMyIxQdSjwZQ+nke9RrYjrBRPQ/PZ8Czc8cR1iKEiNTwt1f1k7j3RCiYX0OoM"
    "GgU2qqbbWrSEoiW6tIysGmja17nTld7OHS776/wMli6tNb17Nj2EmbYxa4CZ+m1MiDCqjYkQRl8x3u1l"
    "knr4EqnxwP/fS2R4Fu6oFYPpkMXgDAps3tqVXYuWULREty2tuZOQe5m74aG5U5Fo5DUN2NwhaMQWEILG"
    "bAUhiKk5gqBGU6040GL1EgfDQ3GglLiJA9PoCoBo72ZxgCCm0hCCGplagUQQ1OiqFQcqf73EwfAwPWif"
    "6wGCuB4giOsBgrgeIKhDD1Qne4mD4WF60D7XAwIprgcI4nqAIK6HK8js3reFzDhTttd8qJAZHlbItOKF"
    "7Ay6LWSiJRQt0W1La+7TnuZuePiCKS5cCOLChSAuXAjiwoWgDuEqKsG9LKIl4imsuHQhSnPtYhQXL0Zx"
    "9WJURzlTwD59SNKWiGtac11cULeilk2hbIpaTS1dK2airIv+iLuwRHw9NRPtCqGUZqpdQy6fcYV4RMYV"
    "YVSTAu1o9OW11MVs3W7UWnPDiVHcckKUjAYaccgyJcIjdkWjL/dm7qvEdj1k+bnCKJafa4iS0YAjsj0i"
    "wiM2WdfWRl8eTiHDNBSZAlEiU6D5EpkCufg2gOfVFY2+nJxCtmkoMgWiRKa8x/GFcMRAZAriChrNtrXR"
    "l59TyGEFLAdWGMXUvYYomSlwRLbzRHjErmj0ZfTUxend3vbpQGQKRIlMQSgZDcglMgWiuhxTX9ZRIZ8W"
    "CMcEUEqzK1hDLhkNNOJIZApEdWhD9+UfLRH3GyOeKRjFMwWiRDQwF88UjGo026obui//aIm4fxyxHFhd"
    "ULf+UTaFsilqNbWvoC//qM9PxFqPePSIP0HEKLZbrDGKZUoIUWOubozqUndf/lEjNzcW6oYooW6EGolo"
    "QC6hbojqUndf/lEjNzcW6oYoXvkgl4wG5GI6i/C8msi2M6Uv/6iRm5sw3a4winnuNUTJaMAR+Z0zHrHj"
    "zln35R8tEd8HJvzOGaP4IxWIktFAznDC7zvxiB13Wrov/2iJ+POliTjoQJ5vIo46EEpGA6HuRRWFqK4q"
    "2pd/1Min3YsqClGiiiKUjAbkElUUorqqaF/+kQ5j5bPHe1FFIUpUUej5xJ4CuUQVhSheRc/nwOezyzwt"
    "d+k6PRwqJykezQmvohPHa2tz/qzsmWEDX85P8S79NS532bFyDunWHg5PqLyV5wNkf2C+1MXJnow+FDWd"
    "BduPe/pDiJROLP0B/X5bFPXrF/OQ+/qnFcv/AAAA//8DAFBLAwQUAAYACAAAACEA6aYluGYGAABTGwAA"
    "EwAAAHhsL3RoZW1lL3RoZW1lMS54bWzsWc1uGzcQvhfoOxB7TyzZkmIZkQNLluI2cWLYSoocqV1qlxF3"
    "uSApO7oVybFAgaJp0UuB3noo2gZIgF7Sp3Gbok2BvEKH5EpaWlRsJwb6Fx1sLffj/M9whrp67UHK0CER"
    "kvKsFVQvVwJEspBHNItbwZ1+79J6gKTCWYQZz0grmBAZXNt8/72reEMlJCUI9mdyA7eCRKl8Y2VFhrCM"
    "5WWekwzeDblIsYJHEa9EAh8B3ZStrFYqjZUU0yxAGU6B7O3hkIYE9TXJYHNKvMvgMVNSL4RMHGjSxNlh"
    "sNGoqhFyIjtMoEPMWgHwifhRnzxQAWJYKnjRCirmE6xsXl3BG8UmppbsLe3rmU+xr9gQjVYNTxEPZkyr"
    "vVrzyvaMvgEwtYjrdrudbnVGzwBwGIKmVpYyzVpvvdqe0iyB7NdF2p1KvVJz8SX6awsyN9vtdr1ZyGKJ"
    "GpD9WlvAr1cata1VB29AFl9fwNfaW51Ow8EbkMU3FvC9K81GzcUbUMJoNlpAa4f2egX1GWTI2Y4Xvg7w"
    "9UoBn6MgGmbRpVkMeaaWxVqK73PRA4AGMqxohtQkJ0McQhR3cDoQFGsGeIPg0hu7FMqFJc0LyVDQXLWC"
    "D3MMGTGn9+r596+eP0Wvnj85fvjs+OFPx48eHT/80dJyNu7gLC5vfPntZ39+/TH64+k3Lx9/4cfLMv7X"
    "Hz755efP/UDIoLlEL7588tuzJy+++vT37x574FsCD8rwPk2JRLfIEdrnKehmDONKTgbifDv6CabODpwA"
    "bQ/prkoc4K0JZj5cm7jGuyugePiA18f3HVkPEjFW1MP5RpI6wF3OWZsLrwFuaF4lC/fHWexnLsZl3D7G"
    "hz7eHZw5ru2Oc6ia06B0bN9JiCPmHsOZwjHJiEL6HR8R4tHuHqWOXXdpKLjkQ4XuUdTG1GuSPh04gTTf"
    "tENT8MvEpzO42rHN7l3U5syn9TY5dJGQEJh5hO8T5pjxOh4rnPpI9nHKyga/iVXiE/JgIsIyrisVeDom"
    "jKNuRKT07bktQN+S029gqFdet++ySeoihaIjH82bmPMycpuPOglOc6/MNEvK2A/kCEIUoz2ufPBd7maI"
    "fgY/4Gypu+9S4rj79EJwh8aOSPMA0W/GoqjaTv1Nafa6YswoVON3xXh6Om3B0eRLiZ0TJXgZ7l9YeLfx"
    "ONsjEOuLB8+7uvuu7gb/+bq7LJfPWm3nBRaa5HlfbLrkdGmTPKSMHagJIzel6ZMlHBZRDxZNA2+muNnQ"
    "lCfwtSjuDi4W2OxBgquPqEoOEpxDj101I18sC9KxRDmXMNuZZTN8khO0zThJoc02k2Fdzwy2Hkisdnlk"
    "l9fKs+GMjJkUYzN/ThmtaQJnZbZ25e2YVa1US83mqlY1oplS56g2Uxl8uKgaLM6sCV0Igt4FrNyAEV3L"
    "DrMJZiTSdrdz89QtmvWFukgmOCKFj7Teiz6qGidNY2UaRh4f6TnvFB+VuDU12bfgdhYnldnVlrCbeu9t"
    "vDQdbude0nl7Ih1ZVk5OlqGjVtCsr9YDFOK8FQxhrIWvaQ5el7rxwyyGu6FQCRv2pyazCde5N5v+sKzC"
    "TYW1+4LCTh3IhVTbWCY2NMyrIgRYZoZwI/9qHcx6UQrYSH8DKdbWIRj+NinAjq5ryXBIQlV2dmnF3FEY"
    "QFFK+VgRcZBER2jAxmIfg/t1qII+EZVwO2Eqgn6AqzRtbfPKLc5F0pUvsAzOrmOWJ7gotzpFp5ls4SaP"
    "ZzKYJyutEQ9088pulDu/KiblL0iVchj/z1TR5wlcF6xF2gMh3OQKjHS+tgIuVMKhCuUJDXsCLrlM7YBo"
    "getYeA1BBffJ5r8gh/q/zTlLw6Q1TH1qn8ZIUDiPVCII2YOyZKLvFGLV4uyyJFlByERUSVyZW7EH5JCw"
    "vq6BDX22ByiBUDfVpCgDBncy/tznIoMGsW5y/qmdj03m87YHujuwLZbdf8ZepFYq+qWjoOk9+0xPNSsH"
    "rznYz3nU2oq1oPFq/cxHbQ6XPkj/gfOPipDZHyf0gdrn+1BbEfzWYNsrBFF9yTYeSBdIWx4H0DjZRRtM"
    "mpRtWIru9sLbKLiRLjrdGV/I0jfpdM9p7Flz5rJzcvH13ef5jF1Y2LF1udP1mBqS9mSK6vZoOsgYx5hf"
    "tco/PPHBfXD0Nlzxj5mS9mr/AVzxwZRhfySA5LfONVs3/wIAAP//AwBQSwMEFAAGAAgAAAAhAKSOwAbz"
    "BgAAqkkAAA0AAAB4bC9zdHlsZXMueG1s7Fxfj5s4EH8/6b4DQrqnUxZMgE22Sar8AalSrzqpe9I99IUQ"
    "krVqIAJnm7S6735jA8FRliTscgmuLittwDHjn2fG4xl7zOD9NiTKc5CkOI6GKrrTVSWI/HiBo9VQ/evR"
    "7fRUJaVetPBIHAVDdRek6vvRr78MUrojweenIKAKkIjSofpE6fpB01L/KQi99C5eBxH8soyT0KNwm6y0"
    "dJ0E3iJlD4VEM3Td1kIPR2pG4SH0LyESesnXzbrjx+Hao3iOCaY7TktVQv/hwyqKE29OAOoWmZ6vbJGd"
    "GMo2KRrhpUfthNhP4jRe0jugq8XLJfaDY7h9ra95fkkJKL+OErI03Tjo+zZ5JSVTS4JnzMSnjgbRJnRD"
    "mip+vInoULX2RUr2y4cFyNg2VSWTyjReAJ/0O/03VSsePqhpHdWsqGgfVvzyu/7uS0d/p1dUvz+qfscf"
    "gP+ApuIh0EQRNrQBwFkz7KvohJazYDRYxlHJCQQAueQfvkbxt8hlvwErgD+s2miQfleePQIliDXuxyRO"
    "FAp6DOzhJZEXBlmNqUfwPMGs2tILMdllxQYr4Kqf1wsxKCIr1LIWsv9zVqtoq1u2lazmQ9XNP7z/+wbH"
    "CfbISUK9Izom/7ydTgaoLp1+Q/0ShVGbQQWXj8GgsT3pTi/pVDUNnX/ezpj7mTXWM4U/I/BqLL0JGqNx"
    "XSyIt5pp+quZe6x5l2tM0aFjGvUFdEzjcqYcMLYc5y8Pu2opGPxzWgrcFKRgODAheyNtcCsEJaMBTGg0"
    "SCIXbpT8+nG3BhsUwdybmQBejz1/ovYq8XbIsC5/II0JXjBbuJpyy5frAxobE5MrliYgYxbtEhQVRB3H"
    "NV2XY2uOqOFYep8rdINISzVskGiplw0StWam3r1vmKfmGJk9biYbRAouX99qmujE7dp600Qd20KZZW5S"
    "+Xuu5fAZqUGiYHJnTtPSB6Ld6axhlXJdZ+I4DROd9hx7+h/wtOdwz69JQU0d02laT12L/TUuKDaPVxLl"
    "EwDMY/M4WUDoWIQbJkwhWdFoQIIlBV83wasn9k3jNfyfx5RCeDUaLLC3iiOPMPe4eEJ8EkJOiC6HKn2C"
    "6LDwx70NjXN3XGPkc+pn63IMHMLZqgCzQHm2btaZ83055MJNIbSRVbn8QZv8gJDPTO5/L/cqxXyj7VKI"
    "SWF9gkVSLJBll+CI5JeZ+mQ3TK1Eahltgez9/avoKtvlvoEqVN0SlaEqJSoEAWj+tOKt12THYlCmzvkd"
    "9KS8m/CBVd6PCV5FYZA9MBpACJrdKk9xgr8DIRa7+vB7AOscsJpDsS+WfEu89WOw5c0x3myX1VxliwM5"
    "V+FSavx9yfFDiCa3AGQfALAAIaEABAsk5QgW8MP6o3wWSMAPq45S44d1Wqnxw5Kx1PilnMEE/ZdzBhM7"
    "IOUMwHZjch8OtWUOftHvBO4W3jzbEiowgwBODNtPm3AeJC7fVxQc6Ku50zV6ImNHsu3BLMRiW26FUE77"
    "Eq0UCuh+oV6S9wTZfCv2p5CKEMILwWZbXL06duq0bW3lkBCYD87FfnC3mvsvK0xbfOs6CnOaza1UGJOp"
    "Sb7iJvscLRpR6fvCU26yCUHyvrRpQjg3mi+esdpu/YXV8zPGv2JlGjhVtQYuMOk2VrqGFE8DlEiKZzh9"
    "oRSRzXPzfjZHk63kFs7ObXydc/suLFC/cdB+FqLAxBaNa2H6QOLSTUvlLBhHlpXaivXJs5K/IWi21c8z"
    "EY52LZHNU4J/OlslMrvVy3fiyLsh6GoNqQLYln2NWuPuyqDrc7Utuy21uHpl0BdxVRhLVx7/HN4bM0gQ"
    "W6osVo2vvIPVNH7jyikYjeAXQrt2qE+NnCM5Uy6qUtGA/YdJX2Vi14nwuYoaOLevoCYYEym384W46PR2"
    "WvsT6qRMZxGs+ZXDlQpjWMOaSJl/IwxYKfNXBIW5ssK/XWGkT7mUPef1/5zpIiP8ijnfwrqelPxno7aI"
    "OKTvwC1s5tmIWWCwlBZG1JBbeDF1GCznHCRyWMq0c7EDckahYg+kDDVuOErrHZ7ip8vgPJlwaO3gyNr+"
    "8JnCXigwVD+xZFUS7YQAfr7BhOLohQNrQHWxLY/A8WPtlL3Shh+O27cDJnkRLL0NoY/7H4dqef1HsMCb"
    "EKx1XutP/BxTTmKoltcf2QFN2IcCAwUHxD6mcKoSvpVNgofqD2dy3585rtHp6ZNex+wGVqdvTWYdy5xO"
    "ZjO3rxv69B/hxTpveK0Ofw8QZBQg8yEl8PKdJO9sDv5zWTZUhZsMPt8eAtgi9r5h62M4wt1xuzrqmLbX"
    "6/TsrtVxLWTMbHPiwKlZAbv1ytfv6BpC2Yt8GHjrgeIwIDgqZFVISCwFIcHtiU5ohSS08iVLo38BAAD/"
    "/wMAUEsDBBQABgAIAAAAIQCPmK7gpQsAAOMkAAAUAAAAeGwvc2hhcmVkU3RyaW5ncy54bWykWktvI8cR"
    "vhPgf2jIWEALi6RI7fqFlYyhOFTGpDgTchhBuhhNsiWNSM4w81gueQhsxWvn7DXi2A6wCXLNzUgOsW8S"
    "r/kR+wvyE/L1DKmVuppSHoAhS1PV3dXVVfXVY599/GI8Ys9FGHmBv7tRLm5vMOH3g4Hnn+1udN164YMN"
    "FsXcH/BR4IvdjZmINj7ey+eeRVHMsNaPdjfO43jyUakU9c/FmEfFYCJ8UE6DcMxj/BmelaJJKPggOhci"
    "Ho9Kle3t90pj7vkbrB8kfry78eR9nJv43q8TsZ99qXy4vbH3LPL2nsV7HddwjzvuccNitXZ38dNxyz4y"
    "2ZvPvmGfmDWzxQ7N/ZNnpXjvWUnyZ2v+9fr1n9nip6tXi8/dY7bLJjyMRMg2oyQMpoINuC8eM8b+9fqb"
    "v7GWZVYts4Ptd5mUOllccsaTOEgZXv+JVbvtzokrj5VbjUUczoacNRwLDG/+8BLf/NWm7HnEOt3DY1We"
    "Wpgsfp75PJ8rscaUhzFXOdonXUh6ZDdbpko6MViFOd1WwyX7grSzhuQ0u52CTXer2VXLbVjkkKplt6ED"
    "XEglGR1X97njtvEwrMTqRreJB6ELHbtjGTWjZRnqlk7DVT+ZdRMXPGrZi++vv1SJrapRMpt4arct7QBn"
    "aY6ru0Y+FwwGRTbnrKxuUXcP87m+GPliDZmNvSjK53xPpFx0/aN8LhomsejP/WDxw/VXKkfFuX1ChZIf"
    "OKHiZPLDOFmUjImJVJyHJNi5I8GOKsGO84AEOw9IsPOgBJV3y/kcH/YvBDl8LcX1JmzMB2L5POrClHzv"
    "00gOKCyfW1xev8TreORwcDykO7ttVqVh5XPzXpEFp8KPisSN7/AMhJbphmetNEuHgqKiWRTPyIVtGFoc"
    "8njGqm3iOfncJJyLcMp9j1hIvZbPnRbZdDZOornwiRqcenZD8EwQpxeXPuI+4ao2GxlbbxQMPVWAjtvM"
    "qFKM/vl0Ri8Av+9AzCDyOCItFdPYd0HXWkn9AD5wWjwL+IjxOCYv4MiVk8Qf0lONUnZuum8Jp9PFjoPF"
    "w3gNESdLolx//Z16bVE/eLSMO/mcOIUG8YG8XOctTxwmcOPzIKCXcO1bbEnoB0BgslfdDW/Oq7sM5kAe"
    "ym7HZzc8gZQJXIBuRSwJ1B9FE94HgAOJAYTPxcYesOv3f2fMPbI/MVbASuK09Uv1ipbuk+bbr8hKm4R8"
    "IKV64r2yto+PjCa1x5v0wGrVjo+sWtcAiLIT48iutazG1asjcoea+iVFX7qzBLjSenhzO+qSQ6ulA8IV"
    "bi1hTA9f3dqJZSw+V7c84dNg4HtD9fs7+dwwiObJaKgBihsg0qDQCiQJ6RZ6aKBjtYyQbgGrBnRXywhp"
    "GWXUa63C8LrQuz5Ykki5CnY8GRHXUaIYCRVZJEUgJZTlE+dzY89PaBS6FSfWRwd6CaBTvIwEKrFutUCd"
    "zobB4ov+XCy+UBkW37eL+ycGAuNmJIaPVfK7pYI2q2sZTQv5Y6ex+NyowUvY04LdsavIcvd/kabXdt1s"
    "deByzGJ2tW23SHTIlhKDRbqq+UbtRqaNVBPqF5ln3+s3jnf9MkZCvumHLPMIjRLser1UM+tEt64m7qZJ"
    "ksqJ5It+0nIiSVI5l4kXsdBbOEnsLIMzIvABzaVl0pYBlwZqxHgSrDEMAh9a9Nir3UYZ5DxFPVtLxG8Z"
    "FbH/CwS6Y3n/7zY6q30YYf5nEVJ8ylZr/cW1EDwOmAQb6W+yhHXsLOrjFyZ92DLx26pqYpvbbz57VXkS"
    "EZ9e1ZPEWbS1pTubqIz7psTIkix+s18nKI9lRjeYe3xxyfpzHiVss3L17fusl/SHIr76cUrEME9Pia3L"
    "AEgM3XGIAGJUzOe2I/V7C+WXlrCfLigXnqxboiNli54WPli3SEfKFn1YKFfWrdLSlgLuFMrvrRVRS1wu"
    "fL9QWauOspaYLayUCzCRNXrUE//5FyZr3Fn/XF0GCsonWeaSgHR/plhxSHq3Qz8ZTZK/aZM6YwlO5Q8Z"
    "2gxmnVVtq9OgACRp5Os7RBd8PuXoY4XilBRbdi+a81BjnbQukqYp+wcqsxHHsrM05sQ/nFXBoStl8Hoo"
    "9MPZlI80j5c+gZ4qz9Muk+6op0AQPaFm1euoZzQ+2xTTGZMVINucBIMUWcWY3tCusYbdOTEPic+H/D/Z"
    "ID1mcYlCU1x/xTZPedijeqwb7Sp55sX3YTAQEE+/xknPf2jn7JY8EiMPj7E59gbkiodWrdA2WgckV8Fa"
    "VHjDcD4b4Bj94oOrH0POhqMEAVXPIeV8cJvsMr175UyvIoZRXFx/0Izdw1DjIzH02Eg+vFYTS4ZJqlkt"
    "RyqDz8Ng8TNKFraJEED0uVNIe5mqvaja1C3NtLm4TIbAJh0DUeY6phl7QMqmeWC2akaG5o7VYUedE3Sh"
    "GxaS44bd7B62kBZndZ16E7NpHpotWu022levyFf3mECjjePUPbtHxoEFvEYzHFK1SQH75pvP8J+UzG4f"
    "s+wvdY87TXGVaLWcLhFO9snZVFwsfuh7aHGra44mXoTwBR6ZOcxXXXb8MvLwQnMOevHFKHqhiW9MDGZx"
    "MOVw+vD6O23T7m53Xt2jbrcPuyhd1O9Bb+TB2XxSFRto6WMkMUvPSiVeMi6lN18gut8rqUahd6YDKh2l"
    "i/ppOT0gnDJABPC7aTTki39I30FzEU2uqWw190kTzvD5yEuvQq755q9/ZJ0To01mBa0S0dUkiWhXCTDH"
    "BoHcfMZicYHIP+55Pu9fECmWRpfNYzAI0ZudHJG4ptGEx50tLpFGCtLLkDuoKsFMBYEzGCVj+pQ1aXPI"
    "U4Ke6EsLZUE4O4OMSGnG6fwldVtU7/yMX1y/TOcz8iJyM+j4tn2qx570vCAcelsMTdothnx9i6FTsQXH"
    "a24xhxSQWSsHOfuh0TYWv90/QfmsMYNlBbvGGuApHG4U4cb0rofpvMkDhqBbnO2DIdZI8DP0FwdB6qKD"
    "gCGLAVj7iXqEhPYt5nbwQ9Z7W2j1yZ+o1lhb1yzcWz7qzaglDSzqrsg0RD/2ngtdF3QTgxD2LisXn159"
    "i57SY4QtfDLwCfMFAgeo7cFgvr0TBkoWRXtXtlM7sp2KBqeu8aqKiE5xejC7+pZtoq2M47eLT57IPzEo"
    "ekzksHElP5JX0utFak93BvYrb2/jKNl9Tu0ua0/D3FPCsg1O2zby4vz5GftNubxdKJefqpvX0vHCeoHk"
    "Q6prbMf5VF78Rqj0w0ow31v8HMGrd4HxExgMC3phQC1uvW1kNT45NK37f/d1alqpBnx+kZ5wO55dcDSw"
    "EMqQMWHAOiCGqtqBeoq7bI8x+vi25tPtB1kOU6QE7BEmH+lcYvEF25RnLn4Ii3iDnUfEIuqhEMw9x8RY"
    "mgRBPzToVRlhWHIeCmOTtT3OQx19/V2EUfJgOQBm80HQA2Yy5JsenTJkStDZvxN4fhyVnCCKRCTH9Xp7"
    "vzHCtLlTkhZIhIR33fE4Xchgsh+QVkwEYFBLlOaasdEhRnjS8CZx9OlzPoK7pqr4VMeLOCfxC127KYwh"
    "ojZ4G1owQcZ4WRuHltE6G9cBLMkjyXayer801dG1RBt4qLnohbIBMgkkwKRV8piF86QPVMxknY3VDd98"
    "+TWCq1lFjI1EP/AHrH/O/T7SopjU5SuJs+GhTuSaRuSHJEvLQZoKQC65G+TSBQtDO30E7hEAg15QA6Zq"
    "kS2irHmQqiUWWm1gk5Jrp+OoQN3tUNs3B9KqjPvoQbGzcAZ4zEYgPHWrYNK/QNTSFOLyHRwRFnbeA1Tj"
    "/0+2txi63+q+Vd1YEyCvu3aKqkMAM0wCVrAsvLWalgMN6TVnfEzM0Fk7K0VeoVX3ckHfk7aIH0MaLOR1"
    "5Rxj3aF1Ofi4b9BL0xmyJN18pX+d3deDZMTiMEh6dMjijJKoJJ+beIHmVdpXP8qisc9Zip/IE9nZ4O3b"
    "s94MmgCm9gIPvRFVlKM015P52ltKCf9Iae/fAAAA//8DAFBLAwQUAAYACAAAACEARMUWsSUTAAD9dgAA"
    "EAAAAHhsL2NhbGNDaGFpbi54bWx0nVuP3DiSRt8XmP9g1Pu0u66zM2j3IMspZUpZUrVTt5HeDHdtdwO+"
    "NGxjsPvvVxgxyNB3lI91KoLiJUgGySDzp3/+76ePr/798vXbH18+v7m6/uHHq1cvnz98+fWPz7+9uera"
    "/K//ffXq2/f3n399//HL55c3V//38u3qnz//5b9++vD+44e3v7//4/OrOYXP395c/f79+5//eP3624ff"
    "Xz69//bDlz9fPs//+Z8vXz+9/z7/+fW319/+/Pry/tdvv7+8fP/08fXNjz8+vP40J3D1808fXn19c3W6"
    "ubl69cebq/urVx/nrFy9DrxKfP5M4sfITfIAkoOcbq7DV1L6Rl6t0s8heYxkLXm6mavtPzlPaRpZSx4h"
    "eQDJQU7Xf9f0I5E8Q/J4QfJ0PbfsKs9HkBzkEMn6u9X13yS1J5ATyBHkAFJHoiXVL1bXD5KHE8hTJJqa"
    "6h4vSFbXs32u6u0JpI5E7ASSR5Ac5HAxtTtYhREtnUoer7clT9e3SNOIWDIkc5BDJGot1tOtvzxdKzmB"
    "HEEOIPlGOujv19v9/RC55SoHOV7Qra51BDiBPEWiNam6OXQPF3TrC/yEAcPA+tsHlcsVHLcVKx09Tgqe"
    "DIgtqlyFkUPBSUEOCe29TwbWH6+38VHVDwpyBZWOAicMCwaktVUuV3DYVnzSPnxC99/u00eVOyjIFTR3"
    "i0nexUn4DNJFsi5gc7tYjNMF6SJR3cU6km53a2QteY7cOmtzu1iF/66S9nZpxyTTgJxB+kg0t0tL+S8q"
    "6W+NrHX726XOk24L0oCcQbpINP1l2PYlVdLdGtG6peQyPPvUlJxvlfS3y0DrW1NJA5k2Es0VU1MrbW5h"
    "tyB9JNKaN7BbkP4G9gnSgjQg5410YKs3ShqQM0gfiZZOLf98g75wo9bbRiJtAcnmRu35DNLcqF2dQfrg"
    "Yrt+AdKAnEG6SLQeYCHB3XZfDC6zs3aQM0gfiXwxuMxulADpI5FeHJxfnzdYBWTOIF0kmj5sILi9rs9G"
    "IjYAySY4qr6kGA+jjOZE7acLLqpvBYyZUUbzxtQwHgbn1OdWZfooo22K0S84nim1PriKviZ1BGsh04Cc"
    "N9JRG26DG+jrCnYOmT4SKR0GQgWdDoONgv7CtA1FTNoKOgOSSYxnMGXMzgq67bm5wVCmYPzbUru30T/q"
    "QFqQBqQGqUBKkAIkB8lApoelFVOed48Rret2hOgA0oN0IC1IA1KDVCAlSAGSg2Qg48NifqkKBpAOZIpk"
    "XUtz9Vly63+0SKIBqUEqkBKkAMlBMpDxYelCzlpBWpAGpAapQEqQAiQHyUCmh6U7r6zVkFqrig5Q7kE6"
    "kBakAalBKpASpADJQTKQ6WEZnlZVYEirQEUHKPcgHUgL0oDUIBVICVKA5CAZyPSwjLyrKjCkVaCiA5R7"
    "kA6kBWlAapAKpAQpQHKQDGR8WJwS33mVTFEGI5SJ6gilSTT4TA1SgZQgBUgOkoGMD4s/5QuppIVMA1KD"
    "VCAlSAGSg2Qg08Piza1s05DapooOUO5BOpAWpAGpQSqQEqQAyUEykPFB3aABpAOZIoG1WnJqrfqZBonW"
    "IBVICVKA5CAZyHivnlQH0oI0IDVIBVKCFCA5SAaye7xX32cCGUEGkB6kA2lBGpAapAIpQQqQHCQDmQuv"
    "HtAEMoIMID1IB9KCNCA1SAVSghQgOUi2UVJ6UvcXPKnIbVN1AOlA+guptZBsQGqQCqQEKUBykAxkvMdM"
    "CjJFomPThX+0SKIBqUEqkBKkAMlBMpDxHjMpSAvSgNQgFUgJUoDkIBnIdM+ZNCKZSSE6gPQgHUgL0oDU"
    "IBVICVKA5CDZVoF1jhvvMbmCdCBTJDKHQrIBqUEqkBKkAMlBMpDxDnMoSAvSgNQgFUgJUoDkIBnI7vEO"
    "cyjICDKA9CAdSAvSgNQgFUgJUoDkINlGSXUKnavDkHTSyOM0AtKDdCAtSANSg1QgJUgBkoNkIOOdzqQd"
    "yBSJTiAX/tEiiQakBqlASpACJAfJQMY7bCWA9CAdyBQJKmJ7D6NFEg1IDVKBlCAFSA6SgYx3uqHQgbQg"
    "DUgNUoGUIAVIDpKBTHfq4syddHsLYIToANKDdCAtSANSg1QgJUgBkoNkIOOdOkADSAcyRQJrteRkPkUS"
    "DUgNUoGUIAVIDpKBjHfqQ3UgLUgDUoNUICVIAZKDZCBTiCDxOygR6ZQCZwjKPYiFqKQPtJCxUJckU0Om"
    "AilBCpAcJAOZQiCMr4KIpAogOoD0IBZp46oAMhax46oAMhVICVKA5CAZyBiia9zeH8gUiXbPC/+wQJ+U"
    "qAX6+ELqdF7hwyVIAZKDZCBjCANKXx9AOpApEhT7whyKJBqQGqQCKUEKkBwkAxlDdJJvW51VLe7JN5XK"
    "1EinAilBCpAcJAOZQlTUqktemEMhOoD0IBZ25bsktuOhVYNUICVIAZKDZCBjCOXy1optBchMkcBaL8yh"
    "SKIBqUEqkBKkAMlBMpAxRJ55a9VZ1aLTvLWqTI10KpASpADJQTKQKUSzrax1ext9hOgAYsFxvgp05m2h"
    "ZWF3fmxVrQpaJUgBkoNkIFMI0/NVEJHMoRAdQHqQDqQFaUBqkAqkBClAcpAMZAzxhK7ZQCx2cR5qVtdc"
    "LF7RWTR0a5AKpAQpQHKQDGQMEYxu8AHpQKZIdPC58I8WSVh4pbNiyFQgJUgBkoNkIGMIx/RNqP5JC5kG"
    "pAapQEqQAiQHyUCmEAa66nkXIhQgOoD0IB2IRaR6a9UNixpaFUgJUoDkIBnIGKJcvbVi3wAyUySw1u3o"
    "hhZJWMCtt1b9cAWtEqQAyUEykDEE7nprVR+mhYwFAPs8q1YFrRKkAMlBMpDdY7gMmD4/gYwgA0gP0oG0"
    "IA1IDWJXHVMOS8gUIHbJMWllkNk9hvBkV3iQEWQAsTBn1/KQaUEakBqkAilBCpAcJNsol+7bdyHg2tfG"
    "djDfXHEXovyQRANSg1QgJUgBYhcxXQtDZgwx3r5hcOodZXTcufCPFok2IHZDM33YboM6M4ZWAZKDZCBj"
    "CDr3hcQsCZkGpAax26M+z5pyAa0cJAOZQsh7Snm2qQuzJEQHkB6kA2lBLDbfDb2Qsautvgp0bi2gZRdW"
    "vW2q1hji9JPMAGLx/r5Lbk+Gc/VdmCWRaANSg1QgJUgBkoNkIGO4PuCtFbMkZBqQGqQCKUEKELuh65sK"
    "sYThIsPKWrfX7yNEB5AepANpQRqQGqQCKUEKELsg7KtAtzamcONiVQUXIvsgOoDYRRBvBdhWgJZdBPEd"
    "VrUqaJUgBYjdavZVoClP4WrIqgoubCtAdACxmya+CrCtAK0GxK47p3Ts4rUfszTlAunYBWtfBYi/QICC"
    "gsmATqbbvFX9RkGtoFJQKigUhAvcrmQqMcIZUtAqCNd7nC2qRLgP7ppBJQoFuYIM0RUMObgQcaCCg4Ie"
    "VyAUtAoaBbWCcHXdFVolCgXh6rprHpUYeZvhP08yuL6jEtN2ON98QSVoyo6P6jcKwj15181UooRvpCBc"
    "mXclVYlR3YRwHyuptCrRYK2vIFzMd20CF0ZBuIHvcooQiLBXfJPuw4P8AvIMYrvZKR3bzU7kCVonENvx"
    "Tlq2453IEVoHENsDT1pvIfMIYve6k9a7sC+diO1dJ7KHTBfJ2j5tEzrpdmHT139xGa8TsW3peTZfP+MD"
    "3T2IbWCrru0Fu5yEvVf3XRDbC3a5hYzt6rr6gYzt2PqvL0NkIna33H9LZWzX1X9LZWwXmDWwjE8+fSW2"
    "velzZTLrttiHbc8k2YdNRa+7jBP+i0psc9LXjMnoF1XXdhl9+kt/T8R2/nyNqUwfNsV8OsuyxKejxDby"
    "tJ73SK0PG1g+/cVZ9+kr2UPLdqt8OovHm4jtVqHvhH0kXw+qa/tatJzFqfRfUfIu3NL3JVIZu8nPGlPJ"
    "Ptzk9yXVUcIe1/IlUpkuprO2pSFsRyRdu2mfyC+QeQaxDZOkZRsmidjzWYnY81mJlEjZNlWSjD2xlYg9"
    "sZWIbbMk8hYpP4L0YaPD1TaIbY8kmS7KSD+Fru1z+PTRT8NuhGtNkD4sy3066JVRRuYO6NouBawRkrb1"
    "gH4RluU+P+jXkNmD2Mra1632zXdh/ZtkbK2Nng7JPYitylki9PSw6PRlVBlbhvr8m4zaBnp6WNH59OEP"
    "RBltU5XcI7Xugm6/rKVclhWERV6SCIs8l1NVCSs2Z8MqER5TcGlgilewVxCebEDbYXrH7K5gryC80ICk"
    "MY9jGlewx2StPf4dhgAFewVhzaDZ29nLiOlFyd0hHB04ZKcJDu0p9ZbIDlyc4g5SE4idwSS1fzFtOyqZ"
    "+8XK17XjmqTcQdkOcFTXDm2S7jvo/gLyDGIHOykde74yETvYcY95Ih076tF8so3sCCilxubYaI3ncITi"
    "GshOZxyydzIdsvMZh+ztzJUtLV3QITvq0SLt7HzHyT7F3MniCLk+g9hBlH4npRnfKb3wFZYnFXqdnwxf"
    "37NezQdaVfUyjKyqGshcnFXtQ8qcnFXtq1QfPJgkZN5bIna4xYrT1MxjU0nzwLxlq675W0lmj7ztnoND"
    "tKqyZXhcVQaQPVy6qgxIlTF52bw0l82p22GZltU8QF9/+qEzimFPsCYtO3rT9FkSO1zzNYei2TGWK4E9"
    "s7qqk2WecMhOslZFNynthaptB2zevFSGObO3Y725qFYWXFpfaOTdDrFc3k/BJV3Vw+L5ruoB6DkqaqFV"
    "1HxcX2iVYc6ekDF7jnbVGJaQGKgd4Km12PGbryYUzU5MVnWyeJerOgGyYxQM3VlwdZ36nugt0SPRDsiO"
    "olLqI2T+BWJnUEmrg4w9PZZk7JxKi/gOuvYkmUr+AslnkBrkCeQEYudaKbf2Wm8itqDwfcnacW1DbI2N"
    "xngOawDXsHbY5FBJqSORPd+7su9lGeKQnW5xnFVROxnz/U5lmFl7gNhXEDJh7xaze6moLZt8BS0rF9+7"
    "QMKTw74qIBMOwlARKnhWEFZbqhjO0dI3n/DJ8GIy5nMVDO8hp6Qyldgj7XA45oocHkd2JLyf7E0L/mM4"
    "QvPVC5kDyLMRnekhWW9LhuWnMzbVDOeDcJtULlbEOiso11EVw9FgykJY33rboz8JEo7jfAVC5kAn1IhW"
    "ICSftiXD4aOrQNUMZ42oQLid2x84qlwsqDjrKhfW8q5CnulPgZzoTtIzpWMKEp7ZTl/v4T4qCEeVqCiV"
    "qwysKyC89O0Hn23BTBMMWxNJ02L7Zlds9XMVO85wu/AcuftsOGH1GcGkl3wOW67tODNOuMC0e4xIdgcg"
    "aq9QuxPXsHjz7+8uRpN2n+y12qTV82kVpGNnae491HDi5b6FV71tRX4fT15tEZvI7hkjLsd3blTsuAux"
    "S5sO8Vlx7r6EmIvUdhar41rTAsZcn0dCNou77cQQX+62D0HajZ3ypYVSe0xb7z9ZHIgYBUR7KtNj57YY"
    "d8XCrOCLpxvU0zLQp5zvHo1INlUwTEquzLhJMgd8bydmD7o7A8bz9nZTOBlnGFK8JS779olYoN48Jqz2"
    "5QodTSwkKZnHxjqCzrMFoDpDs984WY0kOsyHwApnVjiMt4fDU50MvIaLZzrnEBOzPWkviPYki6rr/gDx"
    "9fz4QyH4cZw6Elmt4v1xO3tNH9xx/T2PCrq4Drv+rmYQfD0tOt6OjWh3k2CeXjV7jIFBxA+ci1IihzAo"
    "eFNcbE5NcReimPyUjxIPvE+2pOaet1bQAsSnvq3x7OjdpYLH1+3w1U9Ay8ojES6wJzzPOFvmMtjofbcB"
    "oj0JXvO2qFXXCLDNetluSG0Q3LIEgm/pWmmxeG2k4A6m/jobqg4hITbKGbPFZqem7XBq2IJYuLH7mLk2"
    "KSG7dpWEBjziuXvWYSfEESalCW9/zU/zXWgmiPYk+AmDDsTuzDl70gOiDjETdnXJz8N6RtstDZJEQgdw"
    "Oiox4W7AHMq+CMFKIWq7rX7SMt3YwZY53hmpgnA+iTHXT9J6Or4LUaduyOB+zITHaOanZ7bD9AeI9nxZ"
    "B75hT4PjztGEmPe5hrcv7w8Qtdf+/UCvnpXdtUgV2CDqxO6IO7vfuox2IVsQtUCglC129HlhoLNWi3zZ"
    "wiQNP9Y5nG/PhyrwRFTo1s6G+OMHIWjQjfSLLTjTVNAqmLZeYLvgxUK0BxmwSBkssbi+spsf3rfXXvzE"
    "FUE64VzP9hPu68+Ls2VPEh0eohZm6VxcPOq/e4xI/AyI9vytAaw4zbrdGIZ4Grt46kcMXSFOuCo8++IX"
    "BjqIWkhgKrcN0N5vWHqP9xJAaIP4iagOpAUJm1FuFl9WFW4WN7BuAxuenODSSd3SlXuy8xpUd1rtOpcb"
    "UPCY0jzaXjAsiPZ8Kwq+st23caudpbO4whhYl/oQTnfcKBN/7Sv9UOeS1yTD9bKtc92IETqea/T4C1aW"
    "crfx6106B9pq1rVoSNm7z9uls6W+9wyWsrgxFNMyboZZ8KibzPEcSY9Hvno8o9njceoeP/9gLp+vNvWo"
    "n/BbpzbKOQPAb6naBWasMSx6ca4Z3RtTH7HGl+2Az/USRnikCJg4dqfAE/noasB+HX+49uf/BwAA//8D"
    "AFBLAwQUAAYACAAAACEAJcRnYVMBAABlAgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAjJLNTsMwEITvSLxD5HviuIECVppKgHoplZAIAnGz7G0bEf/INqTlyLvxXjhpG1LB"
    "gaN3xp9nVs6nG1lH72BdpdUEkSRFESiuRaVWE/RYzuJLFDnPlGC1VjBBW3BoWpye5NxQri3cW23A+gpc"
    "FEjKUW4maO29oRg7vgbJXBIcKohLbSXz4WhX2DD+ylaAR2k6xhI8E8wz3AJj0xPRHil4jzRvtu4AgmOo"
    "QYLyDpOE4B+vByvdnxc6ZeCUld+a0Gkfd8gWfCf27o2remPTNEmTdTFCfoKfF3cPXdW4Uu2uOKAiF5xy"
    "C8xrW7T9zXZT53gwbBdYM+cXYdfLCsT1tphrJ1k011+f/CPHv/XA7CrswCCiEIruKhyUp+zmtpyhYpSO"
    "xnGaxeSsJCklF5SQl/b5o/ttyN1A7kP8h3heplc0zWg2GhAPgKLLffwxim8AAAD//wMAUEsDBBQABgAI"
    "AAAAIQA5rZb7ugEAAJUDAAAQAAgBZG9jUHJvcHMvYXBwLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AJyT32/aMBDH3yftf7D8XpJ2E5qQ4yqCrCDxIyKh0vbmORewGuzINgj61++SqDS07GVvd/c9f/Xxnc0e"
    "T/uKHME6ZXRE7wchJaClKZTeRnST/7z7QYnzQheiMhoiegZHH/nXLyy1pgbrFTiCFtpFdOd9PQoCJ3ew"
    "F26AskalNHYvPKZ2G5iyVBImRh72oH3wEIbDAE4edAHFXX0xpJ3j6Oj/17QwsuFzz/m5RmDO4rqulBQe"
    "b8kXSlrjTOlJcpJQsaAvMqTLQB6s8mcesqCfskyKCsZozEtROWDBe4FNQTRDS4WyjrOjHx1BemOJU684"
    "tgdK/ggHDU5Ej8IqoT1iNW1d0sZV7bzlsX05uFd0R7WrtGG/sR+r73zYNmBw3dgYdBQoXPPlylfgVmUq"
    "rL+BO+zjtgwdbIeTJ/GCPCXLZB3P+5QX3nQe/0rW2U1tPlsmm/S2lk1XOclni9ny6ebZ8Wqzzsnv1TL5"
    "h3eCVJP40+TaTeAMPtx6rvSL29S5mQgPbyu9LrJsJywU+AouK78U2BS3aavGZLwTegvFW89noXmAz90v"
    "4/fDQfgtxLfVq7Hg/T/xvwAAAP//AwBQSwECLQAUAAYACAAAACEAxcTpJYIBAAC4BwAAEwAAAAAAAAAA"
    "AAAAAAAAAAAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLAQItABQABgAIAAAAIQC1VTAj9AAAAEwCAAALAAAA"
    "AAAAAAAAAAAAALsDAABfcmVscy8ucmVsc1BLAQItABQABgAIAAAAIQA91AX1jAMAAP8IAAAPAAAAAAAA"
    "AAAAAAAAAOAGAAB4bC93b3JrYm9vay54bWxQSwECLQAUAAYACAAAACEAZwxbpSkBAAABBgAAGgAAAAAA"
    "AAAAAAAAAACZCgAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECLQAUAAYACAAAACEAWM/U0xEU"
    "AACydgAAGAAAAAAAAAAAAAAAAAACDQAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAi0AFAAGAAgA"
    "AAAhAB7TQ4jtDwAA+WgAABgAAAAAAAAAAAAAAAAASSEAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbFBL"
    "AQItABQABgAIAAAAIQA8vx1rfSQAAGQxAQAYAAAAAAAAAAAAAAAAAGwxAAB4bC93b3Jrc2hlZXRzL3No"
    "ZWV0My54bWxQSwECLQAUAAYACAAAACEANvAzPSQSAACSeQAAGAAAAAAAAAAAAAAAAAAfVgAAeGwvd29y"
    "a3NoZWV0cy9zaGVldDQueG1sUEsBAi0AFAAGAAgAAAAhACuLE21MCgAAPj4AABgAAAAAAAAAAAAAAAAA"
    "eWgAAHhsL3dvcmtzaGVldHMvc2hlZXQ1LnhtbFBLAQItABQABgAIAAAAIQCD5o3x4QUAAJ8hAAAYAAAA"
    "AAAAAAAAAAAAAPtyAAB4bC93b3Jrc2hlZXRzL3NoZWV0Ni54bWxQSwECLQAUAAYACAAAACEA6aYluGYG"
    "AABTGwAAEwAAAAAAAAAAAAAAAAASeQAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQItABQABgAIAAAAIQCk"
    "jsAG8wYAAKpJAAANAAAAAAAAAAAAAAAAAKl/AAB4bC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhAI+Y"
    "ruClCwAA4yQAABQAAAAAAAAAAAAAAAAAx4YAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAi0AFAAGAAgA"
    "AAAhAETFFrElEwAA/XYAABAAAAAAAAAAAAAAAAAAnpIAAHhsL2NhbGNDaGFpbi54bWxQSwECLQAUAAYA"
    "CAAAACEAJcRnYVMBAABlAgAAEQAAAAAAAAAAAAAAAADxpQAAZG9jUHJvcHMvY29yZS54bWxQSwECLQAU"
    "AAYACAAAACEAOa2W+7oBAACVAwAAEAAAAAAAAAAAAAAAAAB7qAAAZG9jUHJvcHMvYXBwLnhtbFBLBQYA"
    "AAAAEAAQABwEAABrqwAAAAA="
    )


@app.route("/mecz/<int:match_id>/export/xlsx")
def export_match_xlsx(match_id):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m: return redirect(url_for("historia"))

    gtk_name = (m.get("nazwa_gtk","") or "").strip() or get_setting("gtk_name") or "GTK"
    name_opp = m["przeciwnik"]
    dt = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else ""

    cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (match_id,))
    all_stats = list(cur.fetchall())
    cur.execute("SELECT * FROM player_stats WHERE match_id=%s", (match_id,))
    all_players = list(cur.fetchall())
    cur.execute("SELECT * FROM timing_stats WHERE match_id=%s", (match_id,))
    all_timing = list(cur.fetchall())
    cur.close()

    # Wczytaj szablon (wbudowany w kod)
    import base64 as _b64
    tmpl = openpyxl.load_workbook(io.BytesIO(_b64.b64decode(_get_szablon_b64())))

    def qd(druzyna, kwarta):
        r = next((dict(x) for x in all_stats if x["druzyna"]==druzyna and x["kwarta"]==kwarta), {})
        return r

    def td(druzyna, bucket):
        r = next((dict(x) for x in all_timing if x["druzyna"]==druzyna and x["bucket"]==bucket), {})
        return r

    CTR = Alignment(horizontal="center", vertical="center")
    YEL = PatternFill("solid", fgColor="FFF9C4")

    def fill_cell(ws, row, col, val):
        c = ws.cell(row, col, val)
        c.alignment = CTR

    # ── TEAM GENERAL ─────────────────────────────────────────────────────────
    # Kolumny wg szablonu (wiersz 4):
    # A=drużyna/kwarta, B=FTA, C=FTM, D=FT miss, E=FT%,
    # F=2PM, G=2P miss, H=2PA, I=2P%, J=3PM, K=3P miss, L=3PA, M=3P%,
    # N=2+1, O=3+1, P=Tip made, Q=Tip miss, R=Tip sum, S=Tip%,
    # T=OREB, U=DREB, V=REB, W=AST, X=TO/BR, Y=P, Z=FD,
    # AD=POSS, AG=PTS, AK=PPP, AM=eFG%, AN=TS%, AO=TO%, AP=FTr, AQ=ORtg
    ws1 = tmpl['TEAM GENERAL']
    ws1['A1'] = f"STATYSTYKI DRUŻYNOWE — {gtk_name} vs {name_opp} | {dt}"

    Q_NAMES = {1:"IQ", 2:"IIQ", 3:"IIIQ", 4:"IVQ"}

    for team_label, druzyna, start_row in [(gtk_name,"gtk",6),(name_opp,"opp",14)]:
        ws1.cell(start_row-1, 1, f"  ▶  {team_label}")
        for qi, qn in enumerate([1,2,3,4]):
            r = start_row + qi
            q = qd(druzyna, qn)
            ws1.cell(r,1, Q_NAMES[qn])
            fill_cell(ws1, r, 2, q.get("fta",0))    # FTA
            fill_cell(ws1, r, 3, q.get("ftm",0))    # FTM
            fill_cell(ws1, r, 6, q.get("p2m",0))    # 2PM
            fill_cell(ws1, r, 8, q.get("p2a",0))    # 2PA
            fill_cell(ws1, r, 7, q.get("p2a",0)-q.get("p2m",0))  # 2P miss
            fill_cell(ws1, r, 10, q.get("p3m",0))   # 3PM
            fill_cell(ws1, r, 12, q.get("p3a",0))   # 3PA
            fill_cell(ws1, r, 11, q.get("p3a",0)-q.get("p3m",0)) # 3P miss
            fill_cell(ws1, r, 24, q.get("br",0))    # TO/BR
            fill_cell(ws1, r, 26, q.get("fd",0))    # FD
            fill_cell(ws1, r, 30, q.get("poss",0))  # POSS
            fill_cell(ws1, r, 33, q.get("pts",0))   # PTS
        # OT row
        ws1.cell(start_row+4, 1, "OT")

    # ── PLAYERS ──────────────────────────────────────────────────────────────
    # Kolumny: A=Zawodnik, B=#, C=2PM, D=2PA, E=3PM, F=3PA, G=FTM, H=FTA,
    #          I=BLK, J=OREB, K=DREB, L=AST, M=TO, N=PF, O=STL, P=FD,
    #          Q=PTS, R=MIN, ...
    ws2 = tmpl['PLAYERS']
    start = 4
    for druzyna, label in [("gtk", gtk_name), ("opp", name_opp)]:
        ws2.cell(start, 1, f"  ▶  {label}")
        players = sorted([p for p in all_players if p["druzyna"]==druzyna],
                         key=lambda x: x["pts"], reverse=True)
        for i, p in enumerate(players):
            r = start + 1 + i
            fill_cell(ws2, r, 1, "")           # Zawodnik (brak nazwy w DB)
            fill_cell(ws2, r, 2, p["nr"])      # #
            fill_cell(ws2, r, 3, p["p2m"])     # 2PM
            fill_cell(ws2, r, 4, p["p2a"])     # 2PA
            fill_cell(ws2, r, 5, p["p3m"])     # 3PM
            fill_cell(ws2, r, 6, p["p3a"])     # 3PA
            fill_cell(ws2, r, 7, p["ftm"])     # FTM
            fill_cell(ws2, r, 8, p["fta"])     # FTA
            fill_cell(ws2, r, 10, p["oreb"])   # OREB
            fill_cell(ws2, r, 11, p["dreb"])   # DREB
            fill_cell(ws2, r, 12, p["ast"])    # AST
            fill_cell(ws2, r, 13, p["br"])     # TO
            fill_cell(ws2, r, 16, p["fd"])     # FD
            fill_cell(ws2, r, 17, p["pts"])    # PTS
        start = start + len(players) + 2
        if start < 20: start = 20  # separator między drużynami
        ws2.cell(start, 1, f"  ▶  {name_opp if druzyna=='gtk' else ''}")

    # ── SHOT TIMING ────────────────────────────────────────────────────────
    # Kolumny D-Q = celne/niecelne per bucket (0s,1-4s,5-8s,9-12s,13-16s,17-20s,21-24s)
    # Col D=Cel.0s, E=Niec.0s, F=Cel.1-4s, G=Niec.1-4s, ..., P=Cel.21-24s, Q=Niec.21-24s
    ws4 = tmpl['SHOT TIMING']
    bucket_cols = {
        "0s":    (4,5),
        "1-4s":  (6,7),
        "5-8s":  (8,9),
        "9-12s": (10,11),
        "13-16s":(12,13),
        "17-20s":(14,15),
        "21-24s":(16,17),
    }
    Q_LABEL = {1:"IQ", 2:"IIQ", 3:"IIIQ", 4:"IVQ"}

    # Mapowanie wierszy: TWOJA DRUŻYNA IQ 2PT=row4, IQ 3PT=row5, IQ ALL=row6, ...
    def timing_row(druzyna_label, kwarta_label, typ):
        base = 4 if druzyna_label == gtk_name else 22
        q_idx = {"IQ":0,"IIQ":1,"IIIQ":2,"IVQ":3,"OT":4,"SUMA":5}[kwarta_label]
        typ_idx = {"2PT":0,"3PT":1,"ALL":2}[typ]
        return base + q_idx*3 + typ_idx

    for druzyna, label in [("gtk", gtk_name), ("opp", name_opp)]:
        # Wpisz nazwy drużyn i kwart
        for qi, qn in enumerate([1,2,3,4]):
            for typ in ["2PT","3PT","ALL"]:
                r = timing_row(label, Q_LABEL[qn], typ)
                ws4.cell(r, 1, label)
                ws4.cell(r, 2, Q_LABEL[qn])
                ws4.cell(r, 3, typ)

            # Dane per bucket
            t2 = td(druzyna, "0s")  # placeholder
            for b, (col_made, col_miss) in bucket_cols.items():
                t = td(druzyna, b)
                q_timing_all = [x for x in all_timing
                                if x["druzyna"]==druzyna and x["bucket"]==b]
                # 2PT row
                r2 = timing_row(label, Q_LABEL[qn], "2PT")
                r3 = timing_row(label, Q_LABEL[qn], "3PT")

            # Suma per bucket (wszystkie kwarty)
            for b, (col_made, col_miss) in bucket_cols.items():
                t = td(druzyna, b)
                r_sum2 = timing_row(label, "SUMA", "2PT")
                r_sum3 = timing_row(label, "SUMA", "3PT")
                fill_cell(ws4, r_sum2, col_made, t.get("made2",0))
                fill_cell(ws4, r_sum2, col_miss, t.get("att2",0)-t.get("made2",0))
                fill_cell(ws4, r_sum3, col_made, t.get("made3",0))
                fill_cell(ws4, r_sum3, col_miss, t.get("att3",0)-t.get("made3",0))

    buf = io.BytesIO()
    tmpl.save(buf)
    buf.seek(0)
    filename = f"raport_{gtk_name}_vs_{name_opp}_{dt}.xlsx".replace(" ","_").replace("/","")
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route("/mecz/<int:match_id>/export/pdf")
def export_match_pdf(match_id):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m: return redirect(url_for("historia"))

    gtk_name = (m.get("nazwa_gtk","") or "").strip() or get_setting("gtk_name") or "GTK"
    name_opp = m["przeciwnik"]
    dt = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else ""

    cur.execute("SELECT * FROM match_stats WHERE match_id=%s ORDER BY kwarta", (match_id,))
    all_stats = list(cur.fetchall())
    cur.execute("SELECT * FROM player_stats WHERE match_id=%s", (match_id,))
    all_players = list(cur.fetchall())
    cur.execute("SELECT * FROM timing_stats WHERE match_id=%s", (match_id,))
    all_timing = list(cur.fetchall())
    cur.close()

    def build_suma(druzyna):
        s = {"pts":0,"poss":0,"p2m":0,"p2a":0,"p3m":0,"p3a":0,"ftm":0,"fta":0,"br":0,"fd":0}
        for row in all_stats:
            if row["druzyna"]==druzyna:
                for k in s: s[k] += row.get(k,0)
        return s

    suma_gtk = build_suma("gtk"); suma_opp = build_suma("opp")
    kpi_gtk  = calc_kpi(suma_gtk); kpi_opp = calc_kpi(suma_opp)

    winner = gtk_name if m['wynik_gtk']>m['wynik_opp'] else (name_opp if m['wynik_opp']>m['wynik_gtk'] else "Remis")

    def q_row(druzyna, qn, label):
        q = next((dict(r) for r in all_stats if r["druzyna"]==druzyna and r["kwarta"]==qn), {})
        k = calc_kpi(q)
        bg = "#f5f8ff" if qn%2==0 else "#fff"
        return f"""<tr style="background:{bg}">
            <td style="font-weight:700;padding:4px 8px">{label}</td>
            <td style="text-align:center;font-weight:700">{q.get('pts',0)}</td>
            <td style="text-align:center">{q.get('p2m',0)}/{q.get('p2a',0)}</td>
            <td style="text-align:center">{k['p2_pct']}</td>
            <td style="text-align:center">{q.get('p3m',0)}/{q.get('p3a',0)}</td>
            <td style="text-align:center">{k['p3_pct']}</td>
            <td style="text-align:center">{q.get('ftm',0)}/{q.get('fta',0)}</td>
            <td style="text-align:center">{q.get('br',0)}</td>
            <td style="text-align:center">{q.get('poss',0)}</td>
            <td style="text-align:center">{k['efg']}</td>
            <td style="text-align:center">{k['ortg']}</td>
        </tr>"""

    def player_rows(druzyna):
        players = sorted([p for p in all_players if p["druzyna"]==druzyna],
                         key=lambda x: x["pts"], reverse=True)
        rows = ""
        for i, p in enumerate(players):
            fga = p.get("p2a",0)+p.get("p3a",0)
            fta = p.get("fta",0)
            pm2 = p.get("p2m",0); pm3 = p.get("p3m",0)
            pts = p.get("pts",0)
            efg = f"{(pm2+1.5*pm3)/fga:.0%}" if fga else "-"
            ts  = f"{pts/(2*(fga+0.44*fta)):.0%}" if (fga+fta) else "-"
            bg = "#f5f8ff" if i%2==0 else "#fff"
            rows += f"""<tr style="background:{bg}">
                <td style="font-weight:700;padding:3px 6px">#{p['nr']}</td>
                <td style="text-align:center;font-weight:700;color:#1a2b4a">{pts}</td>
                <td style="text-align:center">{pm2}/{p.get('p2a',0)}</td>
                <td style="text-align:center">{pm3}/{p.get('p3a',0)}</td>
                <td style="text-align:center">{p.get('ftm',0)}/{fta}</td>
                <td style="text-align:center;font-weight:600">{efg}</td>
                <td style="text-align:center">{ts}</td>
                <td style="text-align:center">{p.get('ast',0)}</td>
                <td style="text-align:center">{p.get('oreb',0)}</td>
                <td style="text-align:center">{p.get('dreb',0)}</td>
                <td style="text-align:center">{p.get('br',0)}</td>
            </tr>"""
        return rows

    def timing_rows():
        rows = ""
        for i, b in enumerate(BUCKETS):
            gd = next((dict(r) for r in all_timing if r["druzyna"]=="gtk" and r["bucket"]==b), {})
            od = next((dict(r) for r in all_timing if r["druzyna"]=="opp" and r["bucket"]==b), {})
            gm = gd.get("made2",0)+gd.get("made3",0); ga = gd.get("att2",0)+gd.get("att3",0)
            om = od.get("made2",0)+od.get("made3",0); oa = od.get("att2",0)+od.get("att3",0)
            ge = f"{gm/ga:.0%}" if ga else "-"
            oe = f"{om/oa:.0%}" if oa else "-"
            bg = "#f5f8ff" if i%2==0 else "#fff"
            rows += f"""<tr style="background:{bg}">
                <td style="font-weight:700;padding:3px 8px">{b}</td>
                <td style="text-align:center">{gm}/{ga}</td>
                <td style="text-align:center;font-weight:700;color:#1a6b3c">{ge}</td>
                <td style="text-align:center">{om}/{oa}</td>
                <td style="text-align:center;font-weight:700;color:#8b1a1a">{oe}</td>
            </tr>"""
        return rows

    TH = "style='background:#1a2b4a;color:#fff;padding:5px 8px;font-size:10px;text-align:center'"
    THL = "style='background:#1a2b4a;color:#fff;padding:5px 8px;font-size:10px'"
    THG = "style='background:#1a6b3c;color:#fff;padding:5px 8px;font-size:10px;text-align:center'"
    THR = "style='background:#8b1a1a;color:#fff;padding:5px 8px;font-size:10px;text-align:center'"

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  @page{{size:A4 landscape;margin:1.2cm}}
  body{{font-family:Arial,sans-serif;font-size:10px;color:#222;margin:0;padding:0}}
  h2{{font-size:14px;margin:0 0 2px}} h3{{font-size:11px;color:#1a2b4a;margin:10px 0 4px;text-transform:uppercase;letter-spacing:.5px}}
  .hero{{background:#1a2b4a;color:#fff;padding:10px 16px;border-radius:6px;margin-bottom:12px;display:flex;justify-content:space-between;align-items:center}}
  .score{{font-size:26px;font-weight:700;letter-spacing:3px}}
  table{{width:100%;border-collapse:collapse;margin-bottom:10px;font-size:9.5px}}
  th{{background:#1a2b4a;color:#fff;padding:4px 6px}}
  td{{padding:3px 6px;border-bottom:1px solid #eee}}
  .two-col{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:10px}}
  .section{{margin-bottom:12px;page-break-inside:avoid}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(6,1fr);gap:6px;margin-bottom:10px}}
  .kpi-box{{background:#f5f8ff;border-radius:4px;padding:5px 4px;text-align:center}}
  .kpi-val{{font-size:13px;font-weight:700;color:#1a2b4a}}
  .kpi-lbl{{font-size:7.5px;color:#888;text-transform:uppercase;letter-spacing:.3px}}
  .gtk-hdr{{background:#1a6b3c}} .opp-hdr{{background:#8b1a1a}}
  @media print{{.no-print{{display:none}}}}
</style>
</head><body>

<div class="hero">
  <div>
    <h2>{gtk_name} vs {name_opp}</h2>
    <div style="opacity:.7;font-size:9px">Sezon {m['sezon']} · {dt} · Basket Kołcz Analytics</div>
  </div>
  <div class="score">{m['wynik_gtk']} : {m['wynik_opp']}</div>
  <div style="font-size:11px;opacity:.85;font-weight:700">{winner}</div>
</div>

<div class="two-col">
  <div class="section">
    <h3 style="color:#1a6b3c">{gtk_name} — Kluczowe metryki</h3>
    <div class="kpi-grid">
      {''.join(f'<div class="kpi-box"><div class="kpi-val">{v}</div><div class="kpi-lbl">{l}</div></div>' for v,l in [
        (suma_gtk.get('pts',0),'Punkty'),(kpi_gtk['efg'],'eFG%'),(kpi_gtk['ts'],'TS%'),
        (kpi_gtk['ortg'],'ORtg'),(kpi_gtk['ppp'],'PPP'),(kpi_gtk['p2_pct'],'2PT%'),
        (kpi_gtk['p3_pct'],'3PT%'),(kpi_gtk['ft_pct'],'FT%'),(suma_gtk.get('poss',0),'Posiadania'),
        (suma_gtk.get('br',0),'Straty'),(kpi_gtk['ftr'],'FT Rate'),(kpi_gtk['topct'],'TO%'),
      ])}
    </div>
    <table>
      <thead><tr>
        <th {THL}>Q</th><th {TH}>PKT</th><th {TH}>2PM/A</th><th {TH}>2P%</th>
        <th {TH}>3PM/A</th><th {TH}>3P%</th><th {TH}>FTM/A</th>
        <th {TH}>BR</th><th {TH}>POSS</th><th {TH}>eFG%</th><th {TH}>ORtg</th>
      </tr></thead>
      <tbody>
        {q_row('gtk',1,'1Q')}{q_row('gtk',2,'2Q')}{q_row('gtk',3,'3Q')}{q_row('gtk',4,'4Q')}
        <tr style="background:#e8f0fb;font-weight:700">
          <td style="padding:4px 8px">SUMA</td>
          <td style="text-align:center">{suma_gtk.get('pts',0)}</td>
          <td style="text-align:center">{suma_gtk.get('p2m',0)}/{suma_gtk.get('p2a',0)}</td>
          <td style="text-align:center">{kpi_gtk['p2_pct']}</td>
          <td style="text-align:center">{suma_gtk.get('p3m',0)}/{suma_gtk.get('p3a',0)}</td>
          <td style="text-align:center">{kpi_gtk['p3_pct']}</td>
          <td style="text-align:center">{suma_gtk.get('ftm',0)}/{suma_gtk.get('fta',0)}</td>
          <td style="text-align:center">{suma_gtk.get('br',0)}</td>
          <td style="text-align:center">{suma_gtk.get('poss',0)}</td>
          <td style="text-align:center">{kpi_gtk['efg']}</td>
          <td style="text-align:center">{kpi_gtk['ortg']}</td>
        </tr>
      </tbody>
    </table>
  </div>

  <div class="section">
    <h3 style="color:#8b1a1a">{name_opp} — Kluczowe metryki</h3>
    <div class="kpi-grid">
      {''.join(f'<div class="kpi-box"><div class="kpi-val">{v}</div><div class="kpi-lbl">{l}</div></div>' for v,l in [
        (suma_opp.get('pts',0),'Punkty'),(kpi_opp['efg'],'eFG%'),(kpi_opp['ts'],'TS%'),
        (kpi_opp['ortg'],'ORtg'),(kpi_opp['ppp'],'PPP'),(kpi_opp['p2_pct'],'2PT%'),
        (kpi_opp['p3_pct'],'3PT%'),(kpi_opp['ft_pct'],'FT%'),(suma_opp.get('poss',0),'Posiadania'),
        (suma_opp.get('br',0),'Straty'),(kpi_opp['ftr'],'FT Rate'),(kpi_opp['topct'],'TO%'),
      ])}
    </div>
    <table>
      <thead><tr>
        <th {THL}>Q</th><th {TH}>PKT</th><th {TH}>2PM/A</th><th {TH}>2P%</th>
        <th {TH}>3PM/A</th><th {TH}>3P%</th><th {TH}>FTM/A</th>
        <th {TH}>BR</th><th {TH}>POSS</th><th {TH}>eFG%</th><th {TH}>ORtg</th>
      </tr></thead>
      <tbody>
        {q_row('opp',1,'1Q')}{q_row('opp',2,'2Q')}{q_row('opp',3,'3Q')}{q_row('opp',4,'4Q')}
        <tr style="background:#fff0f0;font-weight:700">
          <td style="padding:4px 8px">SUMA</td>
          <td style="text-align:center">{suma_opp.get('pts',0)}</td>
          <td style="text-align:center">{suma_opp.get('p2m',0)}/{suma_opp.get('p2a',0)}</td>
          <td style="text-align:center">{kpi_opp['p2_pct']}</td>
          <td style="text-align:center">{suma_opp.get('p3m',0)}/{suma_opp.get('p3a',0)}</td>
          <td style="text-align:center">{kpi_opp['p3_pct']}</td>
          <td style="text-align:center">{suma_opp.get('ftm',0)}/{suma_opp.get('fta',0)}</td>
          <td style="text-align:center">{suma_opp.get('br',0)}</td>
          <td style="text-align:center">{suma_opp.get('poss',0)}</td>
          <td style="text-align:center">{kpi_opp['efg']}</td>
          <td style="text-align:center">{kpi_opp['ortg']}</td>
        </tr>
      </tbody>
    </table>
  </div>
</div>

<div class="two-col">
  <div class="section">
    <h3 style="color:#1a6b3c">{gtk_name} — Zawodnicy</h3>
    <table>
      <thead><tr>
        <th {THL}>#</th><th {THG}>PTS</th><th {THG}>2PM/A</th><th {THG}>3PM/A</th>
        <th {THG}>FTM/A</th><th {THG}>eFG%</th><th {THG}>TS%</th>
        <th {THG}>AST</th><th {THG}>ORB</th><th {THG}>DRB</th><th {THG}>BR</th>
      </tr></thead>
      <tbody>{player_rows('gtk') or '<tr><td colspan="11" style="text-align:center;color:#aaa">Brak danych</td></tr>'}</tbody>
    </table>
  </div>
  <div class="section">
    <h3 style="color:#8b1a1a">{name_opp} — Zawodnicy</h3>
    <table>
      <thead><tr>
        <th {THL}>#</th><th {THR}>PTS</th><th {THR}>2PM/A</th><th {THR}>3PM/A</th>
        <th {THR}>FTM/A</th><th {THR}>eFG%</th><th {THR}>TS%</th>
        <th {THR}>AST</th><th {THR}>ORB</th><th {THR}>DRB</th><th {THR}>BR</th>
      </tr></thead>
      <tbody>{player_rows('opp') or '<tr><td colspan="11" style="text-align:center;color:#aaa">Brak danych</td></tr>'}</tbody>
    </table>
  </div>
</div>

<div class="section">
  <h3>Timing rzutów</h3>
  <table style="width:40%">
    <thead><tr>
      <th {THL}>Czas</th>
      <th {THG}>{gtk_name} Celne/Att</th><th {THG}>Eff%</th>
      <th {THR}>{name_opp} Celne/Att</th><th {THR}>Eff%</th>
    </tr></thead>
    <tbody>{timing_rows()}</tbody>
  </table>
</div>

<div style="margin-top:14px;text-align:center;font-size:8px;color:#aaa;border-top:1px solid #eee;padding-top:6px">
  Basket Kołcz Analytics · {gtk_name} vs {name_opp} · {dt}
</div>

</body></html>"""

    buf = io.BytesIO(html.encode('utf-8'))
    buf.seek(0)
    filename = f"raport_{gtk_name}_vs_{name_opp}_{dt}.html".replace(" ","_").replace("/","")
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='text/html')


# ══════════════════════════════════════════════════════════════════════════════
# SZABLONY (identyczne jak w app_v2)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/template/zapis")
def template_zapis():
    wb = openpyxl.Workbook()

    # ── Style ─────────────────────────────────────────────────────────────────
    HDR     = PatternFill("solid", fgColor="1A2B4A")
    HDR_F   = Font(color="FFFFFF", bold=True, size=11)
    GREEN_H = PatternFill("solid", fgColor="99FF66")   # nagłówki A,B,C
    YELLOW_H= PatternFill("solid", fgColor="FFFF99")   # D, J, K
    BLUE_H  = PatternFill("solid", fgColor="BBDEFB")   # E-I zawodnicy
    TEAL_H  = PatternFill("solid", fgColor="B2DFDB")   # L,M,N
    META_KEY= PatternFill("solid", fgColor="E8F5E9")
    META_VAL= PatternFill("solid", fgColor="FFFDE7")
    META_INFO=PatternFill("solid", fgColor="E3F2FD")
    YEL_ROW = PatternFill("solid", fgColor="FFFDE7")   # wiersze danych A-C
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    BD   = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    gtk_name = get_setting("gtk_name") or "TWOJA_DRUŻYNA"

    # ── ARKUSZ META ───────────────────────────────────────────────────────────
    ws_meta = wb.active; ws_meta.title = "META"
    ws_meta.column_dimensions["A"].width = 24
    ws_meta.column_dimensions["B"].width = 28

    ws_meta.merge_cells("A1:B1")
    t = ws_meta["A1"]
    t.value = "DANE MECZU — wypełnij przed kodowaniem"
    t.fill = HDR; t.font = HDR_F; t.alignment = CTR
    ws_meta.row_dimensions[1].height = 22

    meta_fields = [
        ("Drużyna A (twoja)",     gtk_name),
        ("Drużyna B (rywal)",     ""),
        ("Data meczu",            ""),
        ("Rozgrywki",             ""),
        ("Runda / Kolejka",       ""),
        ("Miejsce (dom/wyjazd)", "dom"),
        ("Wynik A",               ""),
        ("Wynik B",               ""),
        ("Uwagi",                 ""),
    ]
    for i, (key, val) in enumerate(meta_fields):
        r = i + 2
        ka = ws_meta.cell(r, 1, key);  ka.fill = META_KEY; ka.font = Font(bold=True, size=10); ka.border = BD; ka.alignment = LEFT
        vb = ws_meta.cell(r, 2, val);  vb.fill = META_VAL; vb.font = Font(size=10);            vb.border = BD; vb.alignment = LEFT
        ws_meta.row_dimensions[r].height = 18

    ws_meta.cell(12, 1, "Nazwa pliku: drużyna_a_drużyna_b").fill = META_INFO
    ws_meta.cell(12, 1).font = Font(italic=True, size=9, color="555555")

    # ── ARKUSZ KODY ───────────────────────────────────────────────────────────
    KODY_DATA = [
        ("SCORING",   "2",         "Celny za 2 punkty",               "rzut z gry za 2"),
        ("SCORING",   "3",         "Celny za 3 punkty",               "rzut z gry za 3"),
        ("SCORING",   "2+1",       "2 pkt + celny rzut wolny",        "faulowany i trafia RW"),
        ("SCORING",   "2+0",       "2 pkt + niecelny rzut wolny",     "faulowany, chybia RW"),
        ("SCORING",   "3+1",       "3 pkt + celny rzut wolny",        "faulowana trójka, trafia"),
        ("SCORING",   "3+0",       "3 pkt + niecelny rzut wolny",     "faulowana trójka, chybia"),
        ("SCORING",   "2D",        "Dobitka celna",                   "tip-in po chybieniu"),
        ("SCORING",   "2D+1",      "Dobitka celna + faul (bez RW)",   ""),
        ("SCORING",   "2D+0/1W",   "Dobitka celna + faul (0/1 RW)",   ""),
        ("SCORING",   "2D+1/1W",   "Dobitka celna + faul (1/1 RW)",   ""),
        ("CHYBIENIE", "0/2",       "Niecelny za 2 punkty",            "chybiony rzut z gry za 2"),
        ("CHYBIENIE", "0/3",       "Niecelny za 3 punkty",            "chybiony rzut za 3"),
        ("CHYBIENIE", "0/2D",      "Niecelna dobitka",                "chybiony tip-in"),
        ("CHYBIENIE", "0D+0/2W",   "Niecelna dobitka + faul (0/2 RW)",""),
        ("CHYBIENIE", "0D+1/2W",   "Niecelna dobitka + faul (1/2 RW)",""),
        ("RW",        "1/2W",      "1/2 rzutów wolnych",              ""),
        ("RW",        "2/2W",      "2/2 rzutów wolnych",              ""),
        ("RW",        "0/2W",      "0/2 rzutów wolnych",              ""),
        ("RW",        "1/3W",      "1/3 rzutów wolnych",              ""),
        ("RW",        "2/3W",      "2/3 rzutów wolnych",              ""),
        ("RW",        "3/3W",      "3/3 rzutów wolnych",              ""),
        ("RW",        "0/3W",      "0/3 rzutów wolnych",              ""),
        ("RW OSTATNI","1/2WL",     "1/2 RW — ostatni wolny serii",    ""),
        ("RW OSTATNI","2/2WL",     "2/2 RW — ostatni wolny serii",    ""),
        ("RW OSTATNI","0/2WL",     "0/2 RW — ostatni wolny serii",    ""),
        ("RW TECH.",  "1/1WT",     "1/1 RW techniczny",               ""),
        ("RW TECH.",  "0/1WT",     "0/1 RW techniczny",               ""),
        ("INNE",      "BR",        "Strata / Ball Loss",              ""),
        ("INNE",      "P",         "Przerwanie gry",                  "timeout, faul tech."),
        ("INNE",      "F",         "Faul (bez punktów)",              ""),
    ]

    ws_k = wb.create_sheet("KODY")
    ws_k.merge_cells("A1:E1")
    t = ws_k["A1"]; t.value = "KODY AKCJI — ściągawka"
    t.fill = HDR; t.font = HDR_F; t.alignment = CTR
    ws_k.row_dimensions[1].height = 22

    for ci, (col, w, lbl) in enumerate([("A",12,"Kategoria"),("B",14,"Kod"),("C",38,"Opis"),("D",36,"Przykład/uwagi"),("E",5,"")]):
        ws_k.column_dimensions[col].width = w
        c = ws_k.cell(2, ci+1, lbl if lbl else "")
        c.fill = HDR; c.font = Font(color="FFFFFF", bold=True, size=9); c.alignment = CTR; c.border = BD

    prev_cat = ""
    for i, (cat, kod, opis, przyklad) in enumerate(KODY_DATA):
        r = i + 3
        bg = PatternFill("solid", fgColor="F8F8F8") if i%2==0 else PatternFill("solid", fgColor="FFFFFF")
        cat_fill = PatternFill("solid", fgColor="E8F5E9") if cat != prev_cat else bg
        for ci, val in enumerate([cat if cat != prev_cat else "", kod, opis, przyklad]):
            c = ws_k.cell(r, ci+1, val)
            c.fill = cat_fill if ci==0 else (PatternFill("solid",fgColor="FFF9C4") if ci==1 else bg)
            c.font = Font(bold=True, size=9) if ci==1 else Font(size=9)
            c.alignment = CTR if ci==1 else LEFT
            c.border = BD
        prev_cat = cat

    ws_k.freeze_panes = "A3"

    # ── ARKUSZE DANYCH ────────────────────────────────────────────────────────
    COLS = [
        ("A", "Kwarta\n(*= tip-off)",     10,  GREEN_H),
        ("B", "Czas\n(sek.)",              8,  GREEN_H),
        ("C", "Kod akcji",                18,  GREEN_H),
        ("D", "Strefa",                    8,  YELLOW_H),
        ("E", "1",                         7,  BLUE_H),
        ("F", "2",                         7,  BLUE_H),
        ("G", "3",                         7,  BLUE_H),
        ("H", "4",                         7,  BLUE_H),
        ("I", "5",                         7,  BLUE_H),
        ("J", "Timeout\n(T=tak)",          9,  YELLOW_H),
        ("K", "Kończy\nakcję",            10,  YELLOW_H),
        ("L", "Asysta ★\n(nr)",            8,  TEAL_H),
        ("M", "Zbiorka OFF ★\n(nr)",      10,  TEAL_H),
        ("N", "Zbiorka DEF ★\n(nr)",      10,  TEAL_H),
    ]

    for team_name in [gtk_name, "drużyna_B"]:
        ws = wb.create_sheet(team_name)

        # Nagłówek tytułowy
        ws.merge_cells("A1:N1")
        t = ws["A1"]
        t.value = f"ZAPIS MECZU — {team_name}"
        t.fill = HDR; t.font = Font(color="FFFFFF", bold=True, size=11); t.alignment = CTR
        ws.row_dimensions[1].height = 20

        # Nagłówki kolumn
        for ci, (col, lbl, w, fill) in enumerate(COLS):
            ws.column_dimensions[col].width = w
            c = ws.cell(2, ci+1, lbl)
            c.fill = fill; c.font = Font(bold=True, size=9); c.alignment = CTR; c.border = BD
        ws.row_dimensions[2].height = 32

        # 200 pustych wierszy danych
        for r in range(3, 203):
            for ci, (col, _, _, fill) in enumerate(COLS):
                c = ws.cell(r, ci+1)
                c.border = BD
                c.alignment = CTR
                # Kolumny A,B,C żółtawe, reszta białe
                if col in ("A","B","C"):
                    c.fill = PatternFill("solid", fgColor="FFFDE7")
                elif col == "D":
                    c.fill = PatternFill("solid", fgColor="F9FBE7")

        ws.freeze_panes = "A3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="ZAPIS_MECZU_szablon.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route("/template/szablon")
def template_szablon():
    wb = openpyxl.Workbook()
    HDR=PatternFill("solid",fgColor="1A2B4A"); HDR_F=Font(color="FFFFFF",bold=True,size=10)
    YEL=PatternFill("solid",fgColor="FFF8E1"); CTR=Alignment(horizontal="center",vertical="center")
    BORDER=Border(bottom=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                  left=Side(style="thin",color="CCCCCC"),top=Side(style="thin",color="CCCCCC"))

    def hdr(ws,row,col,val,w=10):
        c=ws.cell(row,col,val); c.fill=HDR; c.font=HDR_F; c.alignment=CTR; c.border=BORDER
        ws.column_dimensions[get_column_letter(col)].width=w

    sheets=[
        ("TEAM GENERAL",["Q","PKT","POSS","2PM","2PA","2P%","3PM","3PA","3P%","FTM","FTA","FT%","eFG%","TS%","ORtg","DRtg","NetRtg","PPP","TO%","FT Rate"]),
        ("PLAYERS",["#","MIN","PTS","2PM","2PA","2P%","3PM","3PA","3P%","FTM","FTA","FT%","eFG%","TS%","AST","OREB","DREB","BR","FD","FIN"]),
        ("LINEUPS",["Skład","POSS","PKT","PPP","eFG%","ORtg","DRtg","NetRtg","BR","FD"]),
        ("SHOT TIMING",["Czas","2PT Made","2PT Att","2PT%","3PT Made","3PT Att","3PT%","Eff% łącznie"]),
    ]
    ws_first=None
    for sheet_name, cols in sheets:
        if ws_first is None:
            ws=wb.active; ws.title=sheet_name; ws_first=ws
        else:
            ws=wb.create_sheet(sheet_name)
        ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
        t=ws["A1"]; t.value=sheet_name; t.fill=HDR; t.font=Font(color="FFFFFF",bold=True,size=12); t.alignment=CTR
        for i,c in enumerate(cols): hdr(ws,2,i+1,c)
        for r in range(3,18):
            for col in range(1,len(cols)+1):
                c=ws.cell(r,col); c.fill=YEL; c.alignment=CTR; c.border=BORDER
        ws.freeze_panes="B3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="SZABLON_MECZ.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════════════════
# ROSTER — zarządzanie zawodnikami
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/roster")
def roster():
    try: init_db()
    except: pass
    sort = request.args.get("sort","nazwisko")
    order = request.args.get("order","asc")
    if sort not in ("nazwisko","imie","aktywny"): sort = "nazwisko"
    sql_order = f"r.{sort} {'DESC' if order=='desc' else 'ASC'}, r.imie ASC"

    db = get_db(); cur = db.cursor()
    cur.execute(f"""
        SELECT r.id, r.imie, r.nazwisko, r.pseudonim, r.aktywny,
               COALESCE(string_agg(DISTINCT pa.nr::text, ', '), '—') as numery
        FROM roster r
        LEFT JOIN player_aliases pa ON pa.roster_id = r.id
        GROUP BY r.id, r.imie, r.nazwisko, r.pseudonim, r.aktywny
        ORDER BY {sql_order}
    """)
    players = cur.fetchall()
    cur.close()

    def sort_arrow(col):
        if sort == col:
            return " ↓" if order=="asc" else " ↑"
        return " ↕"

    def sort_url(col):
        new_order = "desc" if (sort==col and order=="asc") else "asc"
        return f"/roster?sort={col}&order={new_order}"

    rows = ""
    for i, p in enumerate(players):
        bg = "background:#f8f9ff" if i%2==0 else ""
        is_active = p['aktywny']
        # Toggle button
        toggle_label = "aktywny" if is_active else "nieaktywny"
        toggle_style = "background:#c8e6c9;color:#1a6b3c" if is_active else "background:#e0e0e0;color:#555"
        toggle_title = "Kliknij aby dezaktywować" if is_active else "Kliknij aby aktywować"
        rows += f"""<tr style="{bg}" id="row_{p['id']}">
            <td class="fw-bold">{p['nazwisko']} {p['imie']}</td>
            <td style="color:#888;font-size:.82rem">{p['pseudonim'] or '—'}</td>
            <td style="font-size:.8rem">{p['numery']}</td>
            <td>
              <button type="button"
                onclick="toggleStatus({p['id']}, this)"
                data-active="{1 if is_active else 0}"
                title="{toggle_title}"
                style="border:none;border-radius:20px;padding:3px 10px;font-size:.7rem;font-weight:700;cursor:pointer;transition:.2s;{toggle_style}">
                {toggle_label}
              </button>
            </td>
            <td class="text-center">
              <a href="/roster/{p['id']}/edit" class="btn btn-outline-primary btn-sm" style="font-size:.72rem">Edytuj</a>
              <a href="/roster/{p['id']}/delete" class="btn btn-outline-danger btn-sm ms-1" style="font-size:.72rem"
                 onclick="return confirm('Usunąć {p['imie']} {p['nazwisko']}?')">✕</a>
            </td>
        </tr>"""

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3 flex-wrap gap-2">
  <div class="page-title mb-0">👥 Skład drużyny <span style="font-size:.8rem;color:#aaa;font-weight:400">({len(players)} zawodników)</span></div>
  <div class="d-flex gap-2 flex-wrap">
    <a href="/roster/szablon" class="btn btn-outline-secondary btn-sm">📥 Szablon Excel</a>
    <button class="btn btn-outline-success btn-sm" onclick="document.getElementById('importFile').click()">📤 Importuj</button>
    <form method="POST" action="/roster/import" enctype="multipart/form-data" style="display:none">
      <input type="file" id="importFile" name="file" accept=".xlsx" onchange="this.form.submit()">
    </form>
    <a href="/roster/nowy" class="btn btn-primary btn-sm">+ Dodaj</a>
  </div>
</div>
<div class="card">
  <div class="card-body p-2">
    <div class="table-responsive">
      <table class="table table-hover mb-0">
        <thead><tr>
          <th>
            <a href="{sort_url('nazwisko')}" class="text-white text-decoration-none">
              Nazwisko i imię{sort_arrow('nazwisko')}
            </a>
          </th>
          <th>Pseudonim</th>
          <th>Numery koszulek</th>
          <th>
            <a href="{sort_url('aktywny')}" class="text-white text-decoration-none">
              Status{sort_arrow('aktywny')}
            </a>
          </th>
          <th class="text-center">Akcje</th>
        </tr></thead>
        <tbody id="rosterBody">
          {rows if rows else '<tr><td colspan="5" class="text-center text-muted py-4">Brak zawodników</td></tr>'}
        </tbody>
      </table>
    </div>
  </div>
</div>"""

    scripts = """<script>
function toggleStatus(id, btn) {
    const isActive = btn.dataset.active === '1';
    fetch('/roster/' + id + '/toggle', {method:'POST'})
    .then(r => r.json())
    .then(data => {
        if(data.ok) {
            btn.dataset.active = data.aktywny ? '1' : '0';
            btn.textContent = data.aktywny ? 'aktywny' : 'nieaktywny';
            btn.style.background = data.aktywny ? '#c8e6c9' : '#e0e0e0';
            btn.style.color = data.aktywny ? '#1a6b3c' : '#555';
            btn.title = data.aktywny ? 'Kliknij aby dezaktywować' : 'Kliknij aby aktywować';
        }
    });
}
</script>"""

    return render_template_string(base(content, scripts, active="roster"))


@app.route("/roster/szablon")
def roster_szablon():
    """Pobierz szablon Excel do importu zawodników"""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "ZAWODNICY"

    HDR  = PatternFill("solid", fgColor="1A2B4A")
    HDR_F= Font(color="FFFFFF", bold=True, size=10)
    YEL  = PatternFill("solid", fgColor="FFF9C4")
    CTR  = Alignment(horizontal="center", vertical="center")
    BORDER = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        left=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
    )

    # Nagłówek
    ws.merge_cells("A1:F1")
    t = ws["A1"]; t.value = "SZABLON IMPORTU ZAWODNIKÓW — Basket Kołcz Analytics"
    t.fill = HDR; t.font = Font(color="FFFFFF", bold=True, size=12)
    t.alignment = CTR
    ws.row_dimensions[1].height = 24

    # Kolumny
    cols = [
        ("A", "Imię *",          14),
        ("B", "Nazwisko",        16),
        ("C", "Pseudonim",       14),
        ("D", "Numery koszulek", 22),
        ("E", "Sezon",           12),
        ("F", "Status",          12),
    ]
    for col, label, width in cols:
        ws.column_dimensions[col].width = width
        c = ws[f"{col}2"]
        c.value = label; c.fill = HDR; c.font = HDR_F
        c.alignment = CTR; c.border = BORDER
    ws.row_dimensions[2].height = 22

    # Przykładowe dane
    examples = [
        ("Jan",    "Kowalski",  "KOW",  "5, 12",  "2025/26", "aktywny"),
        ("Piotr",  "Nowak",     "NOW",  "7",       "2025/26", "aktywny"),
        ("Marek",  "Wiśniewski","WIS",  "4, 4",    "2025/26", "aktywny"),
        ("Tomasz", "Kowalczyk", "",     "11",      "2025/26", "aktywny"),
        ("",       "",          "",     "",        "",        ""),
    ]
    for i, (imie,nazwisko,pseudo,numery,sezon,status) in enumerate(examples):
        r = 3+i
        for j, v in enumerate([imie,nazwisko,pseudo,numery,sezon,status]):
            c = ws.cell(r, j+1, v)
            c.fill = YEL if v else PatternFill("solid", fgColor="FAFAFA")
            c.alignment = CTR; c.border = BORDER

    # Walidacja kolumny F (Status)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"aktywny,nieaktywny"', allow_blank=True)
    ws.add_data_validation(dv); dv.add("F3:F200")

    # Legenda
    ws["A10"] = "INSTRUKCJA:"
    ws["A10"].font = Font(bold=True, color="1A2B4A")
    notes = [
        ("A11", "* Imię jest wymagane"),
        ("A12", "* Numery: wpisz oddzielone przecinkiem, np. 5, 12"),
        ("A13", "* Sezon: dotyczy wszystkich numerów w tym wierszu, np. 2025/26"),
        ("A14", "* Jeśli zawodnik nosił różne numery w różnych sezonach — dodaj osobny wiersz"),
        ("A15", "* Status: aktywny / nieaktywny"),
    ]
    for cell, text in notes:
        ws[cell] = text
        ws[cell].font = Font(italic=True, color="666666", size=9)

    ws.freeze_panes = "A3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="SZABLON_ZAWODNICY.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route("/roster/import", methods=["POST"])
def roster_import():
    """Importuj zawodników z pliku Excel"""
    if "file" not in request.files:
        flash("Nie wybrano pliku","error"); return redirect(url_for("roster"))
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        flash("Plik musi być w formacie .xlsx","error"); return redirect(url_for("roster"))

    try:
        init_db()
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        # Znajdź arkusz z danymi
        ws = None
        for sn in wb.sheetnames:
            if "zawodnicy" in sn.lower() or "roster" in sn.lower() or sn == wb.sheetnames[0]:
                ws = wb[sn]; break
        if not ws:
            flash("Nie znaleziono arkusza z danymi","error"); return redirect(url_for("roster"))

        db = get_db(); cur = db.cursor()
        dodani = 0; zaktualizowani = 0; bledy = []

        for i, row in enumerate(ws.iter_rows(min_row=3, max_row=500, values_only=True), 3):
            # Pomiń puste wiersze i przykłady
            if not row[0] and not row[1]: continue
            imie = str(row[0] or "").strip()
            if not imie: continue
            if imie.lower() in ("jan","piotr","marek","tomasz","imię"): continue  # pomiń przykłady

            nazwisko  = str(row[1] or "").strip()
            pseudonim = str(row[2] or "").strip()
            numery_raw= str(row[3] or "").strip()
            sezon_raw = str(row[4] or "").strip()
            status    = str(row[5] or "aktywny").strip().lower()
            aktywny   = status != "nieaktywny"

            try:
                # Sprawdź czy istnieje po nazwisku (główny klucz)
                cur.execute("SELECT id FROM roster WHERE LOWER(nazwisko)=LOWER(%s)", (nazwisko,))
                existing = cur.fetchone()
                if not existing and imie:
                    # Próbuj też po imieniu + nazwisku
                    cur.execute("SELECT id FROM roster WHERE LOWER(imie)=LOWER(%s) AND LOWER(nazwisko)=LOWER(%s)", (imie, nazwisko))
                    existing = cur.fetchone()

                if existing:
                    cur.execute("UPDATE roster SET pseudonim=%s, aktywny=%s WHERE id=%s",
                                (pseudonim, aktywny, existing["id"]))
                    roster_id = existing["id"]
                    zaktualizowani += 1
                else:
                    cur.execute("INSERT INTO roster (imie,nazwisko,pseudonim,aktywny) VALUES (%s,%s,%s,%s) RETURNING id",
                                (imie, nazwisko, pseudonim, aktywny))
                    roster_id = cur.fetchone()["id"]
                    dodani += 1

                # Dodaj numery koszulek
                if numery_raw:
                    import re as _re
                    for nr_part in numery_raw.split(","):
                        nr_part = nr_part.strip()
                        if not nr_part: continue
                        try:
                            nr = int(_re.sub(r'[^\d]','', nr_part))
                            sezon = sezon_raw
                            cur.execute("""INSERT INTO player_aliases (roster_id,nr,sezon)
                                          VALUES (%s,%s,%s) ON CONFLICT DO NOTHING""",
                                        (roster_id, nr, sezon))
                        except: pass

            except Exception as e:
                bledy.append(f"Wiersz {i}: {str(e)[:60]}")
                try: get_db().rollback()
                except: pass

        db.commit(); cur.close()

        msg = f"✓ Import zakończony: {dodani} nowych, {zaktualizowani} zaktualizowanych"
        if bledy: msg += f" | Błędy: {', '.join(bledy[:3])}"
        flash(msg, "success" if not bledy else "error")

    except Exception as e:
        try: get_db().rollback()
        except: pass
        flash(f"Błąd importu: {str(e)}", "error")

    return redirect(url_for("roster"))


@app.route("/roster/nowy", methods=["GET","POST"])
@app.route("/roster/<int:player_id>/edit", methods=["GET","POST"])
def roster_edit(player_id=None):
    try: init_db()
    except: pass
    db = get_db(); cur = db.cursor()

    if request.method == "POST":
        imie     = request.form.get("imie","").strip()
        nazwisko = request.form.get("nazwisko","").strip()
        pseudonim= request.form.get("pseudonim","").strip()
        aktywny  = request.form.get("aktywny","1") == "1"
        # Numery — lista par (nr, sezon)
        numery_raw = request.form.get("numery","").strip()

        if not imie:
            flash("Imię jest wymagane","error")
            return redirect(request.url)

        if player_id:
            cur.execute("UPDATE roster SET imie=%s,nazwisko=%s,pseudonim=%s,aktywny=%s WHERE id=%s",
                        (imie,nazwisko,pseudonim,aktywny,player_id))
            cur.execute("DELETE FROM player_aliases WHERE roster_id=%s", (player_id,))
        else:
            cur.execute("INSERT INTO roster (imie,nazwisko,pseudonim,aktywny) VALUES (%s,%s,%s,%s) RETURNING id",
                        (imie,nazwisko,pseudonim,aktywny))
            player_id = cur.fetchone()["id"]

        # Parsuj numery: "5, 12 (2024/25), 7 (2025/26)"
        for part in numery_raw.split(","):
            part = part.strip()
            if not part: continue
            import re as _re
            m = _re.match(r'(\d+)\s*(?:\(([^)]+)\))?', part)
            if m:
                nr = int(m.group(1))
                sezon = (m.group(2) or "").strip()
                try:
                    cur.execute("INSERT INTO player_aliases (roster_id,nr,sezon) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                                (player_id, nr, sezon))
                except: pass

        db.commit(); cur.close()
        flash(f"✓ Zawodnik {imie} {nazwisko} zapisany","success")
        return redirect(url_for("roster"))

    # GET
    player = None; aliases = []
    if player_id:
        cur.execute("SELECT * FROM roster WHERE id=%s", (player_id,))
        player = cur.fetchone()
        cur.execute("SELECT nr,sezon FROM player_aliases WHERE roster_id=%s ORDER BY nr", (player_id,))
        aliases = cur.fetchall()
    cur.close()

    numery_str = ", ".join(
        f"{a['nr']}" + (f" ({a['sezon']})" if a['sezon'] else "")
        for a in aliases
    )

    title = "Edytuj zawodnika" if player_id else "Nowy zawodnik"
    content = f"""
<div class="page-title">{title}</div>
<div class="row justify-content-center">
<div class="col-lg-6">
  <div class="card p-3">
    <form method="POST">
      <div class="row g-3">
        <div class="col-6">
          <label class="form-label fw-bold">Imię *</label>
          <input type="text" name="imie" class="form-control" value="{''+player['imie'] if player else ''}" required>
        </div>
        <div class="col-6">
          <label class="form-label fw-bold">Nazwisko</label>
          <input type="text" name="nazwisko" class="form-control" value="{''+player['nazwisko'] if player else ''}">
        </div>
        <div class="col-6">
          <label class="form-label fw-bold">Pseudonim / inicjały</label>
          <input type="text" name="pseudonim" class="form-control" value="{''+player['pseudonim'] if player else ''}" placeholder="np. KOW, Kowal">
        </div>
        <div class="col-6">
          <label class="form-label fw-bold">Status</label>
          <select name="aktywny" class="form-select">
            <option value="1" {'selected' if not player or player['aktywny'] else ''}>Aktywny</option>
            <option value="0" {'selected' if player and not player['aktywny'] else ''}>Nieaktywny</option>
          </select>
        </div>
        <div class="col-12">
          <label class="form-label fw-bold">Numery koszulek</label>
          <input type="text" name="numery" class="form-control" value="{numery_str}"
                 placeholder="np. 5, 12 (2024/25), 7 (2025/26)">
          <div class="form-text">Wpisz numery oddzielone przecinkiem. Opcjonalnie dodaj sezon w nawiasie.</div>
        </div>
      </div>
      <div class="d-flex gap-2 mt-3">
        <button type="submit" class="btn btn-primary">Zapisz</button>
        <a href="/roster" class="btn btn-outline-secondary">Anuluj</a>
      </div>
    </form>
  </div>
</div></div>"""
    return render_template_string(base(content, active="players"))


@app.route("/roster/<int:player_id>/toggle", methods=["POST"])
def roster_toggle(player_id):
    from flask import jsonify
    try:
        db = get_db(); cur = db.cursor()
        cur.execute("SELECT aktywny FROM roster WHERE id=%s", (player_id,))
        row = cur.fetchone()
        if not row: return jsonify({"ok": False})
        new_status = not row["aktywny"]
        cur.execute("UPDATE roster SET aktywny=%s WHERE id=%s", (new_status, player_id))
        db.commit(); cur.close()
        return jsonify({"ok": True, "aktywny": new_status})
    except Exception as e:
        try: get_db().rollback()
        except: pass
        return jsonify({"ok": False, "error": str(e)})


@app.route("/roster/<int:player_id>/delete")
def roster_delete(player_id):
    db = get_db(); cur = db.cursor()
    cur.execute("DELETE FROM roster WHERE id=%s", (player_id,))
    db.commit(); cur.close()
    flash("Zawodnik usunięty","success")
    return redirect(url_for("roster"))


# ══════════════════════════════════════════════════════════════════════════════
# EDYTOR MECZU — przypisanie numerów do zawodników z rostera
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/mecz/<int:match_id>/edytuj", methods=["GET","POST"])
def mecz_edytuj(match_id):
    try: init_db()
    except: pass
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM matches WHERE id=%s", (match_id,))
    m = cur.fetchone()
    if not m: return redirect(url_for("historia"))

    gtk_name = (m.get("nazwa_gtk","") or "").strip() or get_setting("gtk_name") or "GTK"

    cur.execute("""SELECT id,nr,pts,p2m,p2a,p3m,p3a,ftm,fta,roster_id
                   FROM player_stats WHERE match_id=%s AND druzyna='gtk'
                   ORDER BY nr""", (match_id,))
    players = list(cur.fetchall())

    cur.execute("""SELECT r.id, r.imie, r.nazwisko, r.pseudonim
                   FROM roster r WHERE r.aktywny=TRUE
                   ORDER BY r.nazwisko ASC, r.imie ASC""")
    roster_list = list(cur.fetchall())

    if request.method == "POST":
        sezon = m.get("sezon","")
        for p in players:
            rid = request.form.get(f"roster_{p['id']}","")
            rid_val = int(rid) if rid and rid.isdigit() else None
            cur.execute("UPDATE player_stats SET roster_id=%s WHERE id=%s", (rid_val, p['id']))
            if rid_val and p['nr'] is not None:
                try:
                    cur.execute("""INSERT INTO player_aliases (roster_id,nr,sezon)
                                   VALUES (%s,%s,%s) ON CONFLICT DO NOTHING""",
                                (rid_val, p['nr'], sezon))
                except: pass
        db.commit(); cur.close()
        flash("✓ Przypisania zawodników zapisane","success")
        return redirect(url_for("mecz", match_id=match_id))

    cur.close()

    # JSON roster dla JS
    import json as _json
    roster_json = _json.dumps([
        {"id": r["id"],
         "name": f"{r['nazwisko']} {r['imie']}" + (f" ({r['pseudonim']})" if r['pseudonim'] else "")}
        for r in roster_list
    ])

    # Aktualne przypisania {ps_id: roster_id}
    current = {p['id']: (p.get('roster_id') or "") for p in players}
    current_json = _json.dumps({str(k): str(v) if v else "" for k,v in current.items()})

    # Wiersze tabeli — same numery i statystyki, select wypełniany przez JS
    rows = ""
    for i, p in enumerate(players):
        fga = p.get('p2a',0)+p.get('p3a',0)
        efg = f"{(p.get('p2m',0)+1.5*p.get('p3m',0))/fga:.0%}" if fga else "—"
        bg = "background:#f8f9ff" if i%2==0 else ""
        rows += f"""<tr style="{bg}">
            <td class="fw-bold" style="width:55px;font-size:1rem">#{p['nr']}</td>
            <td class="fw-bold" style="color:#1a2b4a;width:50px">{p['pts']}</td>
            <td style="font-size:.78rem;color:#888;width:90px">{p.get('p2m',0)}/{p.get('p2a',0)} | {p.get('p3m',0)}/{p.get('p3a',0)}</td>
            <td style="font-size:.78rem;color:#888;width:55px">{efg}</td>
            <td>
              <select name="roster_{p['id']}" id="sel_{p['id']}"
                      class="form-select form-select-sm roster-sel"
                      data-ps-id="{p['id']}"
                      onchange="updateSelects()">
              </select>
            </td>
        </tr>"""

    dt = m['data_meczu'].strftime('%d.%m.%Y') if m['data_meczu'] else ""

    content = f"""
<div class="d-flex justify-content-between align-items-center mb-3">
  <div>
    <div class="page-title mb-0">✏️ Przypisz zawodników — {gtk_name}</div>
    <div style="font-size:.8rem;color:#888">{m['przeciwnik']} · {dt}</div>
  </div>
  <a href="/mecz/{match_id}" class="btn btn-outline-secondary btn-sm">← Wróć</a>
</div>

<div class="card p-2 mb-3" style="background:#e8f5e9;border:1px solid #a5d6a7">
  <div style="font-size:.82rem;color:#1a6b3c">
    <b>Wybierz zawodnika</b> z listy dla każdego numeru koszulki.
    Wybrany zawodnik znika z pozostałych list.
    Numer zostanie automatycznie przypisany do zawodnika w rosterze.
    <a href="/roster/nowy" style="color:#1a6b3c">+ Dodaj zawodnika do rostera</a>
  </div>
</div>

<div class="card">
  <div class="card-body p-3">
    <form method="POST" id="assignForm">
      <div class="table-responsive">
        <table class="table table-hover mb-0" style="table-layout:fixed">
          <thead><tr>
            <th style="width:55px">#</th>
            <th style="width:50px">PTS</th>
            <th style="width:90px">2PT|3PT</th>
            <th style="width:55px">eFG%</th>
            <th>Zawodnik</th>
          </tr></thead>
          <tbody>{rows}</tbody>
        </table>
      </div>
      <div class="d-flex gap-2 mt-3">
        <button type="submit" class="btn btn-primary fw-bold">✓ Zapisz</button>
        <a href="/mecz/{match_id}" class="btn btn-outline-secondary">Anuluj</a>
        <button type="button" class="btn btn-outline-warning btn-sm ms-auto"
                onclick="clearAll()">Wyczyść wszystkie</button>
      </div>
    </form>
  </div>
</div>"""

    scripts = f"""<script>
const ROSTER = {roster_json};
const CURRENT = {current_json};

function getSelected() {{
  const sel = {{}};
  document.querySelectorAll('.roster-sel').forEach(s => {{
    if(s.value) sel[s.value] = s.dataset.psId;
  }});
  return sel;
}}

function updateSelects() {{
  const selected = getSelected();
  document.querySelectorAll('.roster-sel').forEach(sel => {{
    const curVal = sel.value;
    sel.innerHTML = '<option value="">— nie przypisany —</option>';
    ROSTER.forEach(r => {{
      const usedBy = selected[r.id.toString()];
      // Pokaż jeśli: nie wybrany przez nikogo LUB wybrany przez ten sam select
      if(!usedBy || usedBy === sel.dataset.psId) {{
        const opt = document.createElement('option');
        opt.value = r.id;
        opt.textContent = r.name;
        if(r.id.toString() === curVal) opt.selected = true;
        sel.appendChild(opt);
      }}
    }});
  }});
}}

function clearAll() {{
  document.querySelectorAll('.roster-sel').forEach(s => s.value = '');
  updateSelects();
}}

// Inicjalizacja — wypełnij z aktualnych przypisań
document.addEventListener('DOMContentLoaded', () => {{
  // Najpierw wypełnij wszystkie puste
  document.querySelectorAll('.roster-sel').forEach(sel => {{
    sel.innerHTML = '<option value="">— nie przypisany —</option>';
    ROSTER.forEach(r => {{
      const opt = document.createElement('option');
      opt.value = r.id; opt.textContent = r.name;
      sel.appendChild(opt);
    }});
  }});
  // Ustaw zapisane wartości
  const psIds = Object.keys(CURRENT);
  psIds.forEach(psId => {{
    const rid = CURRENT[psId];
    if(rid) {{
      const sel = document.querySelector(`[data-ps-id="${{psId}}"]`);
      if(sel) sel.value = rid;
    }}
  }});
  // Teraz odśwież listy
  updateSelects();
}});
</script>"""

    return render_template_string(base(content, scripts, active="history"))


if __name__ == "__main__":
    app.run(debug=True)
